"""
Read DHL and GLS price lists from N:\Analityka\TKL\Kurierzy
"""
import openpyxl
import os

files = [
    r"N:\Analityka\TKL\Kurierzy\Aktualny cennik wysyłek na rok 2026.xlsx",
    r"N:\Analityka\TKL\Kurierzy\Cenniki_GLS-DHL_PL_2026.xlsx",
    r"N:\Analityka\TKL\Kurierzy\DHL\Tabela kraj - oferta na 2026.xlsx",
]

for fp in files:
    if not os.path.exists(fp):
        print(f"\n!!! FILE NOT FOUND: {fp}")
        continue
    
    print(f"\n{'='*80}")
    print(f"FILE: {os.path.basename(fp)}")
    print(f"{'='*80}")
    
    wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- Sheet: {sheet_name} (rows={ws.max_row}, cols={ws.max_column}) ---")
        row_count = 0
        for row in ws.iter_rows(max_row=50, values_only=False):
            row_count += 1
            vals = []
            for cell in row:
                v = cell.value
                if v is not None:
                    vals.append(str(v)[:60])
                else:
                    vals.append("")
            line = " | ".join(vals)
            if any(v for v in vals):
                print(f"  R{row_count:3d}: {line}")
    wb.close()

print("\nDone.")

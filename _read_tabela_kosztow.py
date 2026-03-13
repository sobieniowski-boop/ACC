# Read "00. Tabela Kosztow Logistycznych" - product to box mapping
import sys, os, glob
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "apps", "api"))
import openpyxl

tkl_dir = "N:\\Analityka\\TKL"
# Find the file
for f in os.listdir(tkl_dir):
    if 'tabela' in f.lower() and 'koszt' in f.lower():
        path = os.path.join(tkl_dir, f)
        break

print(f"Reading: {path}")
print(f"Size: {os.path.getsize(path):,} bytes\n")

wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
print(f"Sheets: {wb.sheetnames}\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"=== Sheet: {sheet_name} ===")
    rows = list(ws.iter_rows(values_only=True, max_row=30))
    for i, row in enumerate(rows):
        if all(v is None for v in row):
            continue
        vals = [str(v)[:40] if v is not None else '' for v in row[:20]]
        print(f"  R{i}: {' | '.join(vals)}")
    # Count total rows
    total = sum(1 for _ in ws.iter_rows(values_only=True))
    print(f"  ... ({total} total rows)\n")

wb.close()

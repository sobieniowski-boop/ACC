# ══════════════════════════════════════════════════════════════════════
# MODEL V3: Weight-based logistics cost estimation
# 
# Pipeline:
#   1. Load product → weight mapping from TKL "00. Tabela Kosztów Logistycznych"
#   2. Map Amazon merchant_sku → internal_sku via acc_amazon_listing_registry
#   3. For each MFN order needing estimation:
#      a) Sum product billing weights for all lines
#      b) Look up carrier price by (country, billing_weight_bracket)
#      c) Use GLS pricing (primary carrier for most routes)
#   4. DELETE old v1/v2 estimates
#   5. INSERT new v3 rows
# ══════════════════════════════════════════════════════════════════════
import sys, os, time, math
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "apps", "api"))
from app.core.db_connection import connect_acc
from collections import defaultdict
from datetime import datetime
import statistics
import openpyxl

DRY_RUN = False   # True = just analyze, no DB writes
BATCH_SIZE = 500

print("=" * 80)
print(f"MODEL V3 — {'DRY RUN' if DRY_RUN else 'LIVE APPLY'}")
print("=" * 80)

# ── 1. LOAD TKL PRODUCT WEIGHTS ──
print("\n[1] Loading TKL product weights...")
t0 = time.time()

tkl_dir = "N:\\Analityka\\TKL"
for f in os.listdir(tkl_dir):
    if 'tabela' in f.lower() and 'koszt' in f.lower():
        tkl_path = os.path.join(tkl_dir, f)
        break

wb = openpyxl.load_workbook(tkl_path, read_only=True, data_only=True)
ws = wb['DHL']

# Columns (0-indexed):
# 0=Nr artykulu, 1=Nazwa, 3=Waga bilingowa GLS po spakowaniu
# 4=Najdluzszy, 5=Sredni, 6=Najmniejszy, 7=Waga(kg)
# 10=Sugestia kartonu, 12=Waga kartonu
# 14=Waga billingowa DHL po spakowaniu
# 18=Czy moze DHL, 19=DE GLS price

product_weight = {}  # internal_sku -> {gls_bw, dhl_bw, carrier_hint}
rows = list(ws.iter_rows(values_only=True, min_row=3))
for row in rows:
    nr_art = row[0]
    if not nr_art:
        continue
    nr_art = str(nr_art).strip()
    
    gls_bw = row[3]   # Waga bilingowa GLS po spakowaniu
    dhl_bw = row[14]   # Waga billingowa DHL po spakowaniu
    waga_produktu = row[7]   # Waga samego produktu (kg)
    waga_kartonu = row[12]   # Waga kartonu
    carrier_hint = str(row[18] or '').strip().upper() if len(row) > 18 else ''  # Czy moze DHL
    
    # Prefer GLS billing weight (most orders go GLS)
    bw = None
    if gls_bw and str(gls_bw) not in ('#N/A', 'None', ''):
        try:
            bw = float(gls_bw)
        except (ValueError, TypeError):
            pass
    
    dhl_bw_val = None
    if dhl_bw and str(dhl_bw) not in ('#N/A', 'None', ''):
        try:
            dhl_bw_val = float(dhl_bw)
        except (ValueError, TypeError):
            pass
    
    prod_wt = None
    if waga_produktu and str(waga_produktu) not in ('#N/A', 'None', ''):
        try:
            prod_wt = float(waga_produktu)
        except (ValueError, TypeError):
            pass
    
    box_wt = None
    if waga_kartonu and str(waga_kartonu) not in ('#N/A', 'None', ''):
        try:
            box_wt = float(waga_kartonu)
        except (ValueError, TypeError):
            pass
    
    if bw is not None or dhl_bw_val is not None or prod_wt is not None:
        product_weight[nr_art] = {
            'gls_bw': bw,
            'dhl_bw': dhl_bw_val,
            'prod_wt': prod_wt,
            'box_wt': box_wt,
            'carrier': carrier_hint,
        }

# Also load Liczność paczek (multi-parcel products)
ws2 = wb['Liczność paczek']
multi_parcel_skus = {}
for row in ws2.iter_rows(values_only=True, min_row=2):
    if row[0] and row[1]:
        multi_parcel_skus[str(row[0]).strip()] = int(row[1])

wb.close()
print(f"  TKL products with weight: {len(product_weight):,}")
print(f"  Multi-parcel products: {len(multi_parcel_skus)}")
print(f"  Time: {time.time()-t0:.1f}s")

# ── 2. LOAD AMAZON SKU → INTERNAL_SKU MAPPING ──
print("\n[2] Loading SKU mapping from DB...")
conn = connect_acc(timeout=120)
cur = conn.cursor()

cur.execute("""
    SELECT merchant_sku, merchant_sku_alt, internal_sku
    FROM dbo.acc_amazon_listing_registry WITH (NOLOCK)
    WHERE ISNULL(internal_sku, '') != ''
      AND ISNULL(merchant_sku, '') != ''
""")

sku_to_internal = {}  # merchant_sku -> internal_sku
for merchant, alt, internal in cur.fetchall():
    internal = str(internal).strip()
    if merchant:
        sku_to_internal[merchant.strip()] = internal
    if alt:
        sku_to_internal[alt.strip()] = internal

print(f"  Merchant→Internal mappings: {len(sku_to_internal):,}")

# ── 3. GLS PRICE LISTS (PLN, all-in with fuel surcharge) ──
print("\n[3] Setting up price tables...")

# GLS all-in prices from 2026 cennik (per weight bracket, PLN)
# Brackets: 1, 2, 3, 5, 10, 15, 25, 30 kg
GLS_PRICES = {
    'DE': {1: 15.96, 2: 17.99, 3: 20.14, 5: 20.14, 10: 21.84, 15: 26.06, 25: 36.76, 30: 48.06},
    'AT': {1: 18.22, 2: 18.22, 3: 21.54, 5: 21.54, 10: 21.54, 15: 29.21, 25: 35.64, 30: 42.41},
    'FR': {1: 24.36, 2: 24.36, 3: 31.97, 5: 31.97, 10: 39.75, 15: 48.23, 25: 56.43, 30: 62.99},
    'IT': {1: 23.67, 2: 23.67, 3: 27.47, 5: 27.47, 10: 37.45, 15: 47.30, 25: 53.93, 30: 56.08},
    'NL': {1: 18.69, 2: 21.19, 3: 23.09, 5: 23.09, 10: 24.80, 15: 28.85, 25: 35.35, 30: 39.41},
    'BE': {1: 21.89, 2: 25.59, 3: 28.70, 5: 28.70, 10: 30.58, 15: 36.37, 25: 42.87, 30: 47.59},
    'ES': {1: 30.36, 2: 30.36, 3: 33.81, 5: 33.81, 10: 42.85, 15: 53.59, 25: 61.19, 30: 67.29},
    'SE': {1: 23.32, 2: 25.32, 3: 31.41, 5: 31.41, 10: 37.73, 15: 40.20, 25: 44.59, 30: 50.16},
    'LU': {1: 24.32, 2: 28.52, 3: 31.05, 5: 31.05, 10: 33.87, 15: 44.49, 25: 52.94, 30: 56.58},
    'DK': {1: 24.38, 2: 24.88, 3: 30.59, 5: 30.59, 10: 38.61, 15: 45.93, 25: 51.27, 30: 56.38},
    'IE': {1: 47.54, 2: 47.54, 3: 47.54, 5: 50.35, 10: 69.26, 15: 80.20, 25: 94.59, 30: 108.97},
    'PL': {1: 10.16, 2: 10.16, 3: 10.16, 5: 10.16, 10: 11.12, 15: 11.12, 25: 11.92, 30: 12.89},
    'PT': {1: 58.38, 2: 58.38, 3: 63.85, 5: 63.85, 10: 72.66, 15: 79.75, 25: 94.66, 30: 107.49},
    'CZ': {1: 15.82, 2: 17.23, 3: 19.68, 5: 19.68, 10: 21.77, 15: 25.92, 25: 31.13, 30: 36.76},
    'HU': {1: 22.97, 2: 24.38, 3: 26.57, 5: 26.57, 10: 32.61, 15: 40.46, 25: 50.19, 30: 56.08},
    'HR': {1: 30.70, 2: 30.70, 3: 33.81, 5: 33.81, 10: 43.78, 15: 52.04, 25: 63.30, 30: 73.10},
    'SI': {1: 27.04, 2: 27.04, 3: 30.58, 5: 30.58, 10: 39.25, 15: 47.70, 25: 54.19, 30: 60.08},
    'FI': {1: 36.37, 2: 38.07, 3: 44.08, 5: 44.08, 10: 56.39, 15: 66.20, 25: 76.78, 30: 87.39},
    'SK': {1: 17.99, 2: 20.14, 3: 22.69, 5: 22.69, 10: 25.36, 15: 30.97, 25: 38.14, 30: 44.88},
    'LT': {1: 22.61, 2: 24.88, 3: 27.22, 5: 27.22, 10: 33.37, 15: 42.60, 25: 51.27, 30: 60.20},
    'LV': {1: 25.36, 2: 25.36, 3: 27.22, 5: 27.22, 10: 33.37, 15: 42.60, 25: 51.27, 30: 60.20},
    'EE': {1: 25.36, 2: 25.36, 3: 27.22, 5: 27.22, 10: 33.37, 15: 42.60, 25: 51.27, 30: 60.20},
    'RO': {1: 33.98, 2: 33.98, 3: 39.25, 5: 39.25, 10: 48.17, 15: 56.77, 25: 67.72, 30: 77.37},
    'BG': {1: 38.14, 2: 38.14, 3: 44.08, 5: 44.08, 10: 54.19, 15: 65.17, 25: 72.34, 30: 81.00},
    'GR': {1: 40.25, 2: 40.25, 3: 46.32, 5: 46.32, 10: 58.58, 15: 71.93, 25: 82.35, 30: 94.59},
}

GLS_BRACKETS = sorted(GLS_PRICES['DE'].keys())

def gls_cost(country, weight_kg):
    """Look up GLS cost for country and weight."""
    prices = GLS_PRICES.get(country)
    if not prices:
        # Fallback: use DE pricing (most expensive EU average)
        prices = GLS_PRICES.get('DE', {})
    for bracket in GLS_BRACKETS:
        if weight_kg <= bracket:
            return prices.get(bracket, prices[max(prices.keys())])
    return prices[max(prices.keys())]  # Over 30kg → max bracket

# ── 4. LOAD ORDERS NEEDING ESTIMATION ──
print("\n[4] Loading orders needing estimation...")

cur.execute("""
    SELECT o.amazon_order_id, o.ship_country, ol.sku, ol.quantity_ordered
    FROM dbo.acc_order o WITH (NOLOCK)
    JOIN dbo.acc_order_line ol WITH (NOLOCK) ON ol.order_id = o.id
    WHERE o.fulfillment_channel = 'MFN'
      AND o.purchase_date >= '2026-02-01' AND o.purchase_date < '2026-03-12'
      AND NOT EXISTS (
          SELECT 1 FROM dbo.acc_order_logistics_fact f WITH (NOLOCK)
          WHERE f.amazon_order_id = o.amazon_order_id
            AND f.calc_version IN ('gls_v1', 'dhl_v1')
      )
""")

order_lines = defaultdict(list)  # amazon_order_id -> [(sku, qty, country)]
order_country = {}
for amazon_id, country, sku, qty in cur.fetchall():
    order_lines[amazon_id].append((sku, int(qty) if qty else 1))
    order_country[amazon_id] = country

print(f"  Orders to estimate: {len(order_lines):,}")
print(f"  Total order lines: {sum(len(v) for v in order_lines.values()):,}")

# ── 5. CALCULATE V3 ESTIMATES ──
print("\n[5] Calculating v3 estimates...")

# Build fallback: SKU×country median from actual billing (for products not in TKL)
cur.execute("""
    SELECT ol.sku, o.ship_country,
           sc.net_amount + ISNULL(sc.fuel_amount, 0) + ISNULL(sc.toll_amount, 0) AS cost
    FROM dbo.acc_order o WITH (NOLOCK)
    JOIN dbo.acc_order_line ol WITH (NOLOCK) ON ol.order_id = o.id
    JOIN dbo.acc_shipment_order_link sol WITH (NOLOCK) ON sol.amazon_order_id = o.amazon_order_id
    JOIN dbo.acc_shipment_cost sc WITH (NOLOCK) ON sc.shipment_id = sol.shipment_id
    WHERE o.fulfillment_channel = 'MFN'
      AND sc.net_amount > 0 AND sc.is_estimated = 0
      AND NOT EXISTS (
          SELECT 1 FROM dbo.acc_order_line ol2 WITH (NOLOCK) 
          WHERE ol2.order_id = o.id AND ol2.id != ol.id
      )
""")
sku_country_fallback = defaultdict(list)
for sku, country, cost in cur.fetchall():
    if sku and country and cost:
        sku_country_fallback[(sku, country)].append(float(cost))
sku_country_median = {k: statistics.median(v) for k, v in sku_country_fallback.items()}

country_fallback = defaultdict(list)
for (sku, c), costs in sku_country_fallback.items():
    country_fallback[c].extend(costs)
country_median = {c: statistics.median(v) for c, v in country_fallback.items()}

conn.close()

# Process each order
results = []
tier_counts = {'weight_v3': 0, 'sku_billing': 0, 'country_fb': 0}
unresolved_skus = set()

for amazon_id, lines in order_lines.items():
    country = order_country.get(amazon_id, 'DE')
    
    total_weight = 0.0
    all_resolved = True
    uses_weight = True
    sku_billing_costs = []
    
    for sku, qty in lines:
        # Try to resolve weight via TKL
        internal = sku_to_internal.get(sku)
        pw = product_weight.get(internal) if internal else None
        
        if pw and pw['gls_bw'] is not None and pw['gls_bw'] > 0:
            # GLS billing weight per item (already includes packaging for single items)
            # For multi-qty: product_weight × qty + box_weight (one box for all)
            if qty == 1:
                total_weight += pw['gls_bw']
            else:
                # Multi qty: weight = prod_wt * qty + box_wt
                prod_w = pw['prod_wt'] if pw['prod_wt'] else (pw['gls_bw'] - (pw['box_wt'] or 0))
                box_w = pw['box_wt'] if pw['box_wt'] else 0.5
                item_weight = prod_w * qty + box_w
                # Also check multi-parcel
                if internal in multi_parcel_skus:
                    item_weight *= multi_parcel_skus[internal]
                total_weight += item_weight
        elif pw and pw['prod_wt'] is not None and pw['prod_wt'] > 0:
            # Only have product weight, estimate box = 0.5kg
            total_weight += (pw['prod_wt'] * qty + 0.5)
        else:
            all_resolved = False
            uses_weight = False
            # Fallback: get from billing history
            key = (sku, country)
            if key in sku_country_median:
                sku_billing_costs.append(sku_country_median[key])
            elif country in country_median:
                sku_billing_costs.append(country_median[country])
            else:
                sku_billing_costs.append(26.0)  # global fallback ~DE median
            unresolved_skus.add(sku)
    
    if uses_weight and total_weight > 0:
        # Weight-based pricing
        cost = gls_cost(country, total_weight)
        tier = 'weight_v3'
    elif all_resolved and total_weight > 0:
        cost = gls_cost(country, total_weight)
        tier = 'weight_v3'
    elif sku_billing_costs:
        # Mix: some weight, some billing
        if total_weight > 0:
            weight_cost = gls_cost(country, total_weight)
            cost = weight_cost + sum(sku_billing_costs)
        else:
            cost = max(sku_billing_costs) if len(sku_billing_costs) > 1 else sku_billing_costs[0]
        tier = 'sku_billing'
    else:
        cost = country_median.get(country, 26.0)
        tier = 'country_fb'
    
    results.append((amazon_id, round(cost, 2), tier))
    tier_counts[tier] = tier_counts.get(tier, 0) + 1

total_est = sum(r[1] for r in results)
print(f"\n  Results:")
print(f"    weight_v3 (TKL weights):  {tier_counts['weight_v3']:,} orders")
print(f"    sku_billing (historical): {tier_counts['sku_billing']:,} orders")
print(f"    country_fb (fallback):    {tier_counts['country_fb']:,} orders")
print(f"    Total: {len(results):,} orders = {total_est:,.2f} PLN")
print(f"    Unresolved SKUs: {len(unresolved_skus)} unique")

# Show some cost samples by country
by_country = defaultdict(list)
for amazon_id, cost, tier in results:
    c = order_country.get(amazon_id, '?')
    by_country[c].append(cost)

print(f"\n  Per country:")
for c in sorted(by_country.keys(), key=lambda x: -len(by_country[x])):
    costs = by_country[c]
    med = statistics.median(costs)
    print(f"    {c}: {len(costs):,} orders | median={med:.2f} PLN | total={sum(costs):,.2f} PLN")

if DRY_RUN:
    print("\nDRY RUN — no DB changes.")
    sys.exit(0)

# ── 6. DELETE OLD v1/v2 ESTIMATES ──
print(f"\n[6] Deleting old estimates (hist_country_v1, sku_country_v2)...")
conn = connect_acc(autocommit=True, timeout=120)
cur = conn.cursor()
cur.execute("SET LOCK_TIMEOUT 30000")

cur.execute("""
    DELETE f FROM dbo.acc_order_logistics_fact f
    WHERE f.calc_version IN ('hist_country_v1', 'sku_country_v2')
      AND f.amazon_order_id IN (
          SELECT amazon_order_id FROM dbo.acc_order WITH (NOLOCK)
          WHERE purchase_date >= '2026-02-01' AND purchase_date < '2026-03-12'
            AND fulfillment_channel = 'MFN'
      )
""")
deleted = cur.rowcount
print(f"  Deleted: {deleted:,} old estimate rows")

# ── 7. INSERT V3 ──
print(f"\n[7] Inserting {len(results):,} weight_v3 rows...")
now = datetime.utcnow()
inserted = 0
errors = 0

for i in range(0, len(results), BATCH_SIZE):
    batch = results[i:i+BATCH_SIZE]
    values_parts = []
    params = []
    for amazon_id, cost, tier in batch:
        values_parts.append("(?, ?, ?, ?, ?)")
        params.extend([amazon_id, 'weight_v3', cost, now, f'tkl_{tier}'])
    
    sql = f"""
        INSERT INTO dbo.acc_order_logistics_fact 
        (amazon_order_id, calc_version, total_logistics_pln, calculated_at, source_system)
        VALUES {', '.join(values_parts)}
    """
    try:
        cur.execute(sql, tuple(params))
        inserted += len(batch)
    except Exception as e:
        print(f"  ERROR batch {i}: {e}")
        errors += len(batch)
    
    if (i // BATCH_SIZE) % 10 == 0:
        print(f"  ... {inserted:,} inserted", flush=True)

print(f"  Inserted: {inserted:,}, Errors: {errors}")

# ── 8. VERIFY ──
print(f"\n[8] Verification...")
cur.execute("""
    SELECT f.calc_version, COUNT(*) as cnt, SUM(f.total_logistics_pln) as total
    FROM dbo.acc_order_logistics_fact f WITH (NOLOCK)
    JOIN dbo.acc_order o WITH (NOLOCK) ON o.amazon_order_id = f.amazon_order_id
    WHERE o.purchase_date >= '2026-02-01' AND o.purchase_date < '2026-03-12'
    GROUP BY f.calc_version
    ORDER BY cnt DESC
""")
print(f"\n  Final logistics_fact state (Feb-Mar):")
for row in cur.fetchall():
    print(f"    {row[0]}: {row[1]:,} rows = {float(row[2]):,.2f} PLN")

conn.close()
print(f"\nTotal time: {time.time()-t0:.0f}s")
print("Done!")

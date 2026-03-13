# Deep read: DHL sheet from Tabela Kosztow - this has product->box mapping
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "apps", "api"))
import openpyxl

tkl_dir = "N:\\Analityka\\TKL"
for f in os.listdir(tkl_dir):
    if 'tabela' in f.lower() and 'koszt' in f.lower():
        path = os.path.join(tkl_dir, f)
        break

wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

# DHL sheet: product -> box mapping with weights
ws = wb['DHL']
print("=== DHL Sheet: columns ===")
rows = list(ws.iter_rows(values_only=True, max_row=2))
header = rows[1] if len(rows) > 1 else rows[0]
for i, col in enumerate(header):
    print(f"  Col {i}: {col}")

print(f"\n=== Sample rows with box mapping ===")
count = 0
rows_all = list(ws.iter_rows(values_only=True, min_row=3, max_row=100))
for i, row in enumerate(rows_all):
    nr_art = row[0]
    nazwa = row[1]
    waga_bil_gls = row[3]    # Waga bilingowa GLS
    waga_kg = row[7]         # Waga (kg)
    sugestia_kartonu = row[10] # Sugestia kartonu
    czy_karton = row[11]     # CZY DOLICZYC KARTON
    waga_kartonu = row[12]   # Waga kartonu
    waga_bil_dhl = row[14]   # Waga billingowa DHL
    
    if sugestia_kartonu and str(sugestia_kartonu).startswith('999'):
        print(f"  Art={nr_art} | Nazwa={str(nazwa)[:40]} | Karton={sugestia_kartonu} | Waga={waga_kg}kg | WagaBilGLS={waga_bil_gls} | WagaBilDHL={waga_bil_dhl} | WagaKarton={waga_kartonu}")
        count += 1
    if count >= 20:
        break

# Count total with box mapping
print(f"\n=== Total products with box mapping ===")
all_rows = list(ws.iter_rows(values_only=True, min_row=3))
with_box = sum(1 for r in all_rows if r[10] and str(r[10]).startswith('999'))
manual = sum(1 for r in all_rows if r[10] and str(r[10]).lower().startswith('r'))
no_box = sum(1 for r in all_rows if not r[10] or str(r[10]) == '#N/A' or str(r[10]) == 'None')
print(f"  With box SKU (9999xxx):  {with_box}")
print(f"  Manual (Ręcznie):        {manual}")
print(f"  No mapping:              {no_box}")
print(f"  Total:                   {len(all_rows)}")

# Licznosc paczek: products that need 2+ parcels
ws2 = wb['Liczność paczek']
print(f"\n=== Liczność paczek (products needing 2+ parcels) ===")
paczki_rows = list(ws2.iter_rows(values_only=True, min_row=2))
print(f"  Total multi-parcel products: {len(paczki_rows)}")
for r in paczki_rows[:10]:
    print(f"  SKU={r[0]} parcels={r[1]}")

wb.close()
print("\nDone.")

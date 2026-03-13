# Read Baza opakowan from N:\Analityka\TKL
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "apps", "api"))

import openpyxl

path = "N:\\Analityka\\TKL\\Baza opakowan.xlsx"
# Try with polish chars if needed
if not os.path.exists(path):
    import glob
    candidates = glob.glob("N:\\Analityka\\TKL\\Baza opakowa*.xlsx")
    if candidates:
        path = candidates[0]
        print(f"Found: {path}")
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
print(f"Sheets: {wb.sheetnames}\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"=== Sheet: {sheet_name} ===")
    rows = list(ws.iter_rows(values_only=True))
    # Print first 50 rows
    for i, row in enumerate(rows[:50]):
        print(f"  Row {i}: {row}")
    if len(rows) > 50:
        print(f"  ... ({len(rows)} total rows)")
    print()

wb.close()

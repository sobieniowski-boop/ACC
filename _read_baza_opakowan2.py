# Read full Baza opakowan
import sys, os, glob
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "apps", "api"))
import openpyxl

# Find the main baza opakowan file (not 'skrocone')
tkl_dir = "N:\\Analityka\\TKL"
for f in os.listdir(tkl_dir):
    if 'opakow' in f.lower() and not 'skr' in f.lower() and not 'kopia' in f.lower():
        path = os.path.join(tkl_dir, f)
        break

print(f"Reading: {path}")
print(f"Size: {os.path.getsize(path):,} bytes\n")

wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
print(f"Sheets: {wb.sheetnames}\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"=== Sheet: {sheet_name} ===")
    rows = list(ws.iter_rows(values_only=True))
    for i, row in enumerate(rows[:80]):
        # Compact: skip empty rows
        if all(v is None for v in row):
            continue
        vals = [str(v) if v is not None else '' for v in row]
        print(f"  R{i}: {' | '.join(vals[:12])}")
    if len(rows) > 80:
        print(f"  ... ({len(rows)} total rows)")
    print()

wb.close()

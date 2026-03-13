"""
COGS Importer — reads purchase price XLSX files from 'cogs from sell' folder.

Handles varying column structures:
  - SKU column: looks for 'SKU', 'Nr art.', 'Symbol', 'SU', 'Nr artykulu'
  - Price column: looks for 'Cena netto', 'Cena', 'Netto', 'Cena zakupu'
  - Name column (optional): 'Nazwa', 'Opis', 'Nazwa produktu'

Flow:
  1. Scan folder for *.xlsx files
  2. Check acc_cogs_import_log to skip already-processed files
  3. For each new/modified file:
     a. Open workbook, detect SKU + price columns
     b. Read all rows with valid SKU + price
     c. Upsert into acc_purchase_price (source='cogs_xlsx', source_document=filename)
     d. Update acc_product.netto_purchase_price_pln for matching products
     e. Log file as processed

Can be run:
  - Manually: python -m app.services.cogs_importer
  - Via scheduler: every 30 min or on-demand
  - Via API endpoint: POST /api/v1/jobs/import-cogs
"""
from __future__ import annotations

import glob
import hashlib
import os
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Optional

import structlog

log = structlog.get_logger(__name__)

# Folder where purchase dept drops XLSX files
COGS_FOLDER = r"C:\ACC\cogs from sell"

# Column name variations (case-insensitive matching)
SKU_COLUMNS = {"sku", "nr art.", "nr art", "symbol", "su", "nr artykulu", "nr artykułu",
               "kod", "kod produktu", "indeks", "numer"}
PRICE_COLUMNS = {"cena netto", "cena netto ", "netto", "cena", "cena zakupu",
                 "cena netto pln", "cena_netto", "purchase price", "unit price"}
NAME_COLUMNS = {"nazwa", "opis", "nazwa produktu", "name", "opis produktu"}


def _normalize_header(val: Any) -> str:
    """Normalize header cell value for matching."""
    if val is None:
        return ""
    return str(val).strip().lower()


def _file_hash(filepath: str) -> str:
    """Compute MD5 hash of file for change detection."""
    h = hashlib.md5()
    with open(filepath, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def _detect_columns(ws) -> tuple[Optional[int], Optional[int], Optional[int]]:
    """
    Scan first 5 rows of worksheet to find SKU and Price column indices.
    Returns (sku_col_idx, price_col_idx, name_col_idx) — 0-based.
    Returns None for columns not found.
    """
    import openpyxl

    sku_col = None
    price_col = None
    name_col = None

    for row_idx, row in enumerate(ws.iter_rows(max_row=5, values_only=False)):
        for cell in row:
            header = _normalize_header(cell.value)
            if not header:
                continue

            col_idx = cell.column - 1  # 0-based

            if header in SKU_COLUMNS and sku_col is None:
                sku_col = col_idx
            elif header in PRICE_COLUMNS and price_col is None:
                price_col = col_idx
            elif header in NAME_COLUMNS and name_col is None:
                name_col = col_idx

        # If we found both required columns, stop scanning
        if sku_col is not None and price_col is not None:
            break

    return sku_col, price_col, name_col


def _parse_sku(val: Any) -> Optional[str]:
    """Parse SKU value — could be int, float, or string."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        # SKU like 33969 or 33969.0
        s = str(int(val))
    else:
        s = str(val).strip()
    return s if s and s != "0" else None


def _parse_price(val: Any) -> Optional[float]:
    """Parse price value — float, int, or string with comma."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        price = float(val)
    elif isinstance(val, str):
        cleaned = val.strip().replace(",", ".").replace(" ", "")
        if not cleaned:
            return None
        try:
            price = float(cleaned)
        except ValueError:
            return None
    else:
        return None

    # Sanity: price must be positive and reasonable
    # Cap: 2000 PLN — max legitimate single-unit purchase price for KADAX
    # (highest known product ~389 PLN). Values like 99999, 102998 are xlsx errors.
    if price <= 0 or price > 2_000:
        return None
    return round(price, 4)


def read_cogs_file(filepath: str) -> list[dict]:
    """
    Read a single COGS XLSX file.
    Returns list of {'sku': str, 'price': float, 'name': str|None}.
    """
    import openpyxl

    filename = os.path.basename(filepath)
    log.info("cogs_import.reading", file=filename)

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    results = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        sku_col, price_col, name_col = _detect_columns(ws)
        if sku_col is None or price_col is None:
            log.warning("cogs_import.columns_not_found",
                        file=filename, sheet=sheet_name,
                        sku_col=sku_col, price_col=price_col)
            continue

        log.info("cogs_import.columns_detected",
                 file=filename, sheet=sheet_name,
                 sku_col=sku_col, price_col=price_col, name_col=name_col)

        # Read data rows (skip header rows — start from row where we found headers + 1)
        header_row_found = False
        for row in ws.iter_rows(values_only=False):
            vals = [cell.value for cell in row]

            # Skip header row(s)
            header_check = _normalize_header(vals[sku_col] if sku_col < len(vals) else None)
            if header_check in SKU_COLUMNS:
                header_row_found = True
                continue

            if not header_row_found:
                continue

            sku_val = vals[sku_col] if sku_col < len(vals) else None
            price_val = vals[price_col] if price_col < len(vals) else None
            name_val = vals[name_col] if name_col is not None and name_col < len(vals) else None

            sku = _parse_sku(sku_val)
            price = _parse_price(price_val)

            if sku and price:
                results.append({
                    "sku": sku,
                    "price": price,
                    "name": str(name_val).strip()[:300] if name_val else None,
                })

    wb.close()

    log.info("cogs_import.file_parsed", file=filename, rows=len(results))
    return results


# ── Database operations ──

def _ensure_import_log_table(conn) -> None:
    """Create acc_cogs_import_log if it doesn't exist."""
    cur = conn.cursor()
    cur.execute("""
        IF NOT EXISTS (
            SELECT 1 FROM INFORMATION_SCHEMA.TABLES
            WHERE TABLE_NAME = 'acc_cogs_import_log'
        )
        CREATE TABLE acc_cogs_import_log (
            id            INT IDENTITY(1,1) PRIMARY KEY,
            filename      NVARCHAR(300)  NOT NULL,
            file_hash     NVARCHAR(32)   NOT NULL,
            rows_imported INT            DEFAULT 0,
            rows_updated  INT            DEFAULT 0,
            imported_at   DATETIME2      NOT NULL DEFAULT GETUTCDATE(),
            UNIQUE (filename, file_hash)
        )
    """)
    conn.commit()
    cur.close()


def _get_processed_files(conn) -> dict[str, str]:
    """Return {filename: file_hash} of already-processed files."""
    cur = conn.cursor()
    try:
        cur.execute("SELECT filename, file_hash FROM acc_cogs_import_log")
        result = {r[0]: r[1] for r in cur.fetchall()}
    except Exception:
        result = {}
    cur.close()
    return result


def _log_import(conn, filename: str, file_hash: str,
                rows_imported: int, rows_updated: int) -> None:
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO acc_cogs_import_log (filename, file_hash, rows_imported, rows_updated) "
        "VALUES (?, ?, ?, ?)",
        filename, file_hash, rows_imported, rows_updated,
    )
    conn.commit()
    cur.close()


def upsert_purchase_prices(conn, records: list[dict], source_document: str) -> tuple[int, int]:
    """
    Upsert records into acc_purchase_price + update acc_product cache.
    Returns (new_count, updated_count).
    """
    cur = conn.cursor()
    new_count = 0
    updated_count = 0
    today = date.today()

    for rec in records:
        sku = rec["sku"]
        price = rec["price"]

        # Check if this SKU already has a price from this source
        cur.execute(
            "SELECT id, netto_price_pln FROM acc_purchase_price "
            "WHERE internal_sku = ? AND source = 'cogs_xlsx' "
            "AND (valid_to IS NULL OR valid_to >= ?)",
            sku, today,
        )
        existing = cur.fetchone()

        if existing:
            old_price = float(existing[1]) if existing[1] else 0
            if abs(old_price - price) > 0.001:
                # Price changed — expire old record, insert new
                cur.execute(
                    "UPDATE acc_purchase_price SET valid_to = ?, updated_at = GETUTCDATE() "
                    "WHERE id = ?",
                    today, existing[0],
                )
                cur.execute(
                    "INSERT INTO acc_purchase_price "
                    "(internal_sku, netto_price_pln, valid_from, valid_to, "
                    " source, source_document, created_at, updated_at) "
                    "VALUES (?, ?, ?, NULL, 'cogs_xlsx', ?, GETUTCDATE(), GETUTCDATE())",
                    sku, price, today, source_document,
                )
                updated_count += 1
                # --- Controlling: log price change ---
                try:
                    from app.services.controlling import log_price_change
                    log_price_change(
                        conn,
                        internal_sku=sku,
                        old_price=old_price,
                        new_price=price,
                        source="cogs_xlsx",
                        source_document=source_document,
                    )
                except Exception:
                    pass  # controlling is non-blocking
            # else: same price, skip
        else:
            # New price record
            cur.execute(
                "INSERT INTO acc_purchase_price "
                "(internal_sku, netto_price_pln, valid_from, valid_to, "
                " source, source_document, created_at, updated_at) "
                "VALUES (?, ?, ?, NULL, 'cogs_xlsx', ?, GETUTCDATE(), GETUTCDATE())",
                sku, price, today, source_document,
            )
            new_count += 1
            # --- Controlling: log new price ---
            try:
                from app.services.controlling import log_price_change
                log_price_change(
                    conn,
                    internal_sku=sku,
                    old_price=None,
                    new_price=price,
                    source="cogs_xlsx",
                    source_document=source_document,
                )
            except Exception:
                pass  # controlling is non-blocking

        # Commit every 50 records
        if (new_count + updated_count) % 50 == 0:
            conn.commit()

    conn.commit()

    # Update acc_product.netto_purchase_price_pln cache for affected SKUs
    # Match internal_sku from acc_purchase_price to internal_sku in acc_product
    skus = [r["sku"] for r in records]
    if skus:
        # Batch update: set product price = latest valid purchase price
        cur.execute("""
            UPDATE p
            SET p.netto_purchase_price_pln = pp.netto_price_pln
            FROM acc_product p
            INNER JOIN (
                SELECT internal_sku, netto_price_pln,
                       ROW_NUMBER() OVER (
                           PARTITION BY internal_sku
                           ORDER BY valid_from DESC, created_at DESC
                       ) AS rn
                FROM acc_purchase_price
                WHERE (valid_to IS NULL OR valid_to >= GETDATE())
                  AND source = 'cogs_xlsx'
            ) pp ON pp.internal_sku = p.internal_sku AND pp.rn = 1
            WHERE p.internal_sku IS NOT NULL
        """)
        product_updates = cur.rowcount
        conn.commit()
        log.info("cogs_import.product_cache_updated", count=product_updates)

    cur.close()
    return new_count, updated_count


# ── Main scan & import function ──

def scan_and_import(folder: str = COGS_FOLDER) -> dict:
    """
    Scan folder for XLSX files, import new/changed ones.
    Returns summary dict.
    """
    from app.core.db_connection import connect_acc

    if not os.path.isdir(folder):
        log.warning("cogs_import.folder_missing", folder=folder)
        return {"error": f"Folder not found: {folder}", "files_processed": 0}

    conn = connect_acc(autocommit=False, timeout=60)
    _ensure_import_log_table(conn)
    processed_files = _get_processed_files(conn)

    xlsx_files = sorted(glob.glob(os.path.join(folder, "*.xlsx")))
    if not xlsx_files:
        conn.close()
        return {"files_found": 0, "files_processed": 0, "message": "No XLSX files in folder"}

    stats = {
        "files_found": len(xlsx_files),
        "files_processed": 0,
        "files_skipped": 0,
        "total_new": 0,
        "total_updated": 0,
        "errors": [],
    }

    for fpath in xlsx_files:
        filename = os.path.basename(fpath)
        fhash = _file_hash(fpath)

        # Skip if same file+hash already imported
        if filename in processed_files and processed_files[filename] == fhash:
            stats["files_skipped"] += 1
            log.debug("cogs_import.skip", file=filename, reason="already imported")
            continue

        try:
            records = read_cogs_file(fpath)
            if not records:
                log.warning("cogs_import.empty", file=filename)
                stats["errors"].append(f"{filename}: no valid rows found")
                continue

            new, updated = upsert_purchase_prices(conn, records, filename)
            _log_import(conn, filename, fhash, new, updated)

            stats["files_processed"] += 1
            stats["total_new"] += new
            stats["total_updated"] += updated

            log.info("cogs_import.file_done",
                     file=filename, new=new, updated=updated,
                     total_rows=len(records))

        except Exception as e:
            error_msg = f"{filename}: {str(e)[:200]}"
            stats["errors"].append(error_msg)
            log.error("cogs_import.file_error", file=filename, error=str(e))
            conn.rollback()

    conn.close()

    log.info("cogs_import.scan_complete", **{k: v for k, v in stats.items() if k != "errors"})
    return stats


# ── CLI entry point ──

def main():
    """Run COGS import from command line."""
    import json

    print(f"Scanning COGS folder: {COGS_FOLDER}")
    print()

    result = scan_and_import()

    print()
    print(f"{'=' * 50}")
    print(f"  Files found:     {result.get('files_found', 0)}")
    print(f"  Files processed: {result.get('files_processed', 0)}")
    print(f"  Files skipped:   {result.get('files_skipped', 0)}")
    print(f"  New prices:      {result.get('total_new', 0)}")
    print(f"  Updated prices:  {result.get('total_updated', 0)}")
    if result.get("errors"):
        print(f"  Errors:")
        for e in result["errors"]:
            print(f"    - {e}")
    print(f"{'=' * 50}")


if __name__ == "__main__":
    main()

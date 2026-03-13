from __future__ import annotations

import csv
import hashlib
import json
import os
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

import openpyxl
import structlog

from app.core.config import settings
from app.core.db_connection import connect_acc
from app.services.bl_order_lookup import resolve_bl_orders_to_acc_orders
from app.services.dhl_integration import ensure_dhl_schema
from app.services.dhl_registry_sync import _upsert_shipment, _upsert_shipment_links
from app.services.gls_integration import ensure_gls_schema

log = structlog.get_logger(__name__)

_BATCH_SIZE = 500
_CSV_FILENAME_RE = re.compile(r"GLS_(?P<document_number>\d+)\.csv$", re.IGNORECASE)
_BILLING_PERIOD_RE = re.compile(r"^\d{4}\.\d{2}$")
_YEAR_RE = re.compile(r"^\d{4}$")
_CORRECTION_MONTH_RE = re.compile(r"^(?P<month>\d{2})\.(?P<year>\d{4})$")
_CORRECTION_PATH_FRAGMENT = "korekty kosztowe"


@dataclass
class LocalPackageCandidate:
    amazon_order_id: str
    acc_order_id: str | None
    bl_order_id: int | None
    package_order_id: int | None
    courier_package_nr: str | None
    courier_inner_number: str | None


@dataclass
class ImportedGLSShipmentSeed:
    parcel_number: str
    row_date: date | None
    delivery_date: date | None
    parcel_status: str | None
    service_code: str | None
    note1: str | None
    recipient_name: str | None
    recipient_country: str | None
    billing_period: str | None
    total_amount: float
    line_count: int


def _connect():
    return connect_acc(autocommit=False, timeout=60)


def _normalize_text(value: Any) -> str:
    return str(value or "").strip()


def _normalize_identifier(value: Any) -> str:
    text = _normalize_text(value)
    if not text:
        return ""
    if text.endswith(".0"):
        prefix = text[:-2]
        if prefix.isdigit():
            return prefix
    return text.replace(" ", "")


def _normalize_int(value: Any) -> int | None:
    text = _normalize_identifier(value)
    if not text:
        return None
    try:
        return int(float(text))
    except Exception:
        return None


def _parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _normalize_text(value)
    if not text:
        return None
    for candidate in (text, text[:10]):
        try:
            return datetime.fromisoformat(candidate).date()
        except ValueError:
            continue
    return None


def _parse_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except Exception:
        text = _normalize_text(value).replace(",", ".")
        if not text:
            return None
        try:
            return float(text)
        except Exception:
            return None


def _json_dump(payload: Any) -> str:
    return json.dumps(payload, ensure_ascii=False, sort_keys=True)


def _hash_payload(payload: Any) -> str:
    return hashlib.sha256(_json_dump(payload).encode("utf-8")).hexdigest()


def _chunks(items: list[Any], size: int) -> Iterable[list[Any]]:
    for idx in range(0, len(items), size):
        yield items[idx : idx + size]


def _file_mtime_utc(path: Path) -> datetime:
    return datetime.utcfromtimestamp(path.stat().st_mtime).replace(microsecond=0)


def _billing_document_number_from_csv(path: Path) -> str | None:
    match = _CSV_FILENAME_RE.search(path.name)
    return match.group("document_number") if match else None


def _billing_period_from_path(path: Path) -> str | None:
    for part in reversed(path.parts):
        if _BILLING_PERIOD_RE.match(part):
            return part
        if _YEAR_RE.match(part):
            return part
    return None


def _billing_period_from_correction_path(path: Path) -> str | None:
    for part in reversed(path.parts):
        match = _CORRECTION_MONTH_RE.match(part)
        if match:
            return f"{match.group('year')}.{match.group('month')}"
    return None


def _is_gls_correction_path(path: Path) -> bool:
    return _CORRECTION_PATH_FRAGMENT in str(path).lower()


def _discover_gls_csv_files(root: Path) -> list[Path]:
    if not root.exists():
        return []
    files = []
    for path in root.rglob("GLS_*.csv"):
        if not path.is_file():
            continue
        lowered = str(path).lower()
        if "korekty kosztowe" in lowered:
            continue
        files.append(path.resolve())
    return sorted({path for path in files})


def _discover_gls_correction_xlsx_files(root: Path) -> list[Path]:
    if not root.exists():
        return []
    files = []
    for path in root.rglob("*.xlsx"):
        if not path.is_file():
            continue
        if not _is_gls_correction_path(path):
            continue
        if path.name.strip().lower() == "gls - bl.xlsx":
            continue
        files.append(path.resolve())
    return sorted({path for path in files})


def _read_csv_rows(path: str | os.PathLike[str]) -> tuple[list[str], list[list[str]]]:
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "utf-8", "cp1250", "latin-1"):
        try:
            with Path(path).open("r", encoding=encoding, newline="") as handle:
                reader = csv.reader(handle, delimiter=";")
                rows = list(reader)
            if not rows:
                return [], []
            header = [str(cell or "").strip() for cell in rows[0]]
            return header, rows[1:]
        except UnicodeDecodeError as exc:
            last_error = exc
            continue
    if last_error is not None:
        raise last_error
    return [], []


def _normalize_header_label(value: Any) -> str:
    text = _normalize_text(value).lower()
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return " ".join(text.split())


def _find_header_index(header_row: list[Any], *substrings: str, from_end: bool = False) -> int | None:
    needles = [part.lower() for part in substrings if part]
    indexed = list(enumerate(header_row))
    if from_end:
        indexed = list(reversed(indexed))
    for idx, value in indexed:
        label = _normalize_header_label(value)
        if label and all(needle in label for needle in needles):
            return idx
    return None


def parse_gls_billing_csv(path: str | os.PathLike[str]) -> list[dict[str, Any]]:
    source_path = Path(path)
    header_row, data_rows = _read_csv_rows(source_path)
    if not header_row:
        return []

    header_map = {name: idx for idx, name in enumerate(header_row) if name}
    document_number = _normalize_identifier(
        _billing_document_number_from_csv(source_path)
    )
    billing_period = _billing_period_from_path(source_path)

    def cell(row: list[str], name: str) -> str | None:
        idx = header_map.get(name)
        if idx is None or idx >= len(row):
            return None
        value = row[idx]
        return value if value is not None else None

    parsed: list[dict[str, Any]] = []
    for row_no, row in enumerate(data_rows, start=2):
        parcel_number = _normalize_identifier(cell(row, "parcel_num"))
        row_document_number = _normalize_identifier(cell(row, "invoice_num")) or document_number
        if not parcel_number or not row_document_number:
            continue
        parsed.append(
            {
                "document_number": row_document_number,
                "billing_period": billing_period,
                "row_date": _parse_date(cell(row, "date")),
                "delivery_date": _parse_date(cell(row, "delivery_date_x")),
                "parcel_number": parcel_number,
                "recipient_name": _normalize_text(cell(row, "rname1")) or None,
                "recipient_postal_code": _normalize_text(cell(row, "rpost")) or None,
                "recipient_city": _normalize_text(cell(row, "rcity")) or None,
                "recipient_country": _normalize_text(cell(row, "rcountry")) or None,
                "weight": _parse_float(cell(row, "weight")),
                "declared_weight": _parse_float(cell(row, "weight_declared")),
                "billing_weight": _parse_float(cell(row, "weight_billing")),
                "net_amount": _parse_float(cell(row, "netto")),
                "toll_amount": _parse_float(cell(row, "toll")),
                "fuel_amount": _parse_float(cell(row, "fuel_surcharge")),
                "storewarehouse_amount": _parse_float(cell(row, "storewarehouse_price")),
                "surcharge_amount": _parse_float(cell(row, "surcharge")),
                "billing_type": _normalize_text(cell(row, "billing_type")) or None,
                "note1": _normalize_identifier(cell(row, "note1")) or None,
                "dimension_combined": _normalize_text(cell(row, "dim_combined")) or None,
                "volumetric_weight": _parse_float(cell(row, "weight_volumetric")),
                "parcel_status": _normalize_text(cell(row, "parcel_status")) or None,
                "service_code": _normalize_text(cell(row, "srv")) or None,
                "source_row_no": row_no,
            }
        )
    return parsed


def parse_gls_billing_correction_xlsx(path: str | os.PathLike[str]) -> list[dict[str, Any]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        rows = worksheet.iter_rows(min_row=1, values_only=True)
        header_row = next(rows, None)
        if header_row is None:
            return []
        header = list(header_row)
        parcel_idx = _find_header_index(header, "nr", "pacz")
        document_idx = _find_header_index(header, "dok", "faktur")
        issue_date_idx = _find_header_index(header, "data", "wystaw")
        sales_date_idx = _find_header_index(header, "data", "sprzed")
        recipient_name_idx = _find_header_index(header, "odbiorca")
        recipient_postal_idx = _find_header_index(header, "kod", "poczt", from_end=True)
        recipient_city_idx = _find_header_index(header, "miejsc", from_end=True)
        recipient_country_idx = _find_header_index(header, "kraj", from_end=True)
        original_net_idx = _find_header_index(header, "warto", "netto")
        corrected_net_idx = _find_header_index(header, "prawid", "kwota")
        delta_idx = _find_header_index(header, "rozn")
        fuel_rate_idx = _find_header_index(header, "stawka", "paliw")
        fuel_correction_idx = _find_header_index(header, "korekt", "paliw")
        toll_idx = _find_header_index(header, "drogow")

        if parcel_idx is None or document_idx is None or original_net_idx is None or corrected_net_idx is None:
            raise ValueError("unsupported_correction_layout")

        parsed: list[dict[str, Any]] = []
        billing_period = _billing_period_from_correction_path(Path(path))
        for row_no, row in enumerate(rows, start=2):
            parcel_number = _normalize_identifier(row[parcel_idx] if parcel_idx < len(row) else None)
            document_number = _normalize_identifier(row[document_idx] if document_idx < len(row) else None)
            if not parcel_number or not document_number:
                continue
            fuel_rate_raw = (
                row[fuel_rate_idx] if fuel_rate_idx is not None and fuel_rate_idx < len(row) else None
            )
            fuel_rate_pct = _parse_float(fuel_rate_raw)
            if fuel_rate_pct is None and fuel_rate_raw not in (None, ""):
                fuel_rate_text = _normalize_text(fuel_rate_raw).replace("%", "").replace(",", ".")
                try:
                    fuel_rate_pct = float(fuel_rate_text)
                except Exception:
                    fuel_rate_pct = None
            if fuel_rate_pct is not None and fuel_rate_pct > 1:
                fuel_rate_pct = fuel_rate_pct / 100.0
            parsed.append(
                {
                    "document_number": document_number,
                    "billing_period": billing_period,
                    "issue_date": _parse_date(
                        row[issue_date_idx]
                        if issue_date_idx is not None and issue_date_idx < len(row)
                        else None
                    ),
                    "sales_date": _parse_date(
                        row[sales_date_idx]
                        if sales_date_idx is not None and sales_date_idx < len(row)
                        else None
                    ),
                    "parcel_number": parcel_number,
                    "recipient_name": _normalize_text(
                        row[recipient_name_idx]
                        if recipient_name_idx is not None and recipient_name_idx < len(row)
                        else None
                    )
                    or None,
                    "recipient_postal_code": _normalize_text(
                        row[recipient_postal_idx]
                        if recipient_postal_idx is not None and recipient_postal_idx < len(row)
                        else None
                    )
                    or None,
                    "recipient_city": _normalize_text(
                        row[recipient_city_idx]
                        if recipient_city_idx is not None and recipient_city_idx < len(row)
                        else None
                    )
                    or None,
                    "recipient_country": _normalize_text(
                        row[recipient_country_idx]
                        if recipient_country_idx is not None and recipient_country_idx < len(row)
                        else None
                    )
                    or None,
                    "original_net_amount": _parse_float(
                        row[original_net_idx] if original_net_idx < len(row) else None
                    ),
                    "corrected_net_amount": _parse_float(
                        row[corrected_net_idx] if corrected_net_idx < len(row) else None
                    ),
                    "net_delta_amount": _parse_float(
                        row[delta_idx] if delta_idx is not None and delta_idx < len(row) else None
                    ),
                    "fuel_rate_pct": fuel_rate_pct,
                    "fuel_correction_amount": _parse_float(
                        row[fuel_correction_idx]
                        if fuel_correction_idx is not None and fuel_correction_idx < len(row)
                        else None
                    ),
                    "toll_amount": _parse_float(
                        row[toll_idx] if toll_idx is not None and toll_idx < len(row) else None
                    ),
                    "source_row_no": row_no,
                }
            )
        return parsed
    finally:
        workbook.close()


def parse_gls_bl_map_xlsx(path: str | os.PathLike[str]) -> list[dict[str, Any]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        rows = worksheet.iter_rows(min_row=1, values_only=True)
        header_row = next(rows, None)
        if header_row is None:
            return []
        header_map = {
            _normalize_text(cell): idx
            for idx, cell in enumerate(header_row)
            if _normalize_text(cell)
        }
        parsed: list[dict[str, Any]] = []
        for row_no, row in enumerate(rows, start=2):
            tracking_number = _normalize_identifier(
                row[header_map["tracking_number"]]
                if "tracking_number" in header_map and header_map["tracking_number"] < len(row)
                else None
            )
            bl_order_id = _normalize_int(
                row[header_map["order_id"]]
                if "order_id" in header_map and header_map["order_id"] < len(row)
                else None
            )
            if not tracking_number or bl_order_id is None:
                continue
            custom_1 = (
                _normalize_text(row[header_map["custom_1"]])
                if "custom_1" in header_map and header_map["custom_1"] < len(row)
                else ""
            )
            parsed.append(
                {
                    "tracking_number": tracking_number,
                    "bl_order_id": bl_order_id,
                    "map_source": custom_1 or None,
                    "source_row_no": row_no,
                }
            )
        return parsed
    finally:
        workbook.close()


def _load_import_file_state(
    cur, *, source_kind: str, file_path: str
) -> tuple[int | None, datetime | None, str | None] | None:
    cur.execute(
        """
        SELECT file_size_bytes, file_mtime_utc, status
        FROM dbo.acc_gls_import_file WITH (NOLOCK)
        WHERE source_kind = ? AND file_path = ?
        """,
        [source_kind, file_path],
    )
    row = cur.fetchone()
    if not row:
        return None
    size = int(row[0]) if row[0] is not None else None
    mtime = row[1]
    if isinstance(mtime, datetime):
        mtime = mtime.replace(microsecond=0)
    status = str(row[2] or "").strip().lower() or None
    return size, mtime, status


def _should_skip_import_file(
    state: tuple[int | None, datetime | None, str | None] | None,
    *,
    file_size: int,
    file_mtime: datetime,
    force_reimport: bool,
) -> bool:
    if force_reimport or state is None:
        return False
    prev_size, prev_mtime, prev_status = state
    return prev_status == "imported" and prev_size == file_size and prev_mtime == file_mtime


def _upsert_import_file(
    cur,
    *,
    source_kind: str,
    file_path: str,
    file_name: str,
    document_number: str | None,
    file_size_bytes: int,
    file_mtime_utc: datetime,
    status: str,
    rows_imported: int,
    error_message: str | None,
) -> None:
    cur.execute(
        """
        SELECT CAST(id AS NVARCHAR(40))
        FROM dbo.acc_gls_import_file WITH (NOLOCK)
        WHERE source_kind = ? AND file_path = ?
        """,
        [source_kind, file_path],
    )
    row = cur.fetchone()
    if row:
        cur.execute(
            """
            UPDATE dbo.acc_gls_import_file
            SET file_name = ?,
                document_number = ?,
                file_size_bytes = ?,
                file_mtime_utc = ?,
                status = ?,
                rows_imported = ?,
                error_message = ?,
                last_imported_at = SYSUTCDATETIME(),
                updated_at = SYSUTCDATETIME()
            WHERE id = CAST(? AS UNIQUEIDENTIFIER)
            """,
            [
                file_name,
                document_number,
                file_size_bytes,
                file_mtime_utc,
                status,
                rows_imported,
                error_message,
                str(row[0]),
            ],
        )
        return

    cur.execute(
        """
        INSERT INTO dbo.acc_gls_import_file (
            source_kind, file_path, file_name, document_number, file_size_bytes, file_mtime_utc,
            status, rows_imported, error_message, last_imported_at, updated_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, SYSUTCDATETIME(), SYSUTCDATETIME())
        """,
        [
            source_kind,
            file_path,
            file_name,
            document_number,
            file_size_bytes,
            file_mtime_utc,
            status,
            rows_imported,
            error_message,
        ],
    )


def _upsert_billing_document(
    cur,
    *,
    document_number: str,
    billing_period: str | None,
    source_file: str,
    detail_rows_count: int,
) -> None:
    cur.execute(
        """
        SELECT document_number
        FROM dbo.acc_gls_billing_document WITH (NOLOCK)
        WHERE document_number = ?
        """,
        [document_number],
    )
    exists = cur.fetchone() is not None
    if exists:
        cur.execute(
            """
            UPDATE dbo.acc_gls_billing_document
            SET billing_period = COALESCE(?, billing_period),
                source_file = ?,
                detail_rows_count = ?,
                last_imported_at = SYSUTCDATETIME(),
                updated_at = SYSUTCDATETIME()
            WHERE document_number = ?
            """,
            [billing_period, source_file, detail_rows_count, document_number],
        )
        return
    cur.execute(
        """
        INSERT INTO dbo.acc_gls_billing_document (
            document_number, billing_period, source_file, detail_rows_count,
            last_imported_at, updated_at
        )
        VALUES (?, ?, ?, ?, SYSUTCDATETIME(), SYSUTCDATETIME())
        """,
        [document_number, billing_period, source_file, detail_rows_count],
    )


def _replace_gls_billing_lines(cur, *, source_file: str, rows: list[dict[str, Any]]) -> int:
    cur.execute("DELETE FROM dbo.acc_gls_billing_line WHERE source_file = ?", [source_file])
    if not rows:
        return 0
    insert_sql = """
        INSERT INTO dbo.acc_gls_billing_line (
            document_number, billing_period, row_date, delivery_date, parcel_number,
            recipient_name, recipient_postal_code, recipient_city, recipient_country,
            weight, declared_weight, billing_weight,
            net_amount, toll_amount, fuel_amount, storewarehouse_amount, surcharge_amount,
            billing_type, note1, dimension_combined, volumetric_weight, parcel_status,
            service_code, source_file, source_row_no, source_hash, imported_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, SYSUTCDATETIME())
    """
    payloads: list[list[Any]] = []
    for row in rows:
        source_hash = _hash_payload(
            {
                "document_number": row.get("document_number"),
                "parcel_number": row.get("parcel_number"),
                "source_row_no": row.get("source_row_no"),
            }
        )
        payloads.append(
            [
                row.get("document_number"),
                row.get("billing_period"),
                row.get("row_date"),
                row.get("delivery_date"),
                row.get("parcel_number"),
                row.get("recipient_name"),
                row.get("recipient_postal_code"),
                row.get("recipient_city"),
                row.get("recipient_country"),
                row.get("weight"),
                row.get("declared_weight"),
                row.get("billing_weight"),
                row.get("net_amount"),
                row.get("toll_amount"),
                row.get("fuel_amount"),
                row.get("storewarehouse_amount"),
                row.get("surcharge_amount"),
                row.get("billing_type"),
                row.get("note1"),
                row.get("dimension_combined"),
                row.get("volumetric_weight"),
                row.get("parcel_status"),
                row.get("service_code"),
                source_file,
                row.get("source_row_no"),
                source_hash,
            ]
        )
    for chunk in _chunks(payloads, _BATCH_SIZE):
        cur.executemany(insert_sql, chunk)
    return len(payloads)


def _replace_gls_billing_correction_lines(cur, *, source_file: str, rows: list[dict[str, Any]]) -> int:
    cur.execute("DELETE FROM dbo.acc_gls_billing_correction_line WHERE source_file = ?", [source_file])
    if not rows:
        return 0
    insert_sql = """
        INSERT INTO dbo.acc_gls_billing_correction_line (
            document_number, issue_date, sales_date, parcel_number,
            recipient_name, recipient_postal_code, recipient_city, recipient_country,
            original_net_amount, corrected_net_amount, net_delta_amount,
            fuel_rate_pct, fuel_correction_amount, toll_amount,
            source_file, source_row_no, source_hash, imported_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, SYSUTCDATETIME())
    """
    payloads: list[list[Any]] = []
    for row in rows:
        source_hash = _hash_payload(
            {
                "document_number": row.get("document_number"),
                "parcel_number": row.get("parcel_number"),
                "source_row_no": row.get("source_row_no"),
            }
        )
        payloads.append(
            [
                row.get("document_number"),
                row.get("issue_date"),
                row.get("sales_date"),
                row.get("parcel_number"),
                row.get("recipient_name"),
                row.get("recipient_postal_code"),
                row.get("recipient_city"),
                row.get("recipient_country"),
                row.get("original_net_amount"),
                row.get("corrected_net_amount"),
                row.get("net_delta_amount"),
                row.get("fuel_rate_pct"),
                row.get("fuel_correction_amount"),
                row.get("toll_amount"),
                source_file,
                row.get("source_row_no"),
                source_hash,
            ]
        )
    for chunk in _chunks(payloads, _BATCH_SIZE):
        cur.executemany(insert_sql, chunk)
    return len(payloads)


def _replace_gls_bl_map_rows(cur, *, source_file: str, rows: list[dict[str, Any]]) -> int:
    cur.execute("DELETE FROM dbo.acc_gls_bl_map WHERE source_file = ?", [source_file])
    if not rows:
        return 0
    insert_sql = """
        INSERT INTO dbo.acc_gls_bl_map (
            tracking_number, bl_order_id, map_source, source_file, source_row_no, source_hash, imported_at
        )
        VALUES (?, ?, ?, ?, ?, ?, SYSUTCDATETIME())
    """
    payloads: list[list[Any]] = []
    for row in rows:
        source_hash = _hash_payload(
            {
                "tracking_number": row.get("tracking_number"),
                "bl_order_id": row.get("bl_order_id"),
                "source_row_no": row.get("source_row_no"),
            }
        )
        payloads.append(
            [
                row.get("tracking_number"),
                row.get("bl_order_id"),
                row.get("map_source"),
                source_file,
                row.get("source_row_no"),
                source_hash,
            ]
        )
    for chunk in _chunks(payloads, _BATCH_SIZE):
        cur.executemany(insert_sql, chunk)
    return len(payloads)


def _normalize_token(value: Any) -> str:
    return _normalize_identifier(value).upper()


def _to_datetime(value: date | datetime | None) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    return datetime(value.year, value.month, value.day)


def _is_gls_package(courier_code: Any, courier_other_name: Any) -> bool:
    code = _normalize_text(courier_code).lower()
    other = _normalize_text(courier_other_name).lower()
    if code == "blconnectpackages":
        code = other
    return "gls" in code or "gls" in other


def _is_delivered_status(status: str | None, delivery_date: date | None) -> bool:
    if delivery_date is not None:
        return True
    normalized = _normalize_text(status).lower()
    return "dor" in normalized or "deliver" in normalized


def _append_tracking_candidate(
    result_bucket: dict[str, list[LocalPackageCandidate]],
    *,
    key: str,
    candidate: LocalPackageCandidate,
) -> None:
    if not key:
        return
    result_bucket.setdefault(key, [])
    marker = (
        candidate.amazon_order_id,
        candidate.acc_order_id or "",
        candidate.bl_order_id or 0,
        candidate.package_order_id or 0,
    )
    existing = {
        (
            item.amazon_order_id,
            item.acc_order_id or "",
            item.bl_order_id or 0,
            item.package_order_id or 0,
        )
        for item in result_bucket[key]
    }
    if marker not in existing:
        result_bucket[key].append(candidate)


def _collect_tracking_candidates(
    cur,
    *,
    tracking_values: list[str],
    result_bucket: dict[str, list[LocalPackageCandidate]],
) -> None:
    normalized_values = sorted(
        {
            _normalize_identifier(value)
            for value in tracking_values
            if _normalize_identifier(value)
        }
    )
    if not normalized_values:
        return
    for batch in _chunks(normalized_values, _BATCH_SIZE):
        placeholders = ",".join("?" for _ in batch)
        params = [*batch, *batch, *batch, *batch, *batch]
        cur.execute(
            f"""
WITH matched_packages AS (
    SELECT DISTINCT
        p.courier_package_nr AS matched_tracking,
        p.order_id AS package_order_id,
        p.courier_package_nr,
        p.courier_inner_number,
        p.courier_code,
        p.courier_other_name,
        COALESCE(dm.holding_order_id, p.order_id) AS resolved_bl_order_id
    FROM dbo.acc_cache_packages p WITH (NOLOCK)
    LEFT JOIN dbo.acc_cache_dis_map dm WITH (NOLOCK)
      ON dm.dis_order_id = p.order_id
    WHERE p.courier_package_nr IN ({placeholders})
    UNION
    SELECT DISTINCT
        p.courier_inner_number AS matched_tracking,
        p.order_id AS package_order_id,
        p.courier_package_nr,
        p.courier_inner_number,
        p.courier_code,
        p.courier_other_name,
        COALESCE(dm.holding_order_id, p.order_id) AS resolved_bl_order_id
    FROM dbo.acc_cache_packages p WITH (NOLOCK)
    LEFT JOIN dbo.acc_cache_dis_map dm WITH (NOLOCK)
      ON dm.dis_order_id = p.order_id
    WHERE p.courier_inner_number IN ({placeholders})
    UNION
    SELECT DISTINCT
        p.courier_package_nr AS matched_tracking,
        CAST(NULL AS BIGINT) AS package_order_id,
        p.courier_package_nr,
        p.courier_inner_number,
        p.courier_code,
        p.courier_other_name,
        COALESCE(dm.holding_order_id, p.order_id) AS resolved_bl_order_id
    FROM dbo.acc_bl_distribution_package_cache p WITH (NOLOCK)
    LEFT JOIN dbo.acc_cache_dis_map dm WITH (NOLOCK)
      ON dm.dis_order_id = p.order_id
    WHERE p.courier_package_nr IN ({placeholders})
    UNION
    SELECT DISTINCT
        p.courier_inner_number AS matched_tracking,
        CAST(NULL AS BIGINT) AS package_order_id,
        p.courier_package_nr,
        p.courier_inner_number,
        p.courier_code,
        p.courier_other_name,
        COALESCE(dm.holding_order_id, p.order_id) AS resolved_bl_order_id
    FROM dbo.acc_bl_distribution_package_cache p WITH (NOLOCK)
    LEFT JOIN dbo.acc_cache_dis_map dm WITH (NOLOCK)
      ON dm.dis_order_id = p.order_id
    WHERE p.courier_inner_number IN ({placeholders})
    UNION
    SELECT DISTINCT
        o.delivery_package_nr AS matched_tracking,
        CAST(NULL AS BIGINT) AS package_order_id,
        o.delivery_package_nr AS courier_package_nr,
        CAST(NULL AS NVARCHAR(255)) AS courier_inner_number,
        o.delivery_package_module AS courier_code,
        o.delivery_method AS courier_other_name,
        COALESCE(dm.holding_order_id, o.order_id) AS resolved_bl_order_id
    FROM dbo.acc_bl_distribution_order_cache o WITH (NOLOCK)
    LEFT JOIN dbo.acc_cache_dis_map dm WITH (NOLOCK)
      ON dm.dis_order_id = o.order_id
    WHERE o.delivery_package_nr IN ({placeholders})
)
SELECT DISTINCT
    mp.matched_tracking,
    mp.resolved_bl_order_id,
    mp.package_order_id,
    mp.courier_package_nr,
    mp.courier_inner_number,
    mp.courier_code,
    mp.courier_other_name
FROM matched_packages mp
            """,
            params,
        )
        rows = cur.fetchall()
        resolved_orders = resolve_bl_orders_to_acc_orders(
            cur,
            bl_order_ids=[
                int(row[1])
                for row in rows
                if row[1] is not None
            ],
        )
        for row in rows:
            if not _is_gls_package(row[5], row[6]):
                continue
            resolved = resolved_orders.get(int(row[1])) if row[1] is not None else None
            if not resolved:
                continue
            candidate = LocalPackageCandidate(
                amazon_order_id=resolved.amazon_order_id,
                acc_order_id=resolved.acc_order_id,
                bl_order_id=resolved.bl_order_id,
                package_order_id=int(row[2]) if row[2] is not None else None,
                courier_package_nr=_normalize_text(row[3]) or None,
                courier_inner_number=_normalize_text(row[4]) or None,
            )
            if not candidate.amazon_order_id:
                continue
            _append_tracking_candidate(
                result_bucket,
                key=_normalize_token(row[0]),
                candidate=candidate,
            )


def _load_bl_order_candidates(
    cur,
    *,
    bl_order_values: list[str],
) -> dict[str, list[LocalPackageCandidate]]:
    result: dict[str, list[LocalPackageCandidate]] = {}
    normalized_values = sorted({_normalize_int(value) for value in bl_order_values if _normalize_int(value) is not None})
    if not normalized_values:
        return result
    resolved_orders = resolve_bl_orders_to_acc_orders(cur, bl_order_ids=normalized_values)
    for bl_order_id, resolved in resolved_orders.items():
        key = str(bl_order_id)
        candidate = LocalPackageCandidate(
            amazon_order_id=resolved.amazon_order_id,
            acc_order_id=resolved.acc_order_id,
            bl_order_id=resolved.bl_order_id,
            package_order_id=None,
            courier_package_nr=None,
            courier_inner_number=None,
        )
        result.setdefault(key, [])
        marker = (
            candidate.amazon_order_id,
            candidate.acc_order_id or "",
            candidate.bl_order_id or 0,
        )
        existing = {
            (
                item.amazon_order_id,
                item.acc_order_id or "",
                item.bl_order_id or 0,
            )
            for item in result[key]
        }
        if marker not in existing:
            result[key].append(candidate)
    return result


def _load_bl_map_tracking_candidates(
    cur,
    *,
    tracking_values: list[str],
) -> dict[str, list[LocalPackageCandidate]]:
    result: dict[str, list[LocalPackageCandidate]] = {}
    normalized_values = sorted({_normalize_identifier(value) for value in tracking_values if _normalize_identifier(value)})
    if not normalized_values:
        return result
    for batch in _chunks(normalized_values, _BATCH_SIZE):
        placeholders = ",".join("?" for _ in batch)
        cur.execute(
            f"""
SELECT DISTINCT
    m.tracking_number,
    m.bl_order_id
FROM dbo.acc_gls_bl_map m WITH (NOLOCK)
WHERE m.tracking_number IN ({placeholders})
            """,
            batch,
        )
        rows = cur.fetchall()
        resolved_orders = resolve_bl_orders_to_acc_orders(
            cur,
            bl_order_ids=[
                int(row[1])
                for row in rows
                if row[1] is not None
            ],
        )
        for row in rows:
            key = _normalize_token(row[0])
            if not key:
                continue
            resolved = resolved_orders.get(int(row[1])) if row[1] is not None else None
            if not resolved:
                continue
            candidate = LocalPackageCandidate(
                amazon_order_id=resolved.amazon_order_id,
                acc_order_id=resolved.acc_order_id,
                bl_order_id=resolved.bl_order_id,
                package_order_id=None,
                courier_package_nr=_normalize_text(row[0]) or None,
                courier_inner_number=None,
            )
            if not candidate.amazon_order_id:
                continue
            result.setdefault(key, [])
            marker = (
                candidate.amazon_order_id,
                candidate.acc_order_id or "",
                candidate.bl_order_id or 0,
            )
            existing = {
                (
                    item.amazon_order_id,
                    item.acc_order_id or "",
                    item.bl_order_id or 0,
                )
                for item in result[key]
            }
            if marker not in existing:
                result[key].append(candidate)
    return result


def _append_link_candidate(
    bucket: dict[tuple[str, str, str], dict[str, Any]],
    *,
    candidate: LocalPackageCandidate,
    link_method: str,
    confidence: float,
) -> None:
    key = (candidate.amazon_order_id, link_method, candidate.acc_order_id or "")
    payload = {
        "amazon_order_id": candidate.amazon_order_id,
        "acc_order_id": candidate.acc_order_id,
        "bl_order_id": candidate.bl_order_id,
        "link_method": link_method,
        "link_confidence": confidence,
        "is_primary": False,
    }
    existing = bucket.get(key)
    if existing is None or confidence > float(existing.get("link_confidence") or 0):
        bucket[key] = payload


def _collect_seed_links(
    *,
    seed: ImportedGLSShipmentSeed,
    by_tracking: dict[str, list[LocalPackageCandidate]],
    by_tracking_bl_map: dict[str, list[LocalPackageCandidate]],
    by_bl_order: dict[str, list[LocalPackageCandidate]],
) -> list[dict[str, Any]]:
    bucket: dict[tuple[str, str, str], dict[str, Any]] = {}

    for candidate in by_tracking.get(_normalize_token(seed.parcel_number), []):
        _append_link_candidate(
            bucket,
            candidate=candidate,
            link_method="billing_parcel_tracking",
            confidence=1.0,
        )

    for candidate in by_tracking_bl_map.get(_normalize_token(seed.parcel_number), []):
        _append_link_candidate(
            bucket,
            candidate=candidate,
            link_method="billing_tracking_bl_map",
            confidence=0.96,
        )

    if seed.note1:
        for candidate in by_bl_order.get(_normalize_identifier(seed.note1), []):
            _append_link_candidate(
                bucket,
                candidate=candidate,
                link_method="billing_note1_bl_order",
                confidence=0.92,
            )

    links = list(bucket.values())
    if not links:
        return links
    best_confidence = max(float(item.get("link_confidence") or 0) for item in links)
    best_orders = {
        _normalize_text(item.get("amazon_order_id"))
        for item in links
        if float(item.get("link_confidence") or 0) == best_confidence
    }
    if len(best_orders) == 1:
        primary_order = next(iter(best_orders))
        primary_set = False
        for item in links:
            if (
                not primary_set
                and _normalize_text(item.get("amazon_order_id")) == primary_order
                and float(item.get("link_confidence") or 0) == best_confidence
            ):
                item["is_primary"] = True
                primary_set = True
    return links


def _load_gls_shipment_seeds(
    cur,
    *,
    parcel_numbers: list[str],
) -> list[ImportedGLSShipmentSeed]:
    if not parcel_numbers:
        return []
    seeds: list[ImportedGLSShipmentSeed] = []
    for batch in _chunks(parcel_numbers, _BATCH_SIZE):
        placeholders = ",".join("?" for _ in batch)
        cur.execute(
            f"""
SELECT
    l.parcel_number,
    MIN(l.row_date) AS row_date,
    MAX(l.delivery_date) AS delivery_date,
    MAX(l.parcel_status) AS parcel_status,
    MAX(l.service_code) AS service_code,
    MAX(l.note1) AS note1,
    MAX(l.recipient_name) AS recipient_name,
    MAX(l.recipient_country) AS recipient_country,
    MAX(l.billing_period) AS billing_period,
    CAST(
        SUM(
            ISNULL(l.net_amount, 0)
            + ISNULL(l.toll_amount, 0)
            + ISNULL(l.fuel_amount, 0)
            + ISNULL(l.storewarehouse_amount, 0)
            + ISNULL(l.surcharge_amount, 0)
        ) AS FLOAT
    ) AS total_amount,
    COUNT(*) AS line_count
FROM dbo.acc_gls_billing_line l WITH (NOLOCK)
WHERE l.parcel_number IN ({placeholders})
GROUP BY l.parcel_number
            """,
            batch,
        )
        for row in cur.fetchall():
            seeds.append(
                ImportedGLSShipmentSeed(
                    parcel_number=_normalize_identifier(row[0]),
                    row_date=row[1],
                    delivery_date=row[2],
                    parcel_status=_normalize_text(row[3]) or None,
                    service_code=_normalize_text(row[4]) or None,
                    note1=_normalize_identifier(row[5]) or None,
                    recipient_name=_normalize_text(row[6]) or None,
                    recipient_country=_normalize_text(row[7]) or None,
                    billing_period=_normalize_text(row[8]) or None,
                    total_amount=float(row[9] or 0),
                    line_count=int(row[10] or 0),
                )
            )
    return seeds


def _load_all_billing_parcel_numbers(cur) -> set[str]:
    cur.execute(
        """
SELECT DISTINCT parcel_number
FROM dbo.acc_gls_billing_line WITH (NOLOCK)
WHERE parcel_number IS NOT NULL
  AND parcel_number <> ''
        """
    )
    return {_normalize_identifier(row[0]) for row in cur.fetchall() if _normalize_identifier(row[0])}


def _load_windowed_billing_parcel_numbers(
    cur,
    *,
    created_from: date | None = None,
    created_to: date | None = None,
    limit_parcels: int | None = None,
) -> set[str]:
    where = [
        "parcel_number IS NOT NULL",
        "parcel_number <> ''",
    ]
    params: list[Any] = []
    if created_from:
        where.append("CAST(COALESCE(delivery_date, row_date) AS DATE) >= ?")
        params.append(created_from.isoformat())
    if created_to:
        where.append("CAST(COALESCE(delivery_date, row_date) AS DATE) <= ?")
        params.append(created_to.isoformat())

    top_sql = f"TOP {int(limit_parcels)} " if limit_parcels else ""
    cur.execute(
        f"""
SELECT DISTINCT {top_sql} parcel_number
FROM dbo.acc_gls_billing_line WITH (NOLOCK)
WHERE {' AND '.join(where)}
ORDER BY parcel_number
        """,
        params,
    )
    return {_normalize_identifier(row[0]) for row in cur.fetchall() if _normalize_identifier(row[0])}


def _resolve_seed_parcel_numbers(
    cur,
    *,
    changed_parcels: set[str],
    seed_all_existing: bool,
) -> set[str]:
    if seed_all_existing:
        return _load_all_billing_parcel_numbers(cur)
    return {_normalize_identifier(value) for value in changed_parcels if _normalize_identifier(value)}


def _seed_shipments_from_billing(
    conn,
    cur,
    *,
    parcel_numbers: set[str],
    commit_every: int = 500,
) -> dict[str, int]:
    # Billing import only seeds shipment identities and order links from invoice
    # parcels. Actual acc_shipment_cost rows are materialized later by
    # gls_sync_costs so late invoices can backfill costs without reseeding.
    parcel_list = sorted({_normalize_identifier(value) for value in parcel_numbers if _normalize_identifier(value)})
    if not parcel_list:
        return {
            "shipments_seeded": 0,
            "links_written": 0,
            "shipments_linked": 0,
            "shipments_unlinked": 0,
        }

    log.info("gls_seed.stage.start", parcel_numbers=len(parcel_list))
    seeds = _load_gls_shipment_seeds(cur, parcel_numbers=parcel_list)
    log.info("gls_seed.stage.seeds_loaded", seeds=len(seeds))
    tracking_values = sorted({_normalize_identifier(seed.parcel_number) for seed in seeds if seed.parcel_number})
    note1_values = sorted({_normalize_identifier(seed.note1) for seed in seeds if seed.note1})
    log.info(
        "gls_seed.stage.lookup_values",
        tracking_values=len(tracking_values),
        note1_values=len(note1_values),
    )
    by_tracking: dict[str, list[LocalPackageCandidate]] = {}
    log.info("gls_seed.stage.collect_tracking.start", tracking_values=len(tracking_values))
    _collect_tracking_candidates(cur, tracking_values=tracking_values, result_bucket=by_tracking)
    log.info("gls_seed.stage.collect_tracking.done", matched_tracking=len(by_tracking))
    log.info("gls_seed.stage.bl_map.start", tracking_values=len(tracking_values))
    by_tracking_bl_map = _load_bl_map_tracking_candidates(cur, tracking_values=tracking_values)
    log.info("gls_seed.stage.bl_map.done", matched_tracking=len(by_tracking_bl_map))
    log.info("gls_seed.stage.bl_order.start", note1_values=len(note1_values))
    by_bl_order = _load_bl_order_candidates(cur, bl_order_values=note1_values)
    log.info("gls_seed.stage.bl_order.done", matched_bl_orders=len(by_bl_order))

    stats = {
        "shipments_seeded": 0,
        "links_written": 0,
        "shipments_linked": 0,
        "shipments_unlinked": 0,
    }
    safe_commit_every = max(1, int(commit_every or 500))
    for idx, seed in enumerate(seeds, start=1):
        delivered = _is_delivered_status(seed.parcel_status, seed.delivery_date)
        payload_json = {
            "source": "gls_billing_files",
            "parcel_number": seed.parcel_number,
            "note1": seed.note1,
            "billing_period": seed.billing_period,
            "line_count": seed.line_count,
            "total_amount": seed.total_amount,
        }
        shipment_payload = {
            "carrier": "GLS",
            "carrier_account": None,
            "shipment_number": seed.parcel_number,
            "piece_id": seed.parcel_number,
            "tracking_number": seed.parcel_number,
            "cedex_number": None,
            "service_code": seed.service_code,
            "ship_date": seed.row_date,
            "created_at_carrier": _to_datetime(seed.row_date),
            "status_code": "DELIVERED" if delivered else "BILLING_IMPORTED",
            "status_label": seed.parcel_status or ("Delivered from billing files" if delivered else "Imported from billing files"),
            "received_by": None,
            "is_delivered": delivered,
            "delivered_at": _to_datetime(seed.delivery_date),
            "recipient_name": seed.recipient_name,
            "recipient_country": seed.recipient_country,
            "shipper_name": None,
            "shipper_country": None,
            "source_system": "gls_billing_files",
            "source_payload_json": _json_dump(payload_json),
            "source_payload_hash": _hash_payload(payload_json),
        }
        shipment_id = _upsert_shipment(cur, shipment_payload)
        stats["shipments_seeded"] += 1
        links = _collect_seed_links(
            seed=seed,
            by_tracking=by_tracking,
            by_tracking_bl_map=by_tracking_bl_map,
            by_bl_order=by_bl_order,
        )
        if links:
            stats["links_written"] += _upsert_shipment_links(cur, shipment_id=shipment_id, links=links)
            stats["shipments_linked"] += 1
        else:
            stats["shipments_unlinked"] += 1
        if idx % safe_commit_every == 0:
            conn.commit()
            log.info(
                "gls_seed.stage.progress",
                processed=idx,
                total=len(seeds),
                linked=stats["shipments_linked"],
                unlinked=stats["shipments_unlinked"],
            )
    conn.commit()
    log.info(
        "gls_seed.stage.done",
        processed=len(seeds),
        linked=stats["shipments_linked"],
        unlinked=stats["shipments_unlinked"],
        links_written=stats["links_written"],
    )
    return stats


def import_gls_billing_files(
    *,
    invoice_root: str | None = None,
    bl_map_path: str | None = None,
    include_shipment_seed: bool = True,
    seed_all_existing: bool = False,
    force_reimport: bool = False,
    limit_invoice_files: int | None = None,
    job_id: str | None = None,
) -> dict[str, Any]:
    ensure_gls_schema()
    ensure_dhl_schema()

    from app.connectors.mssql.mssql_store import set_job_progress

    invoice_root_path = Path(invoice_root or settings.GLS_BILLING_ROOT_PATH)
    bl_map_file = Path(bl_map_path or settings.GLS_BILLING_BL_MAP_PATH)
    invoice_files = _discover_gls_csv_files(invoice_root_path)
    correction_files = _discover_gls_correction_xlsx_files(invoice_root_path)
    if limit_invoice_files is not None:
        invoice_files = invoice_files[:limit_invoice_files]
        correction_files = correction_files[:limit_invoice_files]

    total_steps = len(invoice_files) + len(correction_files) + (1 if bl_map_file.exists() else 0)
    total_steps = max(total_steps, 1)
    processed_steps = 0

    stats: dict[str, Any] = {
        "invoice_files_discovered": len(invoice_files),
        "invoice_files_imported": 0,
        "invoice_files_skipped": 0,
        "invoice_line_rows_imported": 0,
        "billing_documents_upserted": 0,
        "correction_files_discovered": len(correction_files),
        "correction_files_imported": 0,
        "correction_files_skipped": 0,
        "correction_files_unsupported": 0,
        "correction_rows_imported": 0,
        "bl_map_imported": 0,
        "bl_map_rows_imported": 0,
        "shipments_seeded": 0,
        "links_written": 0,
        "shipments_linked": 0,
        "shipments_unlinked": 0,
        "seed_scope": "all_existing" if seed_all_existing else "changed_only",
        "seed_parcel_numbers": 0,
        "errors": 0,
    }
    changed_parcels: set[str] = set()

    conn = _connect()
    try:
        cur = conn.cursor()

        if bl_map_file.exists():
            file_path = str(bl_map_file.resolve())
            file_size = bl_map_file.stat().st_size
            file_mtime = _file_mtime_utc(bl_map_file)
            state = _load_import_file_state(cur, source_kind="bl_map", file_path=file_path)
            if _should_skip_import_file(
                state,
                file_size=file_size,
                file_mtime=file_mtime,
                force_reimport=force_reimport,
            ):
                processed_steps += 1
            else:
                try:
                    rows = parse_gls_bl_map_xlsx(file_path)
                    written_rows = _replace_gls_bl_map_rows(cur, source_file=file_path, rows=rows)
                    _upsert_import_file(
                        cur,
                        source_kind="bl_map",
                        file_path=file_path,
                        file_name=bl_map_file.name,
                        document_number=None,
                        file_size_bytes=file_size,
                        file_mtime_utc=file_mtime,
                        status="imported",
                        rows_imported=written_rows,
                        error_message=None,
                    )
                    conn.commit()
                    stats["bl_map_imported"] = 1
                    stats["bl_map_rows_imported"] = written_rows
                except Exception as exc:
                    conn.rollback()
                    _upsert_import_file(
                        cur,
                        source_kind="bl_map",
                        file_path=file_path,
                        file_name=bl_map_file.name,
                        document_number=None,
                        file_size_bytes=file_size,
                        file_mtime_utc=file_mtime,
                        status="failed",
                        rows_imported=0,
                        error_message=str(exc),
                    )
                    conn.commit()
                    stats["errors"] += 1
                    log.exception("gls_billing_import.bl_map_failed", path=file_path)
                processed_steps += 1
                if job_id:
                    pct = 10 + int((processed_steps / total_steps) * 80)
                    set_job_progress(
                        job_id,
                        progress_pct=min(pct, 90),
                        records_processed=stats["bl_map_rows_imported"],
                        message=f"GLS billing import files {processed_steps}/{total_steps}",
                    )

        for path in invoice_files:
            file_path = str(path)
            file_size = path.stat().st_size
            file_mtime = _file_mtime_utc(path)
            document_number = _billing_document_number_from_csv(path)
            state = _load_import_file_state(cur, source_kind="invoice", file_path=file_path)
            if _should_skip_import_file(
                state,
                file_size=file_size,
                file_mtime=file_mtime,
                force_reimport=force_reimport,
            ):
                stats["invoice_files_skipped"] += 1
                processed_steps += 1
                continue
            try:
                rows = parse_gls_billing_csv(file_path)
                for row in rows:
                    if row.get("parcel_number"):
                        changed_parcels.add(str(row["parcel_number"]))
                written_rows = _replace_gls_billing_lines(cur, source_file=file_path, rows=rows)
                billing_period = rows[0].get("billing_period") if rows else _billing_period_from_path(path)
                _upsert_billing_document(
                    cur,
                    document_number=_normalize_identifier(document_number),
                    billing_period=billing_period,
                    source_file=file_path,
                    detail_rows_count=written_rows,
                )
                _upsert_import_file(
                    cur,
                    source_kind="invoice",
                    file_path=file_path,
                    file_name=path.name,
                    document_number=document_number,
                    file_size_bytes=file_size,
                    file_mtime_utc=file_mtime,
                    status="imported",
                    rows_imported=written_rows,
                    error_message=None,
                )
                conn.commit()
                stats["invoice_files_imported"] += 1
                stats["invoice_line_rows_imported"] += written_rows
                stats["billing_documents_upserted"] += 1
            except Exception as exc:
                conn.rollback()
                _upsert_import_file(
                    cur,
                    source_kind="invoice",
                    file_path=file_path,
                    file_name=path.name,
                    document_number=document_number,
                    file_size_bytes=file_size,
                    file_mtime_utc=file_mtime,
                    status="failed",
                    rows_imported=0,
                    error_message=str(exc),
                )
                conn.commit()
                stats["errors"] += 1
                log.exception("gls_billing_import.invoice_failed", path=file_path)
            processed_steps += 1
            if job_id:
                pct = 10 + int((processed_steps / total_steps) * 80)
                set_job_progress(
                    job_id,
                    progress_pct=min(pct, 95),
                    records_processed=stats["invoice_line_rows_imported"] + stats["bl_map_rows_imported"],
                    message=f"GLS billing import files {processed_steps}/{total_steps}",
                )

        for path in correction_files:
            file_path = str(path)
            file_size = path.stat().st_size
            file_mtime = _file_mtime_utc(path)
            state = _load_import_file_state(cur, source_kind="invoice_correction", file_path=file_path)
            if _should_skip_import_file(
                state,
                file_size=file_size,
                file_mtime=file_mtime,
                force_reimport=force_reimport,
            ):
                stats["correction_files_skipped"] += 1
                processed_steps += 1
                continue
            try:
                rows = parse_gls_billing_correction_xlsx(file_path)
                if not rows:
                    raise ValueError("unsupported_correction_layout")
                for row in rows:
                    if row.get("parcel_number"):
                        changed_parcels.add(str(row["parcel_number"]))
                written_rows = _replace_gls_billing_correction_lines(cur, source_file=file_path, rows=rows)
                _upsert_import_file(
                    cur,
                    source_kind="invoice_correction",
                    file_path=file_path,
                    file_name=path.name,
                    document_number=_normalize_identifier(rows[0].get("document_number")) if rows else None,
                    file_size_bytes=file_size,
                    file_mtime_utc=file_mtime,
                    status="imported",
                    rows_imported=written_rows,
                    error_message=None,
                )
                conn.commit()
                stats["correction_files_imported"] += 1
                stats["correction_rows_imported"] += written_rows
            except ValueError as exc:
                conn.rollback()
                _upsert_import_file(
                    cur,
                    source_kind="invoice_correction",
                    file_path=file_path,
                    file_name=path.name,
                    document_number=None,
                    file_size_bytes=file_size,
                    file_mtime_utc=file_mtime,
                    status="imported",
                    rows_imported=0,
                    error_message=str(exc),
                )
                conn.commit()
                stats["correction_files_unsupported"] += 1
                log.warning("gls_billing_import.correction_unsupported", path=file_path)
            except Exception as exc:
                conn.rollback()
                _upsert_import_file(
                    cur,
                    source_kind="invoice_correction",
                    file_path=file_path,
                    file_name=path.name,
                    document_number=None,
                    file_size_bytes=file_size,
                    file_mtime_utc=file_mtime,
                    status="failed",
                    rows_imported=0,
                    error_message=str(exc),
                )
                conn.commit()
                stats["errors"] += 1
                log.exception("gls_billing_import.correction_failed", path=file_path)
            processed_steps += 1
            if job_id:
                pct = 10 + int((processed_steps / total_steps) * 80)
                set_job_progress(
                    job_id,
                    progress_pct=min(pct, 95),
                    records_processed=(
                        stats["invoice_line_rows_imported"]
                        + stats["correction_rows_imported"]
                        + stats["bl_map_rows_imported"]
                    ),
                    message=f"GLS billing import files {processed_steps}/{total_steps}",
                )

        seed_parcels = _resolve_seed_parcel_numbers(
            cur,
            changed_parcels=changed_parcels,
            seed_all_existing=seed_all_existing,
        )
        stats["seed_parcel_numbers"] = len(seed_parcels)
        if include_shipment_seed and seed_parcels:
            if job_id:
                set_job_progress(
                    job_id,
                    progress_pct=95,
                    records_processed=(
                        stats["invoice_line_rows_imported"]
                        + stats["correction_rows_imported"]
                        + stats["bl_map_rows_imported"]
                    ),
                    message="GLS billing import seeding shipments",
                )
            seed_stats = _seed_shipments_from_billing(conn, cur, parcel_numbers=seed_parcels)
            conn.commit()
            stats.update(seed_stats)

        if job_id:
            set_job_progress(
                job_id,
                progress_pct=98,
                records_processed=(
                    stats["invoice_line_rows_imported"]
                    + stats["correction_rows_imported"]
                    + stats["bl_map_rows_imported"]
                ),
                message="GLS billing import finished",
            )
        return stats
    finally:
        conn.close()


def seed_gls_shipments_from_staging(
    *,
    created_from: date | None = None,
    created_to: date | None = None,
    seed_all_existing: bool = True,
    limit_parcels: int | None = None,
    job_id: str | None = None,
) -> dict[str, Any]:
    ensure_gls_schema()
    ensure_dhl_schema()

    from app.connectors.mssql.mssql_store import set_job_progress

    stats: dict[str, Any] = {
        "seed_scope": "all_existing" if seed_all_existing else "window",
        "seed_parcel_numbers": 0,
        "shipments_seeded": 0,
        "links_written": 0,
        "shipments_linked": 0,
        "shipments_unlinked": 0,
    }

    conn = _connect()
    try:
        cur = conn.cursor()
        if seed_all_existing:
            parcel_numbers = _load_all_billing_parcel_numbers(cur)
            if limit_parcels:
                parcel_numbers = set(sorted(parcel_numbers)[: int(limit_parcels)])
        else:
            parcel_numbers = _load_windowed_billing_parcel_numbers(
                cur,
                created_from=created_from,
                created_to=created_to,
                limit_parcels=limit_parcels,
            )
        stats["seed_parcel_numbers"] = len(parcel_numbers)
        if not parcel_numbers:
            return stats

        if job_id:
            set_job_progress(
                job_id,
                progress_pct=20,
                records_processed=0,
                message=f"GLS seed shipments scope={len(parcel_numbers)}",
            )

        seed_stats = _seed_shipments_from_billing(conn, cur, parcel_numbers=parcel_numbers)
        conn.commit()
        stats.update(seed_stats)

        if job_id:
            set_job_progress(
                job_id,
                progress_pct=95,
                records_processed=int(stats.get("shipments_seeded", 0) or 0),
                message="GLS seed shipments finished",
            )
        return stats
    finally:
        conn.close()

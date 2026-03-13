from __future__ import annotations

import hashlib
import json
import os
import re
import time
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

import openpyxl
import structlog

from app.core.config import settings
from app.core.db_connection import connect_acc, connect_netfox
from app.services.bl_order_lookup import resolve_bl_orders_to_acc_orders
from app.services.dhl_integration import ensure_dhl_schema
from app.services.dhl_registry_sync import _upsert_shipment, _upsert_shipment_links

log = structlog.get_logger(__name__)

_BATCH_SIZE = 300
_JJ_INSERT_BATCH_SIZE = 100
_LOCK_RETRY_ATTEMPTS = 5
_LOCK_RETRY_BASE_SLEEP_SEC = 1.0
_INVOICE_FILENAME_RE = re.compile(r"DHL_Dokument nr (?P<document_number>\d+)\.xlsx$", re.IGNORECASE)
_DHL_NETFOX_JJD_LOOKUP_ENABLED = str(os.getenv("DHL_NETFOX_JJD_LOOKUP_ENABLED", "false") or "").strip().lower() in {
    "1",
    "true",
    "yes",
    "on",
}
_DHL_NETFOX_JJD_LOOKUP_MAX = max(0, int(str(os.getenv("DHL_NETFOX_JJD_LOOKUP_MAX", "2000") or "2000").strip() or "2000"))


@dataclass
class NormalizedParcel:
    raw: str | None
    base: str | None
    suffix: str | None


@dataclass
class LocalPackageCandidate:
    amazon_order_id: str
    acc_order_id: str | None
    bl_order_id: int | None
    package_order_id: int | None
    courier_package_nr: str | None
    courier_inner_number: str | None


@dataclass
class ImportedShipmentSeed:
    parcel_number: str
    parcel_number_base: str
    jjd_number: str | None
    shipment_type: str | None
    ship_date: datetime | None
    delivery_date: datetime | None
    last_event_code: str | None
    last_event_at: datetime | None
    product_code: str | None
    description: str | None
    issue_date: date | None
    sales_date: date | None
    total_net_amount: float
    line_count: int
    sap_order: str | None = None
    shipper_receiver: str | None = None


def _connect():
    return connect_acc(autocommit=False, timeout=60)


def _normalize_text(value: Any) -> str:
    return str(value or "").strip()


def _normalize_token(value: Any) -> str:
    return _normalize_text(value).upper().replace(" ", "")


def _normalize_document_number(value: Any) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    text = _normalize_text(value)
    if text.endswith(".0"):
        prefix = text[:-2]
        if prefix.isdigit():
            return prefix
    return text


def _normalize_parcel_number(value: Any) -> NormalizedParcel:
    raw = _normalize_token(value)
    if not raw:
        return NormalizedParcel(raw=None, base=None, suffix=None)
    base, suffix = raw, None
    if "/" in raw:
        base, suffix = raw.split("/", 1)
    return NormalizedParcel(raw=raw, base=base or None, suffix=suffix or None)


def _parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _normalize_text(value)
    if not text or text.lower() == "n/a":
        return None
    try:
        return datetime.fromisoformat(text[:10]).date()
    except ValueError:
        return None


def _parse_datetime(value: Any) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    text = _normalize_text(value)
    if not text:
        return None
    for candidate in (text, text[:19], text[:10]):
        try:
            parsed = datetime.fromisoformat(candidate)
            if isinstance(parsed, datetime):
                return parsed
        except ValueError:
            continue
    return None


def _parse_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except Exception:
        text = _normalize_text(value).replace(",", ".")
        try:
            return float(text)
        except Exception:
            return None


def _json_dump(payload: Any) -> str:
    return json.dumps(payload, ensure_ascii=False, sort_keys=True)


def _hash_payload(payload: Any) -> str:
    return hashlib.sha256(_json_dump(payload).encode("utf-8")).hexdigest()


def _chunks(items: list[Any], size: int) -> Iterable[list[Any]]:
    for idx in range(0, len(items), size):
        yield items[idx : idx + size]


def _is_lock_timeout_error(exc: Exception) -> bool:
    text = str(exc or "")
    lowered = text.lower()
    return "1222" in lowered or "lock request time out period exceeded" in lowered


def _run_with_lock_retry(
    conn,
    operation,
    *,
    op_name: str,
    source_file: str,
    extra: dict[str, Any] | None = None,
):
    details = extra or {}
    for attempt in range(1, _LOCK_RETRY_ATTEMPTS + 1):
        try:
            return operation()
        except Exception as exc:
            conn.rollback()
            if not _is_lock_timeout_error(exc) or attempt >= _LOCK_RETRY_ATTEMPTS:
                raise
            sleep_sec = round(_LOCK_RETRY_BASE_SLEEP_SEC * (2 ** (attempt - 1)), 2)
            log.warning(
                "dhl_billing_import.lock_retry",
                op_name=op_name,
                source_file=source_file,
                attempt=attempt,
                sleep_sec=sleep_sec,
                error=str(exc),
                **details,
            )
            time.sleep(sleep_sec)


def _file_mtime_utc(path: Path) -> datetime:
    return datetime.utcfromtimestamp(path.stat().st_mtime).replace(microsecond=0)


def _invoice_file_document_number(path: Path) -> str | None:
    match = _INVOICE_FILENAME_RE.search(path.name)
    return match.group("document_number") if match else None


def _discover_invoice_files(root: Path) -> list[Path]:
    if not root.exists():
        return []
    files = sorted(root.glob("DHL_Dokument nr *.xlsx"))
    for child in sorted(root.iterdir()):
        if child.is_dir() and child.name != "JJ":
            files.extend(sorted(child.glob("DHL_Dokument nr *.xlsx")))
    return sorted({path.resolve() for path in files})


def _discover_jj_files(root: Path) -> list[Path]:
    if not root.exists():
        return []
    return sorted(path.resolve() for path in root.glob("*.xlsx"))


def _header_index(header_row: tuple[Any, ...]) -> dict[str, int]:
    result: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        name = _normalize_text(cell)
        if name:
            result[name] = idx
    return result


def _cell(row: tuple[Any, ...], header_map: dict[str, int], *names: str) -> Any:
    for name in names:
        idx = header_map.get(name)
        if idx is not None and idx < len(row):
            return row[idx]
    return None


def parse_dhl_manifest_xlsx(path: str | os.PathLike[str]) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(min_row=1, values_only=True)
        header_row = next(rows, None)
        if header_row is None:
            return []
        header_map = _header_index(tuple(header_row))
        parsed: list[dict[str, Any]] = []
        for row in rows:
            document_number = _normalize_document_number(_cell(row, header_map, "Numer dokumentu"))
            if not document_number:
                continue
            parsed.append(
                {
                    "document_number": document_number,
                    "document_type": _normalize_text(_cell(row, header_map, "Typ dokumentu")) or None,
                    "issue_date": _parse_date(_cell(row, header_map, "Data wystawienia")),
                    "ship_date": _parse_date(_cell(row, header_map, "Data wysyłki")),
                    "due_date": _parse_date(_cell(row, header_map, "Termin płatności")),
                    "net_amount": _parse_float(_cell(row, header_map, "Kwota netto")),
                    "vat_amount": _parse_float(_cell(row, header_map, "Kwota VAT")),
                    "gross_amount": _parse_float(_cell(row, header_map, "Kwota brutto")),
                }
            )
        return parsed
    finally:
        wb.close()


def parse_dhl_invoice_xlsx(path: str | os.PathLike[str]) -> list[dict[str, Any]]:
    # DHL billing workbooks lose most columns in openpyxl read_only mode.
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=True)
    try:
        ws = workbook[workbook.sheetnames[0]]
        rows = ws.iter_rows(min_row=1, values_only=True)
        header_row = next(rows, None)
        if header_row is None:
            return []
        header_map = _header_index(tuple(header_row))
        parsed: list[dict[str, Any]] = []
        file_doc = _invoice_file_document_number(Path(path))
        for row_no, row in enumerate(rows, start=2):
            document_number = _normalize_document_number(
                _cell(row, header_map, "Numer załącznika do dokumentu", "Nr załącznika do dokumentu")
            ) or _normalize_document_number(file_doc)
            parcel = _normalize_parcel_number(_cell(row, header_map, "Numer przesyłki"))
            if not document_number or not parcel.base:
                continue
            parsed.append(
                {
                    "document_number": document_number,
                    "issue_date": _parse_date(_cell(row, header_map, "Data wystawienia")),
                    "sales_date": _parse_date(_cell(row, header_map, "Data sprzedaży")),
                    "due_date": _parse_date(_cell(row, header_map, "Termin płatności")),
                    "parcel_number": parcel.raw,
                    "parcel_number_base": parcel.base,
                    "parcel_number_suffix": parcel.suffix,
                    "delivery_date": _parse_date(_cell(row, header_map, "Data doręczenia")),
                    "quantity": _parse_float(_cell(row, header_map, "Ilość")),
                    "product_code": _normalize_text(_cell(row, header_map, "Kod produktu")) or None,
                    "description": _normalize_text(_cell(row, header_map, "Opis")) or None,
                    "weight": _parse_float(_cell(row, header_map, "Waga")),
                    "weight_kind": _normalize_text(_cell(row, header_map, "Rodzaj wagi")) or None,
                    "shipper_receiver": _normalize_text(_cell(row, header_map, "Nadawca/Odbiorca")) or None,
                    "payer_type": _normalize_text(_cell(row, header_map, "Kto płaci?")) or None,
                    "sap_order": _normalize_text(_cell(row, header_map, "SAP zlec.", "SAP zlec")) or None,
                    "mpk": _normalize_text(_cell(row, header_map, "MPK")) or None,
                    "pkwiu": _normalize_text(_cell(row, header_map, "PKWiU")) or None,
                    "notes": _normalize_text(_cell(row, header_map, "Uwagi")) or None,
                    "net_amount": _parse_float(_cell(row, header_map, "Razem netto")),
                    "base_fee": _parse_float(_cell(row, header_map, "Opłata podstawowa")),
                    "base_discount": _parse_float(_cell(row, header_map, "Rabat od opłaty podstawowej/zaokrąglenia")),
                    "non_standard_fee": _parse_float(_cell(row, header_map, "Element niestandardowy")),
                    "seasonal_fee": _parse_float(_cell(row, header_map, "Opłata sezonowa")),
                    "fuel_road_fee": _parse_float(_cell(row, header_map, "Opłata paliwowa i drogowa")),
                    "insurance_fee": _parse_float(_cell(row, header_map, "Przystąpienie do ubezp. DHL")),
                    "cod_fee": _parse_float(_cell(row, header_map, "Pobranie")),
                    "label_fee": _parse_float(_cell(row, header_map, "Dopłata za wydruk etykiety")),
                    "volumetric_fee": _parse_float(
                        _cell(row, header_map, "Dopłata za wagę wolum. pow. 31,5 kg")
                    ),
                    "source_row_no": row_no,
                }
            )
        return parsed
    finally:
        workbook.close()


def parse_dhl_jj_xlsx(path: str | os.PathLike[str]) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(min_row=1, values_only=True)
        header_row = next(rows, None)
        if header_row is None:
            return []
        header_map = _header_index(tuple(header_row))
        parsed: list[dict[str, Any]] = []
        for row_no, row in enumerate(rows, start=2):
            parcel = _normalize_parcel_number(_cell(row, header_map, "NUMER_PRZESYLKI"))
            jjd_number = _normalize_token(_cell(row, header_map, "PACZKA_NUMER_DOWOLNY_2"))
            if not parcel.base or not jjd_number.startswith("JJD"):
                continue
            parsed.append(
                {
                    "parcel_number": parcel.raw,
                    "parcel_number_base": parcel.base,
                    "parcel_number_suffix": parcel.suffix,
                    "jjd_number": jjd_number,
                    "shipment_type": _normalize_text(_cell(row, header_map, "RODZAJ_PRZESYLKI")) or None,
                    "ship_date": _parse_datetime(_cell(row, header_map, "NADANIE_DATACZAS_BORSOFT")),
                    "delivery_date": _parse_datetime(_cell(row, header_map, "DORECZENIE_DATACZAS_BORSOFT")),
                    "last_event_code": _normalize_text(_cell(row, header_map, "OSTAT_ZDARZ_NAZWA")) or None,
                    "last_event_at": _parse_datetime(_cell(row, header_map, "OSTAT_ZDARZ_DATACZAS")),
                    "source_row_no": row_no,
                }
            )
        return parsed
    finally:
        wb.close()


def _is_dhl_package(courier_code: Any, courier_other_name: Any) -> bool:
    code = _normalize_text(courier_code).lower()
    other = _normalize_text(courier_other_name).lower()
    if code == "blconnectpackages":
        code = other
    return "dhl" in code or "dhl" in other


def _upsert_billing_document(
    cur,
    *,
    document_number: str,
    document_type: str | None,
    issue_date: date | None,
    ship_date: date | None,
    due_date: date | None,
    net_amount: float | None,
    vat_amount: float | None,
    gross_amount: float | None,
    source_file: str | None,
    source_manifest_file: str | None,
    detail_rows_count: int | None,
) -> None:
    cur.execute(
        """
        SELECT document_number
        FROM dbo.acc_dhl_billing_document WITH (NOLOCK)
        WHERE document_number = ?
        """,
        [document_number],
    )
    exists = cur.fetchone() is not None
    if exists:
        cur.execute(
            """
            UPDATE dbo.acc_dhl_billing_document
            SET document_type = COALESCE(?, document_type),
                issue_date = COALESCE(?, issue_date),
                ship_date = COALESCE(?, ship_date),
                due_date = COALESCE(?, due_date),
                net_amount = COALESCE(?, net_amount),
                vat_amount = COALESCE(?, vat_amount),
                gross_amount = COALESCE(?, gross_amount),
                source_file = COALESCE(?, source_file),
                source_manifest_file = COALESCE(?, source_manifest_file),
                detail_rows_count = CASE WHEN ? IS NULL OR ? < 1 THEN detail_rows_count ELSE ? END,
                last_imported_at = SYSUTCDATETIME(),
                updated_at = SYSUTCDATETIME()
            WHERE document_number = ?
            """,
            [
                document_type,
                issue_date,
                ship_date,
                due_date,
                net_amount,
                vat_amount,
                gross_amount,
                source_file,
                source_manifest_file,
                detail_rows_count,
                detail_rows_count,
                detail_rows_count,
                document_number,
            ],
        )
        return

    cur.execute(
        """
        INSERT INTO dbo.acc_dhl_billing_document (
            document_number, document_type, issue_date, ship_date, due_date, net_amount,
            vat_amount, gross_amount, currency, source_file, source_manifest_file,
            detail_rows_count, last_imported_at, updated_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'PLN', ?, ?, ?, SYSUTCDATETIME(), SYSUTCDATETIME())
        """,
        [
            document_number,
            document_type,
            issue_date,
            ship_date,
            due_date,
            net_amount,
            vat_amount,
            gross_amount,
            source_file,
            source_manifest_file,
            int(detail_rows_count or 0),
        ],
    )


def _replace_billing_lines(cur, *, document_number: str, source_file: str, rows: list[dict[str, Any]]) -> int:
    cur.execute("DELETE FROM dbo.acc_dhl_billing_line WHERE document_number = ?", [document_number])
    if not rows:
        return 0
    insert_sql = """
        INSERT INTO dbo.acc_dhl_billing_line (
            document_number, issue_date, sales_date, due_date, parcel_number, parcel_number_base,
            parcel_number_suffix, delivery_date, quantity, product_code, description, weight,
            weight_kind, shipper_receiver, payer_type, sap_order, mpk, pkwiu, notes,
            net_amount, base_fee, base_discount, non_standard_fee, seasonal_fee, fuel_road_fee,
            insurance_fee, cod_fee, label_fee, volumetric_fee, source_file, source_row_no,
            source_hash, imported_at
        )
        VALUES (
            ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
            SYSUTCDATETIME()
        )
    """
    payloads: list[list[Any]] = []
    for row in rows:
        source_hash = _normalize_token(
            f"{document_number}|{row['source_row_no']}|{row.get('parcel_number')}|{row.get('net_amount')}"
        )
        payloads.append(
            [
                document_number,
                row.get("issue_date"),
                row.get("sales_date"),
                row.get("due_date"),
                row.get("parcel_number"),
                row.get("parcel_number_base"),
                row.get("parcel_number_suffix"),
                row.get("delivery_date"),
                row.get("quantity"),
                row.get("product_code"),
                row.get("description"),
                row.get("weight"),
                row.get("weight_kind"),
                row.get("shipper_receiver"),
                row.get("payer_type"),
                row.get("sap_order"),
                row.get("mpk"),
                row.get("pkwiu"),
                row.get("notes"),
                row.get("net_amount"),
                row.get("base_fee"),
                row.get("base_discount"),
                row.get("non_standard_fee"),
                row.get("seasonal_fee"),
                row.get("fuel_road_fee"),
                row.get("insurance_fee"),
                row.get("cod_fee"),
                row.get("label_fee"),
                row.get("volumetric_fee"),
                source_file,
                row.get("source_row_no"),
                source_hash,
            ]
        )
    cur.executemany(insert_sql, payloads)
    return len(payloads)


def _replace_parcel_map_rows(cur, *, source_file: str, rows: list[dict[str, Any]]) -> int:
    cur.execute("DELETE FROM dbo.acc_dhl_parcel_map WHERE source_file = ?", [source_file])
    if not rows:
        return 0
    payloads = _build_parcel_map_payloads(source_file=source_file, rows=rows)
    _insert_parcel_map_payloads(cur, payloads)
    return len(payloads)


def _build_parcel_map_payloads(*, source_file: str, rows: list[dict[str, Any]]) -> list[list[Any]]:
    if not rows:
        return []
    payloads: list[list[Any]] = []
    for row in rows:
        source_hash = _normalize_token(
            f"{row.get('jjd_number')}|{row.get('parcel_number_base')}|{row.get('source_row_no')}"
        )
        payloads.append(
            [
                row.get("parcel_number"),
                row.get("parcel_number_base"),
                row.get("parcel_number_suffix"),
                row.get("jjd_number"),
                row.get("shipment_type"),
                row.get("ship_date"),
                row.get("delivery_date"),
                row.get("last_event_code"),
                row.get("last_event_at"),
                source_file,
                row.get("source_row_no"),
                source_hash,
            ]
        )
    return payloads


def _insert_parcel_map_payloads(cur, payloads: list[list[Any]]) -> int:
    if not payloads:
        return 0
    insert_sql = """
        MERGE dbo.acc_dhl_parcel_map AS target
        USING (
            SELECT
                ? AS parcel_number,
                ? AS parcel_number_base,
                ? AS parcel_number_suffix,
                ? AS jjd_number,
                ? AS shipment_type,
                ? AS ship_date,
                ? AS delivery_date,
                ? AS last_event_code,
                ? AS last_event_at,
                ? AS source_file,
                ? AS source_row_no,
                ? AS source_hash
        ) AS src
          ON target.source_file = src.source_file
         AND target.source_row_no = src.source_row_no
        WHEN MATCHED THEN
            UPDATE SET
                parcel_number = src.parcel_number,
                parcel_number_base = src.parcel_number_base,
                parcel_number_suffix = src.parcel_number_suffix,
                jjd_number = src.jjd_number,
                shipment_type = src.shipment_type,
                ship_date = src.ship_date,
                delivery_date = src.delivery_date,
                last_event_code = src.last_event_code,
                last_event_at = src.last_event_at,
                source_hash = src.source_hash,
                imported_at = SYSUTCDATETIME()
        WHEN NOT MATCHED THEN
            INSERT (
                parcel_number, parcel_number_base, parcel_number_suffix, jjd_number, shipment_type,
                ship_date, delivery_date, last_event_code, last_event_at, source_file, source_row_no,
                source_hash, imported_at
            )
            VALUES (
                src.parcel_number, src.parcel_number_base, src.parcel_number_suffix, src.jjd_number, src.shipment_type,
                src.ship_date, src.delivery_date, src.last_event_code, src.last_event_at, src.source_file, src.source_row_no,
                src.source_hash, SYSUTCDATETIME()
            );
    """
    cur.executemany(insert_sql, payloads)
    return len(payloads)


def _load_import_file_state(
    cur, *, source_kind: str, file_path: str
) -> tuple[int | None, datetime | None, str | None] | None:
    cur.execute(
        """
        SELECT file_size_bytes, file_mtime_utc, status
        FROM dbo.acc_dhl_import_file WITH (NOLOCK)
        WHERE source_kind = ? AND file_path = ?
        """,
        [source_kind, file_path],
    )
    row = cur.fetchone()
    if not row:
        return None
    size = int(row[0]) if row[0] is not None else None
    mtime = row[1]
    if isinstance(mtime, datetime):
        mtime = mtime.replace(microsecond=0)
    status = str(row[2] or "").strip().lower() or None
    return size, mtime, status


def _should_skip_import_file(
    state: tuple[int | None, datetime | None, str | None] | None,
    *,
    file_size: int,
    file_mtime: datetime,
    force_reimport: bool,
) -> bool:
    if force_reimport or state is None:
        return False
    prev_size, prev_mtime, prev_status = state
    return prev_status == "imported" and prev_size == file_size and prev_mtime == file_mtime


def _upsert_import_file(
    cur,
    *,
    source_kind: str,
    file_path: str,
    file_name: str,
    document_number: str | None,
    file_size_bytes: int,
    file_mtime_utc: datetime,
    status: str,
    rows_imported: int,
    error_message: str | None,
) -> None:
    cur.execute(
        """
        SELECT CAST(id AS NVARCHAR(40))
        FROM dbo.acc_dhl_import_file WITH (NOLOCK)
        WHERE source_kind = ? AND file_path = ?
        """,
        [source_kind, file_path],
    )
    row = cur.fetchone()
    if row:
        cur.execute(
            """
            UPDATE dbo.acc_dhl_import_file
            SET file_name = ?,
                document_number = ?,
                file_size_bytes = ?,
                file_mtime_utc = ?,
                status = ?,
                rows_imported = ?,
                error_message = ?,
                last_imported_at = SYSUTCDATETIME(),
                updated_at = SYSUTCDATETIME()
            WHERE id = CAST(? AS UNIQUEIDENTIFIER)
            """,
            [
                file_name,
                document_number,
                file_size_bytes,
                file_mtime_utc,
                status,
                rows_imported,
                error_message,
                str(row[0]),
            ],
        )
        return
    cur.execute(
        """
        INSERT INTO dbo.acc_dhl_import_file (
            source_kind, file_path, file_name, document_number, file_size_bytes, file_mtime_utc,
            status, rows_imported, error_message, last_imported_at, updated_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, SYSUTCDATETIME(), SYSUTCDATETIME())
        """,
        [
            source_kind,
            file_path,
            file_name,
            document_number,
            file_size_bytes,
            file_mtime_utc,
            status,
            rows_imported,
            error_message,
        ],
    )


def _collect_package_candidates(
    cur,
    *,
    column_name: str,
    values: list[str],
    result_bucket: dict[str, list[LocalPackageCandidate]],
) -> None:
    if not values:
        return
    include_distribution_orders = column_name == "courier_package_nr"
    for batch in _chunks(values, _BATCH_SIZE):
        placeholders = ",".join("?" for _ in batch)
        distribution_orders_sql = ""
        params: list[Any] = list(batch)
        params.extend(batch)
        if include_distribution_orders:
            distribution_orders_sql = f"""
    UNION ALL
    SELECT
        COALESCE(dm.holding_order_id, o.order_id) AS resolved_bl_order_id,
        CAST(NULL AS BIGINT) AS package_order_id,
        o.delivery_package_nr AS courier_package_nr,
        CAST(NULL AS NVARCHAR(255)) AS courier_inner_number,
        o.delivery_package_module AS courier_code,
        o.delivery_method AS courier_other_name
    FROM dbo.acc_bl_distribution_order_cache o WITH (NOLOCK)
    LEFT JOIN dbo.acc_cache_dis_map dm WITH (NOLOCK)
      ON dm.dis_order_id = o.order_id
    WHERE o.delivery_package_nr IN ({placeholders})
"""
            params.extend(batch)
        cur.execute(
            f"""
WITH direct_packages AS (
    SELECT
        COALESCE(dm.holding_order_id, p.order_id) AS resolved_bl_order_id,
        p.order_id AS package_order_id,
        p.courier_package_nr,
        p.courier_inner_number,
        p.courier_code,
        p.courier_other_name
    FROM dbo.acc_cache_packages p WITH (NOLOCK)
    LEFT JOIN dbo.acc_cache_dis_map dm WITH (NOLOCK)
      ON dm.dis_order_id = p.order_id
    WHERE p.{column_name} IN ({placeholders})
),
distribution_packages AS (
    SELECT
        COALESCE(dm.holding_order_id, p.order_id) AS resolved_bl_order_id,
        CAST(NULL AS BIGINT) AS package_order_id,
        p.courier_package_nr,
        p.courier_inner_number,
        p.courier_code,
        p.courier_other_name
    FROM dbo.acc_bl_distribution_package_cache p WITH (NOLOCK)
    LEFT JOIN dbo.acc_cache_dis_map dm WITH (NOLOCK)
      ON dm.dis_order_id = p.order_id
    WHERE p.{column_name} IN ({placeholders})
)
SELECT DISTINCT
    x.resolved_bl_order_id,
    x.package_order_id,
    x.courier_package_nr,
    x.courier_inner_number,
    x.courier_code,
    x.courier_other_name
FROM (
    SELECT * FROM direct_packages
    UNION ALL
    SELECT * FROM distribution_packages
    {distribution_orders_sql}
) x
            """,
            params,
        )
        rows = cur.fetchall()
        resolved_orders = resolve_bl_orders_to_acc_orders(
            cur,
            bl_order_ids=[
                int(row[0])
                for row in rows
                if row[0] is not None
            ],
        )
        for row in rows:
            if not _is_dhl_package(row[4], row[5]):
                continue
            resolved = resolved_orders.get(int(row[0])) if row[0] is not None else None
            if not resolved:
                continue
            candidate = LocalPackageCandidate(
                amazon_order_id=resolved.amazon_order_id,
                acc_order_id=resolved.acc_order_id,
                bl_order_id=resolved.bl_order_id,
                package_order_id=int(row[1]) if row[1] is not None else None,
                courier_package_nr=_normalize_text(row[2]) or None,
                courier_inner_number=_normalize_text(row[3]) or None,
            )
            if not candidate.amazon_order_id:
                continue
            key_value = row[2] if column_name == "courier_package_nr" else row[3]
            key = _normalize_token(key_value)
            if not key:
                continue
            result_bucket.setdefault(key, [])
            already = {
                (item.amazon_order_id, item.acc_order_id or "", item.bl_order_id or 0, item.package_order_id or 0)
                for item in result_bucket[key]
            }
            marker = (
                candidate.amazon_order_id,
                candidate.acc_order_id or "",
                candidate.bl_order_id or 0,
                candidate.package_order_id or 0,
            )
            if marker not in already:
                result_bucket[key].append(candidate)


def _load_package_lookup(
    cur,
    *,
    tracking_values: set[str],
    inner_values: set[str],
) -> tuple[dict[str, list[LocalPackageCandidate]], dict[str, list[LocalPackageCandidate]]]:
    by_tracking: dict[str, list[LocalPackageCandidate]] = {}
    by_inner: dict[str, list[LocalPackageCandidate]] = {}
    _collect_package_candidates(
        cur,
        column_name="courier_package_nr",
        values=sorted({_normalize_text(value) for value in tracking_values if _normalize_text(value)}),
        result_bucket=by_tracking,
    )
    _collect_package_candidates(
        cur,
        column_name="courier_inner_number",
        values=sorted({_normalize_text(value) for value in inner_values if _normalize_text(value)}),
        result_bucket=by_inner,
    )
    return by_tracking, by_inner


def _load_jjd_netfox_lookup(
    cur,
    *,
    jjd_values: set[str],
) -> dict[str, list[LocalPackageCandidate]]:
    """Resolve JJD tokens via Netfox extras (parcel_num_other -> delivery_package_nr -> order)."""
    result: dict[str, list[LocalPackageCandidate]] = {}
    normalized = sorted({
        _normalize_token(value)
        for value in jjd_values
        if _normalize_token(value).startswith("JJD")
    })
    if not normalized:
        return result
    if not _DHL_NETFOX_JJD_LOOKUP_ENABLED:
        return result
    if _DHL_NETFOX_JJD_LOOKUP_MAX and len(normalized) > _DHL_NETFOX_JJD_LOOKUP_MAX:
        log.warning(
            "dhl_seed.netfox_jjd_lookup.skipped_too_many",
            jjd_values=len(normalized),
            max_values=_DHL_NETFOX_JJD_LOOKUP_MAX,
        )
        return result

    acc_orders: dict[str, tuple[str | None, str | None]] = {}
    netfox_rows: list[tuple[str, int, str | None]] = []
    netfox_conn = None
    try:
        netfox_conn = connect_netfox(timeout=20)
        ncur = netfox_conn.cursor()
        for batch in _chunks(normalized, _BATCH_SIZE):
            placeholders = ",".join("?" for _ in batch)
            ncur.execute(
                f"""
SELECT
    UPPER(LTRIM(RTRIM(e.parcel_num_other))) AS jjd_number,
    z.order_id,
    z.external_order_id
FROM dbo.ITJK_CouriersInvoicesDetails_Extras e WITH (NOLOCK)
JOIN dbo.ITJK_ZamowieniaBaselinkerAPI z WITH (NOLOCK)
  ON UPPER(LTRIM(RTRIM(z.delivery_package_nr))) = UPPER(LTRIM(RTRIM(e.parcel_num)))
WHERE e.parcel_num_other IN ({placeholders})
  AND e.parcel_num_other IS NOT NULL
  AND LTRIM(RTRIM(e.parcel_num_other)) <> ''
  AND z.order_id IS NOT NULL
                """,
                batch,
            )
            for row in ncur.fetchall():
                jjd = _normalize_token(row[0])
                external = _normalize_text(row[2]) or None
                if not jjd or row[1] is None:
                    continue
                netfox_rows.append((jjd, int(row[1]), external))
    except Exception as exc:
        log.warning("dhl_seed.netfox_jjd_lookup.failed", error=str(exc))
        return result
    finally:
        if netfox_conn is not None:
            netfox_conn.close()

    external_values = sorted({
        external
        for _, _, external in netfox_rows
        if external
    })
    for batch in _chunks(external_values, _BATCH_SIZE):
        placeholders = ",".join("?" for _ in batch)
        cur.execute(
            f"""
SELECT amazon_order_id, CAST(id AS NVARCHAR(36)) AS acc_order_id
FROM dbo.acc_order WITH (NOLOCK)
WHERE amazon_order_id IN ({placeholders})
            """,
            batch,
        )
        for row in cur.fetchall():
            amazon_order_id = _normalize_text(row[0])
            if not amazon_order_id:
                continue
            acc_orders[amazon_order_id] = (amazon_order_id, _normalize_text(row[1]) or None)

    dedup: set[tuple[str, str, int]] = set()
    for jjd, bl_order_id, external in netfox_rows:
        if not external:
            continue
        mapped = acc_orders.get(external)
        if not mapped:
            continue
        amazon_order_id, acc_order_id = mapped
        marker = (jjd, amazon_order_id, bl_order_id)
        if marker in dedup:
            continue
        dedup.add(marker)
        result.setdefault(jjd, []).append(
            LocalPackageCandidate(
                amazon_order_id=amazon_order_id,
                acc_order_id=acc_order_id,
                bl_order_id=bl_order_id,
                package_order_id=None,
                courier_package_nr=None,
                courier_inner_number=None,
            )
        )
    log.info(
        "dhl_seed.netfox_jjd_lookup.done",
        jjd_values=len(normalized),
        matched_jjd=len(result),
    )
    return result


def _append_link_candidate(
    bucket: dict[tuple[str, str, str], dict[str, Any]],
    *,
    candidate: LocalPackageCandidate,
    link_method: str,
    confidence: float,
) -> None:
    key = (candidate.amazon_order_id, link_method, candidate.acc_order_id or "")
    payload = {
        "amazon_order_id": candidate.amazon_order_id,
        "acc_order_id": candidate.acc_order_id,
        "bl_order_id": candidate.bl_order_id,
        "link_method": link_method,
        "link_confidence": confidence,
        "is_primary": False,
    }
    existing = bucket.get(key)
    if existing is None or confidence > float(existing.get("link_confidence") or 0):
        bucket[key] = payload


def _collect_seed_links(
    *,
    seed: ImportedShipmentSeed,
    by_tracking: dict[str, list[LocalPackageCandidate]],
    by_inner: dict[str, list[LocalPackageCandidate]],
    by_jjd_netfox: dict[str, list[LocalPackageCandidate]],
) -> list[dict[str, Any]]:
    bucket: dict[tuple[str, str, str], dict[str, Any]] = {}
    tracking_candidates = []
    if seed.jjd_number:
        tracking_candidates.append((seed.jjd_number, "billing_jjd", 1.0))
    if seed.parcel_number_base:
        tracking_candidates.append((seed.parcel_number_base, "billing_parcel_tracking", 0.9))
    if seed.parcel_number and seed.parcel_number != seed.parcel_number_base:
        tracking_candidates.append((seed.parcel_number, "billing_parcel_tracking_raw", 0.86))

    inner_candidates = []
    if seed.parcel_number_base:
        inner_candidates.append((seed.parcel_number_base, "billing_parcel_inner", 0.98))
    if seed.parcel_number and seed.parcel_number != seed.parcel_number_base:
        inner_candidates.append((seed.parcel_number, "billing_parcel_inner_raw", 0.94))

    for token, method, confidence in tracking_candidates:
        for candidate in by_tracking.get(_normalize_token(token), []):
            _append_link_candidate(bucket, candidate=candidate, link_method=method, confidence=confidence)
    for token, method, confidence in inner_candidates:
        for candidate in by_inner.get(_normalize_token(token), []):
            _append_link_candidate(bucket, candidate=candidate, link_method=method, confidence=confidence)
    if seed.jjd_number:
        for candidate in by_jjd_netfox.get(_normalize_token(seed.jjd_number), []):
            _append_link_candidate(
                bucket,
                candidate=candidate,
                link_method="billing_jjd_netfox_parcel_num_other",
                confidence=0.99,
            )

    links = list(bucket.values())
    if not links:
        return links
    best_confidence = max(float(item.get("link_confidence") or 0) for item in links)
    best_orders = {
        _normalize_text(item.get("amazon_order_id"))
        for item in links
        if float(item.get("link_confidence") or 0) == best_confidence
    }
    if len(best_orders) == 1:
        primary_order = next(iter(best_orders))
        primary_set = False
        for item in links:
            if (
                not primary_set
                and _normalize_text(item.get("amazon_order_id")) == primary_order
                and float(item.get("link_confidence") or 0) == best_confidence
            ):
                item["is_primary"] = True
                primary_set = True
    return links


def _load_shipment_seeds(cur, *, parcel_bases: list[str]) -> list[ImportedShipmentSeed]:
    if not parcel_bases:
        return []
    seeds: list[ImportedShipmentSeed] = []
    for batch in _chunks(parcel_bases, _BATCH_SIZE):
        placeholders = ",".join("?" for _ in batch)
        params = batch + batch
        cur.execute(
            f"""
WITH latest_map AS (
    SELECT
        m.parcel_number,
        m.parcel_number_base,
        m.jjd_number,
        m.shipment_type,
        m.ship_date,
        m.delivery_date,
        m.last_event_code,
        m.last_event_at,
        ROW_NUMBER() OVER (
            PARTITION BY m.parcel_number_base
            ORDER BY
                CASE WHEN m.delivery_date IS NOT NULL THEN 0 ELSE 1 END,
                ISNULL(m.last_event_at, ISNULL(m.delivery_date, m.ship_date)) DESC,
                m.imported_at DESC
        ) AS rn
    FROM dbo.acc_dhl_parcel_map m WITH (NOLOCK)
    WHERE m.parcel_number_base IN ({placeholders})
),
billing AS (
    SELECT
        l.parcel_number_base,
        MAX(l.parcel_number) AS parcel_number,
        MAX(l.product_code) AS product_code,
        MAX(l.description) AS description,
        MAX(l.issue_date) AS issue_date,
        MAX(l.sales_date) AS sales_date,
        MAX(l.delivery_date) AS delivery_date,
        MAX(l.sap_order) AS sap_order,
        MAX(l.shipper_receiver) AS shipper_receiver,
        SUM(ISNULL(l.net_amount, 0)) AS total_net_amount,
        COUNT(*) AS line_count
    FROM dbo.acc_dhl_billing_line l WITH (NOLOCK)
    WHERE l.parcel_number_base IN ({placeholders})
    GROUP BY l.parcel_number_base
)
SELECT
    b.parcel_number,
    b.parcel_number_base,
    m.jjd_number,
    m.shipment_type,
    m.ship_date,
    COALESCE(m.delivery_date, CAST(b.delivery_date AS DATETIME2)) AS delivery_date,
    m.last_event_code,
    m.last_event_at,
    b.product_code,
    b.description,
    b.issue_date,
    b.sales_date,
    b.sap_order,
    b.shipper_receiver,
    CAST(b.total_net_amount AS FLOAT) AS total_net_amount,
    b.line_count
FROM billing b
LEFT JOIN latest_map m
  ON m.parcel_number_base = b.parcel_number_base
 AND m.rn = 1
            """,
            params,
        )
        for row in cur.fetchall():
            seeds.append(
                ImportedShipmentSeed(
                    parcel_number=_normalize_text(row[0]) or _normalize_text(row[1]),
                    parcel_number_base=_normalize_text(row[1]),
                    jjd_number=_normalize_text(row[2]) or None,
                    shipment_type=_normalize_text(row[3]) or None,
                    ship_date=row[4],
                    delivery_date=row[5],
                    last_event_code=_normalize_text(row[6]) or None,
                    last_event_at=row[7],
                    product_code=_normalize_text(row[8]) or None,
                    description=_normalize_text(row[9]) or None,
                    issue_date=row[10],
                    sales_date=row[11],
                    sap_order=_normalize_text(row[12]) or None,
                    shipper_receiver=_normalize_text(row[13]) or None,
                    total_net_amount=float(row[14] or 0),
                    line_count=int(row[15] or 0),
                )
            )
    return seeds


def _load_all_billing_parcel_bases(cur) -> set[str]:
    cur.execute(
        """
SELECT DISTINCT parcel_number_base
FROM dbo.acc_dhl_billing_line WITH (NOLOCK)
WHERE parcel_number_base IS NOT NULL
  AND parcel_number_base <> ''
        """
    )
    return {_normalize_token(row[0]) for row in cur.fetchall() if _normalize_token(row[0])}


def _load_windowed_billing_parcel_bases(
    cur,
    *,
    created_from: date | None = None,
    created_to: date | None = None,
    limit_parcels: int | None = None,
) -> set[str]:
    where = [
        "parcel_number_base IS NOT NULL",
        "parcel_number_base <> ''",
    ]
    params: list[Any] = []
    if created_from:
        where.append("CAST(COALESCE(sales_date, issue_date) AS DATE) >= ?")
        params.append(created_from.isoformat())
    if created_to:
        where.append("CAST(COALESCE(sales_date, issue_date) AS DATE) <= ?")
        params.append(created_to.isoformat())

    top_sql = f"TOP {int(limit_parcels)} " if limit_parcels else ""
    cur.execute(
        f"""
SELECT DISTINCT {top_sql} parcel_number_base
FROM dbo.acc_dhl_billing_line WITH (NOLOCK)
WHERE {' AND '.join(where)}
ORDER BY parcel_number_base
        """,
        params,
    )
    return {_normalize_token(row[0]) for row in cur.fetchall() if _normalize_token(row[0])}


def _resolve_seed_parcel_bases(
    cur,
    *,
    changed_parcels: set[str],
    seed_all_existing: bool,
) -> set[str]:
    if seed_all_existing:
        return _load_all_billing_parcel_bases(cur)
    return {_normalize_token(value) for value in changed_parcels if _normalize_token(value)}


def _seed_shipments_from_billing(
    cur,
    *,
    parcel_bases: set[str],
) -> dict[str, int]:
    parcel_list = sorted({_normalize_token(value) for value in parcel_bases if _normalize_token(value)})
    if not parcel_list:
        return {
            "shipments_seeded": 0,
            "links_written": 0,
            "shipments_linked": 0,
            "shipments_unlinked": 0,
        }

    seeds = _load_shipment_seeds(cur, parcel_bases=parcel_list)
    tracking_values = {seed.jjd_number for seed in seeds if seed.jjd_number}
    tracking_values.update({seed.parcel_number for seed in seeds if seed.parcel_number})
    inner_values = {seed.parcel_number_base for seed in seeds if seed.parcel_number_base}
    inner_values.update({seed.parcel_number for seed in seeds if seed.parcel_number})
    by_tracking, by_inner = _load_package_lookup(cur, tracking_values=tracking_values, inner_values=inner_values)
    by_jjd_netfox = _load_jjd_netfox_lookup(cur, jjd_values=tracking_values)

    stats = {
        "shipments_seeded": 0,
        "links_written": 0,
        "shipments_linked": 0,
        "shipments_unlinked": 0,
    }
    for seed in seeds:
        delivered = seed.delivery_date is not None
        created_at_carrier = seed.ship_date or (_parse_datetime(seed.sales_date) if seed.sales_date else None)
        ship_date = seed.ship_date.date() if seed.ship_date else seed.sales_date
        payload_json = {
            "source": "dhl_billing_files",
            "parcel_number": seed.parcel_number,
            "parcel_number_base": seed.parcel_number_base,
            "jjd_number": seed.jjd_number,
            "sap_order": seed.sap_order,
            "shipper_receiver": seed.shipper_receiver,
            "shipment_type": seed.shipment_type,
            "last_event_code": seed.last_event_code,
            "line_count": seed.line_count,
            "issue_date": seed.issue_date.isoformat() if seed.issue_date else None,
            "sales_date": seed.sales_date.isoformat() if seed.sales_date else None,
            "total_net_amount": seed.total_net_amount,
        }
        shipment_payload = {
            "carrier": "DHL",
            "carrier_account": settings.DHL24_API_USERNAME or None,
            "shipment_number": seed.parcel_number_base,
            "piece_id": seed.jjd_number or seed.parcel_number_base,
            "tracking_number": seed.jjd_number or seed.parcel_number_base,
            "cedex_number": None,
            "service_code": seed.product_code or seed.shipment_type,
            "ship_date": ship_date,
            "created_at_carrier": created_at_carrier,
            "status_code": seed.last_event_code or ("DELIVERED" if delivered else "BILLING_IMPORTED"),
            "status_label": seed.last_event_code or ("Delivered from billing files" if delivered else "Imported from billing files"),
            "received_by": None,
            "is_delivered": delivered,
            "delivered_at": seed.delivery_date,
            "recipient_name": None,
            "recipient_country": None,
            "shipper_name": None,
            "shipper_country": None,
            "source_system": "dhl_billing_files",
            "source_payload_json": _json_dump(payload_json),
            "source_payload_hash": _hash_payload(payload_json),
        }
        shipment_id = _upsert_shipment(cur, shipment_payload)
        stats["shipments_seeded"] += 1
        links = _collect_seed_links(
            seed=seed,
            by_tracking=by_tracking,
            by_inner=by_inner,
            by_jjd_netfox=by_jjd_netfox,
        )
        if links:
            stats["links_written"] += _upsert_shipment_links(cur, shipment_id=shipment_id, links=links)
            stats["shipments_linked"] += 1
        else:
            stats["shipments_unlinked"] += 1
    return stats


def import_dhl_billing_files(
    *,
    invoice_root: str | None = None,
    jj_root: str | None = None,
    manifest_path: str | None = None,
    include_shipment_seed: bool = True,
    seed_all_existing: bool = False,
    force_reimport: bool = False,
    limit_invoice_files: int | None = None,
    limit_jj_files: int | None = None,
    job_id: str | None = None,
) -> dict[str, Any]:
    ensure_dhl_schema()

    from app.connectors.mssql.mssql_store import set_job_progress

    invoice_root_path = Path(invoice_root or settings.DHL_BILLING_ROOT_PATH)
    jj_root_path = Path(jj_root or settings.DHL_BILLING_JJ_PATH)
    manifest_value = manifest_path if manifest_path is not None else settings.DHL_BILLING_MANIFEST_PATH
    manifest_file = Path(manifest_value) if manifest_value else None

    invoice_files = _discover_invoice_files(invoice_root_path)
    jj_files = _discover_jj_files(jj_root_path)
    if limit_invoice_files is not None:
        invoice_files = invoice_files[: max(0, int(limit_invoice_files))]
    if limit_jj_files is not None:
        jj_files = jj_files[: max(0, int(limit_jj_files))]

    stats: dict[str, Any] = {
        "manifest_rows_imported": 0,
        "invoice_files_found": len(invoice_files),
        "invoice_files_imported": 0,
        "invoice_files_skipped": 0,
        "invoice_line_rows_imported": 0,
        "invoice_documents_upserted": 0,
        "jj_files_found": len(jj_files),
        "jj_files_imported": 0,
        "jj_files_skipped": 0,
        "jj_rows_imported": 0,
        "shipments_seeded": 0,
        "links_written": 0,
        "shipments_linked": 0,
        "shipments_unlinked": 0,
        "seed_scope": "all_existing" if seed_all_existing else "changed_only",
        "seed_parcel_bases": 0,
        "errors": 0,
    }
    changed_parcels: set[str] = set()

    conn = _connect()
    try:
        cur = conn.cursor()
        if manifest_file and manifest_file.exists():
            file_path = str(manifest_file.resolve())
            file_size = manifest_file.stat().st_size
            file_mtime = _file_mtime_utc(manifest_file)
            state = _load_import_file_state(cur, source_kind="manifest", file_path=file_path)
            if not _should_skip_import_file(
                state,
                file_size=file_size,
                file_mtime=file_mtime,
                force_reimport=force_reimport,
            ):
                manifest_rows = parse_dhl_manifest_xlsx(file_path)
                for row in manifest_rows:
                    _upsert_billing_document(
                        cur,
                        document_number=row["document_number"],
                        document_type=row.get("document_type"),
                        issue_date=row.get("issue_date"),
                        ship_date=row.get("ship_date"),
                        due_date=row.get("due_date"),
                        net_amount=row.get("net_amount"),
                        vat_amount=row.get("vat_amount"),
                        gross_amount=row.get("gross_amount"),
                        source_file=None,
                        source_manifest_file=file_path,
                        detail_rows_count=None,
                    )
                _upsert_import_file(
                    cur,
                    source_kind="manifest",
                    file_path=file_path,
                    file_name=manifest_file.name,
                    document_number=None,
                    file_size_bytes=file_size,
                    file_mtime_utc=file_mtime,
                    status="imported",
                    rows_imported=len(manifest_rows),
                    error_message=None,
                )
                conn.commit()
                stats["manifest_rows_imported"] = len(manifest_rows)

        total_steps = max(len(invoice_files) + len(jj_files), 1)
        processed_steps = 0

        for path in invoice_files:
            file_path = str(path)
            file_size = path.stat().st_size
            file_mtime = _file_mtime_utc(path)
            state = _load_import_file_state(cur, source_kind="invoice", file_path=file_path)
            if _should_skip_import_file(
                state,
                file_size=file_size,
                file_mtime=file_mtime,
                force_reimport=force_reimport,
            ):
                stats["invoice_files_skipped"] += 1
                processed_steps += 1
                continue
            try:
                rows = parse_dhl_invoice_xlsx(file_path)
                grouped: dict[str, list[dict[str, Any]]] = {}
                for row in rows:
                    grouped.setdefault(row["document_number"], []).append(row)
                    changed_parcels.add(row["parcel_number_base"])
                for document_number, document_rows in grouped.items():
                    first_row = document_rows[0]
                    _upsert_billing_document(
                        cur,
                        document_number=document_number,
                        document_type="Załącznik do dokumentu",
                        issue_date=first_row.get("issue_date"),
                        ship_date=first_row.get("sales_date"),
                        due_date=first_row.get("due_date"),
                        net_amount=sum(float(item.get("net_amount") or 0) for item in document_rows),
                        vat_amount=None,
                        gross_amount=None,
                        source_file=file_path,
                        source_manifest_file=None,
                        detail_rows_count=len(document_rows),
                    )
                    _replace_billing_lines(cur, document_number=document_number, source_file=file_path, rows=document_rows)
                    stats["invoice_documents_upserted"] += 1
                _upsert_import_file(
                    cur,
                    source_kind="invoice",
                    file_path=file_path,
                    file_name=path.name,
                    document_number=_invoice_file_document_number(path),
                    file_size_bytes=file_size,
                    file_mtime_utc=file_mtime,
                    status="imported",
                    rows_imported=len(rows),
                    error_message=None,
                )
                conn.commit()
                stats["invoice_files_imported"] += 1
                stats["invoice_line_rows_imported"] += len(rows)
            except Exception as exc:
                conn.rollback()
                _upsert_import_file(
                    cur,
                    source_kind="invoice",
                    file_path=file_path,
                    file_name=path.name,
                    document_number=_invoice_file_document_number(path),
                    file_size_bytes=file_size,
                    file_mtime_utc=file_mtime,
                    status="failed",
                    rows_imported=0,
                    error_message=str(exc),
                )
                conn.commit()
                stats["errors"] += 1
                log.exception("dhl_billing_import.invoice_failed", path=file_path)
            processed_steps += 1
            if job_id:
                pct = 10 + int((processed_steps / total_steps) * 70)
                set_job_progress(
                    job_id,
                    progress_pct=min(pct, 85),
                    records_processed=stats["invoice_line_rows_imported"] + stats["jj_rows_imported"],
                    message=f"DHL billing import invoice {processed_steps}/{total_steps}",
                )

        for path in jj_files:
            file_path = str(path)
            file_size = path.stat().st_size
            file_mtime = _file_mtime_utc(path)
            state = _load_import_file_state(cur, source_kind="jj_map", file_path=file_path)
            if _should_skip_import_file(
                state,
                file_size=file_size,
                file_mtime=file_mtime,
                force_reimport=force_reimport,
            ):
                stats["jj_files_skipped"] += 1
                processed_steps += 1
                continue
            try:
                rows = parse_dhl_jj_xlsx(file_path)
                for row in rows:
                    changed_parcels.add(row["parcel_number_base"])
                _run_with_lock_retry(
                    conn,
                    lambda: cur.execute("DELETE FROM dbo.acc_dhl_parcel_map WHERE source_file = ?", [file_path]),
                    op_name="jj_delete_existing",
                    source_file=file_path,
                    extra={"rows": len(rows)},
                )
                conn.commit()

                written_rows = 0
                for chunk_idx, chunk_rows in enumerate(_chunks(rows, _JJ_INSERT_BATCH_SIZE), start=1):
                    payloads = _build_parcel_map_payloads(source_file=file_path, rows=chunk_rows)
                    _run_with_lock_retry(
                        conn,
                        lambda payloads=payloads: _insert_parcel_map_payloads(cur, payloads),
                        op_name="jj_insert_chunk",
                        source_file=file_path,
                        extra={
                            "chunk_idx": chunk_idx,
                            "chunk_rows": len(chunk_rows),
                        },
                    )
                    conn.commit()
                    written_rows += len(chunk_rows)
                _upsert_import_file(
                    cur,
                    source_kind="jj_map",
                    file_path=file_path,
                    file_name=path.name,
                    document_number=None,
                    file_size_bytes=file_size,
                    file_mtime_utc=file_mtime,
                    status="imported",
                    rows_imported=written_rows,
                    error_message=None,
                )
                conn.commit()
                stats["jj_files_imported"] += 1
                stats["jj_rows_imported"] += written_rows
            except Exception as exc:
                conn.rollback()
                try:
                    _run_with_lock_retry(
                        conn,
                        lambda: cur.execute("DELETE FROM dbo.acc_dhl_parcel_map WHERE source_file = ?", [file_path]),
                        op_name="jj_cleanup_partial",
                        source_file=file_path,
                    )
                    conn.commit()
                except Exception as cleanup_exc:
                    conn.rollback()
                    log.exception(
                        "dhl_billing_import.jj_cleanup_failed",
                        path=file_path,
                        error=str(cleanup_exc),
                    )
                _upsert_import_file(
                    cur,
                    source_kind="jj_map",
                    file_path=file_path,
                    file_name=path.name,
                    document_number=None,
                    file_size_bytes=file_size,
                    file_mtime_utc=file_mtime,
                    status="failed",
                    rows_imported=0,
                    error_message=str(exc),
                )
                conn.commit()
                stats["errors"] += 1
                log.exception("dhl_billing_import.jj_failed", path=file_path)
            processed_steps += 1
            if job_id:
                pct = 10 + int((processed_steps / total_steps) * 70)
                set_job_progress(
                    job_id,
                    progress_pct=min(pct, 85),
                    records_processed=stats["invoice_line_rows_imported"] + stats["jj_rows_imported"],
                    message=f"DHL billing import files {processed_steps}/{total_steps}",
                )

        seed_parcels = _resolve_seed_parcel_bases(
            cur,
            changed_parcels=changed_parcels,
            seed_all_existing=seed_all_existing,
        )
        stats["seed_parcel_bases"] = len(seed_parcels)
        if include_shipment_seed and seed_parcels:
            if job_id:
                set_job_progress(
                    job_id,
                    progress_pct=90,
                    records_processed=stats["invoice_line_rows_imported"] + stats["jj_rows_imported"],
                    message="DHL billing import seeding shipments",
                )
            seed_stats = _seed_shipments_from_billing(cur, parcel_bases=seed_parcels)
            conn.commit()
            stats.update(seed_stats)

        if job_id:
            set_job_progress(
                job_id,
                progress_pct=95,
                records_processed=stats["invoice_line_rows_imported"] + stats["jj_rows_imported"],
                message="DHL billing import finished",
            )
        return stats
    finally:
        conn.close()


def seed_dhl_shipments_from_staging(
    *,
    created_from: date | None = None,
    created_to: date | None = None,
    seed_all_existing: bool = True,
    limit_parcels: int | None = None,
    job_id: str | None = None,
) -> dict[str, Any]:
    ensure_dhl_schema()

    from app.connectors.mssql.mssql_store import set_job_progress

    stats: dict[str, Any] = {
        "seed_scope": "all_existing" if seed_all_existing else "window",
        "seed_parcel_bases": 0,
        "shipments_seeded": 0,
        "links_written": 0,
        "shipments_linked": 0,
        "shipments_unlinked": 0,
    }

    conn = _connect()
    try:
        cur = conn.cursor()
        if seed_all_existing:
            parcel_bases = _load_all_billing_parcel_bases(cur)
            if limit_parcels:
                parcel_bases = set(sorted(parcel_bases)[: int(limit_parcels)])
        else:
            parcel_bases = _load_windowed_billing_parcel_bases(
                cur,
                created_from=created_from,
                created_to=created_to,
                limit_parcels=limit_parcels,
            )
        stats["seed_parcel_bases"] = len(parcel_bases)
        if not parcel_bases:
            return stats

        if job_id:
            set_job_progress(
                job_id,
                progress_pct=20,
                records_processed=0,
                message=f"DHL seed shipments scope={len(parcel_bases)}",
            )

        seed_stats = _seed_shipments_from_billing(cur, parcel_bases=parcel_bases)
        conn.commit()
        stats.update(seed_stats)

        if job_id:
            set_job_progress(
                job_id,
                progress_pct=95,
                records_processed=int(stats.get("shipments_seeded", 0) or 0),
                message="DHL seed shipments finished",
            )
        return stats
    finally:
        conn.close()

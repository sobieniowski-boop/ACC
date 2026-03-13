"""Import Products service — parse CEO's Excel file and manage import product data.

This module handles:
1. Parsing the SharePoint/OneDrive Excel file (headers on row 3)
2. Upserting data into acc_import_products table
3. Querying import SKUs for dashboard flagging

The Excel columns (from row 3):
  FOTO | SKU | Nazwa pełna | Kod K | Kod importu | Aktywny |
  Data pierwszej dostawy | STAN MAGAZYNOWY | w tym FBA | Sprzedaż 30 dni |
  AMAZON | w tym FBA2 | ALLEGRO | SKLEP | INNE | Zasięg dni |
  Estymacja braku stanu MAG | Dynamika 10 do 30 | Data ostatniej dostawy |
  Ilość ostatniej dostawy | Cena zakupu | Wartość magazynu |
  Średnia cena sprzedaży 30 dni | Średnia marża | MARŻA | Miejsc paletowych |
  Koszt składowania 1szt na 30dni | Koszt składowania zapasu na 30 dni |
  Nasycenie ost 12m | Data dostawy | Tempo (Pokrycie 150 dni) |
  Sprzedaż ost 12m | FILTR | Mix
"""
from __future__ import annotations

import io
import math
from datetime import datetime, timezone
from typing import Any, Optional

import structlog
from openpyxl import load_workbook

from app.core.db_connection import connect_acc

log = structlog.get_logger(__name__)

# ---------------------------------------------------------------------------
# Schema bootstrap
# ---------------------------------------------------------------------------

_CREATE_TABLE_SQL = """\
IF OBJECT_ID('dbo.acc_import_products', 'U') IS NULL
BEGIN
    CREATE TABLE dbo.acc_import_products (
        id               INT IDENTITY(1,1) PRIMARY KEY,
        sku              NVARCHAR(120)  NOT NULL,
        nazwa_pelna      NVARCHAR(500)  NULL,
        kod_k            NVARCHAR(120)  NULL,
        kod_importu      NVARCHAR(120)  NULL,
        aktywny          BIT            NULL,
        data_pierwszej_dostawy  DATE    NULL,
        stan_magazynowy  INT            NULL,
        w_tym_fba        INT            NULL,
        sprzedaz_30d     INT            NULL,
        amazon_30d       INT            NULL,
        fba_30d          INT            NULL,
        allegro_30d      INT            NULL,
        sklep_30d        INT            NULL,
        inne_30d         INT            NULL,
        zasieg_dni       INT            NULL,
        estymacja_braku_stanu  DATE     NULL,
        dynamika_10_30   FLOAT          NULL,
        data_ostatniej_dostawy DATE     NULL,
        ilosc_ostatniej_dostawy INT     NULL,
        cena_zakupu      DECIMAL(18,4)  NULL,
        wartosc_magazynu DECIMAL(18,2)  NULL,
        srednia_cena_sprzedazy_30d DECIMAL(18,4) NULL,
        srednia_marza    DECIMAL(18,4)  NULL,
        marza            DECIMAL(18,4)  NULL,
        miejsc_paletowych FLOAT         NULL,
        koszt_skladowania_1szt_30d DECIMAL(18,4) NULL,
        koszt_skladowania_zapasu_30d DECIMAL(18,2) NULL,
        nasycenie_12m    FLOAT          NULL,
        data_dostawy     DATE           NULL,
        tempo_pokrycie_150d INT         NULL,
        sprzedaz_12m     INT            NULL,
        filtr            NVARCHAR(120)  NULL,
        mix              NVARCHAR(120)  NULL,
        is_import        BIT            NOT NULL DEFAULT 1,
        uploaded_at      DATETIME2      NOT NULL DEFAULT SYSUTCDATETIME(),
        updated_at       DATETIME2      NOT NULL DEFAULT SYSUTCDATETIME()
    );
    CREATE UNIQUE INDEX UX_acc_import_products_sku ON dbo.acc_import_products(sku);
    CREATE INDEX IX_acc_import_products_import ON dbo.acc_import_products(is_import);
    CREATE INDEX IX_acc_import_products_kod ON dbo.acc_import_products(kod_importu);
END
"""


def ensure_import_products_schema() -> None:
    """No-op — schema managed by Alembic migration eb020."""


# ---------------------------------------------------------------------------
# Excel parsing
# ---------------------------------------------------------------------------

# Column mapping: Excel header → DB column name
_HEADER_MAP: dict[str, str] = {
    "sku": "sku",
    "nazwa pełna": "nazwa_pelna",
    "kod k": "kod_k",
    "kod importu": "kod_importu",
    "aktywny": "aktywny",
    "data pierwszej dostawy": "data_pierwszej_dostawy",
    "stan magazynowy": "stan_magazynowy",
    "w tym fba": "w_tym_fba",
    "sprzedaż 30 dni": "sprzedaz_30d",
    "amazon": "amazon_30d",
    "w tym fba2": "fba_30d",
    "allegro": "allegro_30d",
    "sklep": "sklep_30d",
    "inne": "inne_30d",
    "zasięg dni": "zasieg_dni",
    "estymacja braku stanu mag": "estymacja_braku_stanu",
    "dynamika 10 do 30": "dynamika_10_30",
    "data ostatniej dostawy": "data_ostatniej_dostawy",
    "ilość ostatniej dostawy": "ilosc_ostatniej_dostawy",
    "cena zakupu": "cena_zakupu",
    "wartość magazynu": "wartosc_magazynu",
    "średnia cena sprzedaży 30 dni": "srednia_cena_sprzedazy_30d",
    "średnia marża": "srednia_marza",
    "marża": "marza",
    "miejsc paletowych": "miejsc_paletowych",
    "koszt składowania 1szt na 30dni": "koszt_skladowania_1szt_30d",
    "koszt składowania zapasu na 30 dni": "koszt_skladowania_zapasu_30d",
    "nasycenie ost 12m": "nasycenie_12m",
    "data dostawy": "data_dostawy",
    "tempo (pokrycie 150 dni)": "tempo_pokrycie_150d",
    "sprzedaż ost 12m": "sprzedaz_12m",
    "filtr": "filtr",
    "mix": "mix",
}

# Columns that hold dates
_DATE_COLS = {
    "data_pierwszej_dostawy",
    "estymacja_braku_stanu",
    "data_ostatniej_dostawy",
    "data_dostawy",
}

# Columns that hold integers
_INT_COLS = {
    "stan_magazynowy", "w_tym_fba", "sprzedaz_30d", "amazon_30d",
    "fba_30d", "allegro_30d", "sklep_30d", "inne_30d", "zasieg_dni",
    "ilosc_ostatniej_dostawy", "tempo_pokrycie_150d", "sprzedaz_12m",
}

# Columns that hold floats/decimals
_FLOAT_COLS = {
    "dynamika_10_30", "cena_zakupu", "wartosc_magazynu",
    "srednia_cena_sprzedazy_30d", "srednia_marza", "marza",
    "miejsc_paletowych", "koszt_skladowania_1szt_30d",
    "koszt_skladowania_zapasu_30d", "nasycenie_12m",
}


def _safe_int(val: Any) -> Optional[int]:
    if val is None:
        return None
    try:
        return int(float(str(val).replace(",", ".").replace(" ", "")))
    except (ValueError, TypeError):
        return None


def _safe_float(val: Any) -> Optional[float]:
    if val is None:
        return None
    try:
        s = str(val).replace(",", ".").replace(" ", "").replace("%", "")
        return round(float(s), 4) if s else None
    except (ValueError, TypeError):
        return None


def _safe_date(val: Any) -> Optional[str]:
    """Convert to ISO date string or None."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    try:
        from dateutil.parser import parse as dparse
        return dparse(str(val), dayfirst=True).strftime("%Y-%m-%d")
    except Exception:
        return None


def _safe_bool(val: Any) -> Optional[bool]:
    if val is None:
        return None
    s = str(val).strip().lower()
    if s in ("tak", "yes", "true", "1", "t", "y"):
        return True
    if s in ("nie", "no", "false", "0", "n"):
        return False
    return None


def parse_import_excel(file_bytes: bytes) -> list[dict[str, Any]]:
    """Parse the CEO's import products Excel file.

    Headers are on row 3. Data starts from row 4.
    Returns list of dicts with DB column names as keys.
    """
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("No active worksheet found")

    # Read headers from row 3
    header_row = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    col_indices: dict[int, str] = {}
    for idx, hdr in enumerate(header_row):
        if hdr is None:
            continue
        key = str(hdr).strip().lower()
        if key in _HEADER_MAP:
            col_indices[idx] = _HEADER_MAP[key]

    if "sku" not in col_indices.values():
        # Try row 1 and 2 as fallback
        for try_row in (1, 2):
            header_row = list(ws.iter_rows(min_row=try_row, max_row=try_row, values_only=True))[0]
            col_indices = {}
            for idx, hdr in enumerate(header_row):
                if hdr is None:
                    continue
                key = str(hdr).strip().lower()
                if key in _HEADER_MAP:
                    col_indices[idx] = _HEADER_MAP[key]
            if "sku" in col_indices.values():
                break

    if "sku" not in col_indices.values():
        raise ValueError(
            "Could not find 'SKU' column in the Excel file. "
            "Expected headers: " + ", ".join(_HEADER_MAP.keys())
        )

    log.info(
        "import_excel.headers_found",
        mapped_cols=len(col_indices),
        columns=list(col_indices.values()),
    )

    # Read data rows (start after header row)
    header_row_num = 3  # default
    for try_row in (3, 1, 2):
        row_data = list(ws.iter_rows(min_row=try_row, max_row=try_row, values_only=True))[0]
        for idx, hdr in enumerate(row_data):
            if hdr and str(hdr).strip().lower() == "sku":
                header_row_num = try_row
                break

    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=header_row_num + 1, values_only=True):
        record: dict[str, Any] = {}
        for idx, db_col in col_indices.items():
            val = row[idx] if idx < len(row) else None

            if db_col == "aktywny":
                record[db_col] = _safe_bool(val)
            elif db_col in _DATE_COLS:
                record[db_col] = _safe_date(val)
            elif db_col in _INT_COLS:
                record[db_col] = _safe_int(val)
            elif db_col in _FLOAT_COLS:
                record[db_col] = _safe_float(val)
            else:
                record[db_col] = str(val).strip() if val is not None else None

        # Skip rows without SKU
        sku = record.get("sku")
        if not sku or str(sku).strip() in ("", "None", "nan"):
            continue

        record["sku"] = str(sku).strip()
        rows.append(record)

    wb.close()
    log.info("import_excel.parsed", total_rows=len(rows))
    return rows


# ---------------------------------------------------------------------------
# Database operations
# ---------------------------------------------------------------------------

def upsert_import_products(rows: list[dict[str, Any]]) -> dict[str, int]:
    """Upsert parsed import products into acc_import_products.

    Uses MERGE for atomic upsert. Returns counts.
    """
    if not rows:
        return {"inserted": 0, "updated": 0, "total": 0}

    ensure_import_products_schema()

    conn = connect_acc(autocommit=False)
    inserted = 0
    updated = 0

    try:
        cur = conn.cursor()
        now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")

        for row in rows:
            sku = row["sku"]

            # Check if exists
            cur.execute(
                "SELECT id FROM dbo.acc_import_products WHERE sku = ?",
                sku,
            )
            existing = cur.fetchone()

            # Build column list (excluding sku for update)
            cols = [k for k in row.keys() if k != "sku" and row[k] is not None]

            if existing:
                # UPDATE
                if cols:
                    set_parts = [f"{c} = ?" for c in cols]
                    set_parts.append("updated_at = ?")
                    set_parts.append("is_import = 1")
                    vals = [row[c] for c in cols] + [now]
                    sql = (
                        f"UPDATE dbo.acc_import_products "
                        f"SET {', '.join(set_parts)} "
                        f"WHERE sku = ?"
                    )
                    cur.execute(sql, *vals, sku)
                updated += 1
            else:
                # INSERT
                all_cols = ["sku"] + cols + ["is_import", "uploaded_at", "updated_at"]
                all_vals = [sku] + [row[c] for c in cols] + [True, now, now]
                placeholders = ", ".join(["?"] * len(all_vals))
                sql = (
                    f"INSERT INTO dbo.acc_import_products ({', '.join(all_cols)}) "
                    f"VALUES ({placeholders})"
                )
                cur.execute(sql, *all_vals)
                inserted += 1

            # Commit every 100 rows
            if (inserted + updated) % 100 == 0:
                conn.commit()

        conn.commit()
        cur.close()
        log.info(
            "import_products.upsert_done",
            inserted=inserted,
            updated=updated,
            total=len(rows),
        )
    except Exception as e:
        conn.rollback()
        log.error("import_products.upsert_error", error=str(e))
        raise
    finally:
        conn.close()

    return {"inserted": inserted, "updated": updated, "total": len(rows)}


def get_import_skus() -> set[str]:
    """Return set of SKUs that are flagged as import products."""
    conn = connect_acc(timeout=10)
    try:
        cur = conn.cursor()
        cur.execute(
            "SELECT sku FROM dbo.acc_import_products WITH (NOLOCK) "
            "WHERE is_import = 1"
        )
        skus = {str(row[0]).strip() for row in cur.fetchall()}
        cur.close()
        return skus
    except Exception as e:
        log.error("import_products.get_skus_error", error=str(e))
        return set()
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Amazon aggregate CTE (reused by list + summary)
# ---------------------------------------------------------------------------
_AMZ_CTE = """
amz AS (
    SELECT
        p.internal_sku,
        SUM(ISNULL(ol.quantity_ordered, 0))          AS amz_units,
        COUNT(DISTINCT o.id)                         AS amz_orders,
        SUM(
            (ISNULL(ol.item_price, 0)
             - ISNULL(ol.item_tax, 0)
             - ISNULL(ol.promotion_discount, 0))
            * ISNULL(fx.rate_to_pln,
                ISNULL(fx_latest.rate_to_pln,
                    CASE WHEN o.currency = 'PLN' THEN 1.0 ELSE NULL END))
        )                                            AS amz_revenue_pln,
        SUM(ISNULL(ol.cogs_pln, 0))                 AS amz_cogs_pln,
        SUM(ISNULL(ol.fba_fee_pln, 0)
          + ISNULL(ol.referral_fee_pln, 0))          AS amz_fees_pln,
        SUM(CASE WHEN ol.cogs_pln IS NOT NULL
                  AND ol.cogs_pln > 0 THEN 1 ELSE 0 END) AS lines_with_cogs,
        COUNT(*)                                     AS total_lines
    FROM dbo.acc_order_line ol WITH (NOLOCK)
    JOIN dbo.acc_order o WITH (NOLOCK) ON o.id = ol.order_id
    JOIN dbo.acc_product p WITH (NOLOCK) ON p.id = ol.product_id
    OUTER APPLY (
        SELECT TOP 1 rate_to_pln
        FROM dbo.acc_exchange_rate er WITH (NOLOCK)
        WHERE er.currency = o.currency
          AND er.rate_date <= CAST(o.purchase_date AS DATE)
        ORDER BY er.rate_date DESC
    ) fx
    OUTER APPLY (
        SELECT TOP 1 rate_to_pln
        FROM dbo.acc_exchange_rate er2 WITH (NOLOCK)
        WHERE er2.currency = o.currency
        ORDER BY er2.rate_date DESC
    ) fx_latest
    WHERE o.status = 'Shipped'
      AND CAST(o.purchase_date AS DATE) >= DATEADD(day, -30, CAST(GETDATE() AS DATE))
      AND p.internal_sku IS NOT NULL
    GROUP BY p.internal_sku
)
"""

_NUMERIC_SORT_COLS = {
    "stan_magazynowy", "w_tym_fba", "sprzedaz_30d", "zasieg_dni",
    "cena_zakupu", "wartosc_magazynu", "marza", "srednia_marza",
    "sprzedaz_12m", "nasycenie_12m", "dynamika_10_30",
    "amz_units_30d", "amz_orders_30d", "amz_revenue_pln_30d",
    "amz_cogs_pln_30d", "amz_fees_pln_30d", "amz_cm1_pln_30d",
    "amz_cm1_pct_30d", "amz_avg_price_pln", "amz_cogs_coverage_pct",
}

_ALL_SORT_COLS = _NUMERIC_SORT_COLS | {
    "sku", "nazwa_pelna", "kod_importu", "aktywny",
}


def get_import_products_list(
    page: int = 1,
    page_size: int = 50,
    sku_search: Optional[str] = None,
    aktywny: Optional[bool] = None,
    kod_importu: Optional[str] = None,
    has_amazon_sales: Optional[bool] = None,
    min_zasieg: Optional[int] = None,
    max_zasieg: Optional[int] = None,
    sort_by: str = "sku",
    sort_dir: str = "asc",
) -> dict[str, Any]:
    """Return paginated import products with Amazon metrics (same logic as profit engine)."""
    conn = connect_acc(timeout=30)
    try:
        cur = conn.cursor()

        # --- Dynamic WHERE on import_products ---
        where_parts = ["ip.is_import = 1"]
        params: list[Any] = []

        if sku_search:
            where_parts.append("(ip.sku LIKE ? OR ip.nazwa_pelna LIKE ?)")
            params.extend([f"%{sku_search}%", f"%{sku_search}%"])
        if aktywny is not None:
            where_parts.append("ip.aktywny = ?")
            params.append(1 if aktywny else 0)
        if kod_importu:
            where_parts.append("ip.kod_importu = ?")
            params.append(kod_importu)
        if min_zasieg is not None:
            where_parts.append("ISNULL(ip.zasieg_dni, 0) >= ?")
            params.append(min_zasieg)
        if max_zasieg is not None:
            where_parts.append("ISNULL(ip.zasieg_dni, 0) <= ?")
            params.append(max_zasieg)

        where_sql = " AND ".join(where_parts)

        # --- Fetch all matching rows with Amazon CTE ---
        sql = f"""
            WITH {_AMZ_CTE}
            SELECT
                ip.*,
                ISNULL(amz.amz_units, 0)           AS amz_units_30d,
                ISNULL(amz.amz_orders, 0)           AS amz_orders_30d,
                ISNULL(amz.amz_revenue_pln, 0)      AS amz_revenue_pln_30d,
                ISNULL(amz.amz_cogs_pln, 0)         AS amz_cogs_pln_30d,
                ISNULL(amz.amz_fees_pln, 0)         AS amz_fees_pln_30d,
                ISNULL(amz.lines_with_cogs, 0)      AS amz_lines_with_cogs,
                ISNULL(amz.total_lines, 0)          AS amz_total_lines
            FROM dbo.acc_import_products ip WITH (NOLOCK)
            LEFT JOIN amz ON ip.sku = amz.internal_sku
            WHERE {where_sql}
        """
        cur.execute(sql, *params)
        columns = [desc[0] for desc in cur.description]
        raw_rows = cur.fetchall()
        cur.close()

        # --- Build items with derived Amazon fields ---
        all_items: list[dict[str, Any]] = []
        for row in raw_rows:
            item: dict[str, Any] = {}
            for i, col in enumerate(columns):
                val = row[i]
                if isinstance(val, datetime):
                    val = val.isoformat()
                elif hasattr(val, "isoformat"):
                    val = val.isoformat()
                item[col] = val

            rev = float(item.get("amz_revenue_pln_30d") or 0)
            cogs = float(item.get("amz_cogs_pln_30d") or 0)
            fees = float(item.get("amz_fees_pln_30d") or 0)
            units = int(item.get("amz_units_30d") or 0)
            tl = int(item.get("amz_total_lines") or 0)
            lc = int(item.get("amz_lines_with_cogs") or 0)
            cm1 = round(rev - cogs - fees, 2)

            item["amz_cm1_pln_30d"] = cm1
            item["amz_cm1_pct_30d"] = round(cm1 / rev * 100, 1) if rev else 0.0
            item["amz_avg_price_pln"] = round(rev / units, 2) if units else 0.0
            item["amz_cogs_coverage_pct"] = round(lc / tl * 100, 1) if tl else 0.0

            all_items.append(item)

        # --- has_amazon_sales filter (post-SQL) ---
        if has_amazon_sales is True:
            all_items = [i for i in all_items if i["amz_units_30d"] > 0]
        elif has_amazon_sales is False:
            all_items = [i for i in all_items if i["amz_units_30d"] == 0]

        # --- Sort ---
        sort_key = sort_by if sort_by in _ALL_SORT_COLS else "sku"
        reverse = sort_dir.lower() == "desc"
        if sort_key in _NUMERIC_SORT_COLS:
            all_items.sort(
                key=lambda x: float(x.get(sort_key) or 0), reverse=reverse
            )
        else:
            all_items.sort(
                key=lambda x: str(x.get(sort_key) or "").lower(), reverse=reverse
            )

        # --- Paginate ---
        total = len(all_items)
        pages = math.ceil(total / page_size) if page_size > 0 else 1
        start = (page - 1) * page_size
        page_items = all_items[start : start + page_size]

        return {
            "total": total,
            "page": page,
            "page_size": page_size,
            "pages": pages,
            "items": page_items,
        }
    except Exception as e:
        log.error("import_products.list_error", error=str(e))
        raise
    finally:
        conn.close()


def get_import_products_summary() -> dict[str, Any]:
    """Summary stats for import products — Holding (CEO Excel) + Amazon (our data)."""
    conn = connect_acc(timeout=20)
    try:
        cur = conn.cursor()

        # ── Holding stats (from CEO's Excel) ──
        cur.execute("""
            SELECT
                COUNT(*)                                      AS total_products,
                SUM(CASE WHEN aktywny = 1 THEN 1 ELSE 0 END) AS active_count,
                SUM(ISNULL(stan_magazynowy, 0))               AS total_stock,
                SUM(ISNULL(wartosc_magazynu, 0))              AS total_stock_value,
                AVG(CASE WHEN marza IS NOT NULL AND marza <> 0
                         THEN marza END)                      AS avg_margin,
                SUM(ISNULL(sprzedaz_30d, 0))                  AS total_sales_30d
            FROM dbo.acc_import_products WITH (NOLOCK)
            WHERE is_import = 1
        """)
        h = cur.fetchone()

        # ── Amazon stats (last 30 days, same logic as profit engine) ──
        cur.execute(f"""
            WITH {_AMZ_CTE}
            SELECT
                ISNULL(SUM(amz.amz_units), 0)        AS amz_units,
                ISNULL(SUM(amz.amz_orders), 0)       AS amz_orders,
                ISNULL(SUM(amz.amz_revenue_pln), 0)  AS amz_revenue,
                ISNULL(SUM(amz.amz_cogs_pln), 0)     AS amz_cogs,
                ISNULL(SUM(amz.amz_fees_pln), 0)     AS amz_fees,
                COUNT(*)                              AS amz_sku_count
            FROM amz
            WHERE amz.internal_sku IN (
                SELECT sku FROM dbo.acc_import_products WITH (NOLOCK)
                WHERE is_import = 1
            )
        """)
        a = cur.fetchone()
        cur.close()

        amz_rev = float(a[2] or 0)
        amz_cogs = float(a[3] or 0)
        amz_fees = float(a[4] or 0)
        amz_cm1 = round(amz_rev - amz_cogs - amz_fees, 2)

        return {
            "total_products": h[0] or 0,
            "active_count": h[1] or 0,
            # Holding
            "holding_total_stock": h[2] or 0,
            "holding_stock_value": round(float(h[3] or 0), 2),
            "holding_avg_margin": round(float(h[4] or 0), 1),
            "holding_sales_30d": h[5] or 0,
            # Amazon
            "amz_units_30d": int(a[0] or 0),
            "amz_orders_30d": int(a[1] or 0),
            "amz_revenue_30d": round(amz_rev, 2),
            "amz_cogs_30d": round(amz_cogs, 2),
            "amz_fees_30d": round(amz_fees, 2),
            "amz_cm1_30d": amz_cm1,
            "amz_cm1_pct_30d": round(amz_cm1 / amz_rev * 100, 1) if amz_rev else 0.0,
            "amz_products_with_sales": int(a[5] or 0),
        }
    except Exception as e:
        log.error("import_products.summary_error", error=str(e))
        return {}
    finally:
        conn.close()


def get_import_filter_options() -> dict[str, Any]:
    """Return distinct filter values for the import products UI."""
    conn = connect_acc(timeout=10)
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT DISTINCT kod_importu
            FROM dbo.acc_import_products WITH (NOLOCK)
            WHERE is_import = 1 AND kod_importu IS NOT NULL
            ORDER BY kod_importu
        """)
        kod_importu_vals = [str(r[0]).strip() for r in cur.fetchall() if r[0]]
        cur.close()
        return {"kod_importu": kod_importu_vals}
    except Exception as e:
        log.error("import_products.filter_options_error", error=str(e))
        return {"kod_importu": []}
    finally:
        conn.close()

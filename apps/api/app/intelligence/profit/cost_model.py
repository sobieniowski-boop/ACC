"""Profit Engine - cost model, TKL logistics, FX, pricing, schemas.

Extracted from the monolithic profit_engine.py (Sprint 3).
Manages logistics cost tables (TKL XLSX), exchange rates, purchase pricing,
official price workbooks, cost model configuration, and schema bootstrapping.
"""
from __future__ import annotations

import csv
import io
import re
import threading
import time
import uuid
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from urllib.request import urlopen

import pyodbc
import structlog
from openpyxl import load_workbook

from app.core.config import settings, MARKETPLACE_REGISTRY
from app.core.db_connection import connect_acc
from app.services.order_logistics_source import (
    profit_logistics_join_sql,
    profit_logistics_value_sql,
)

from app.intelligence.profit.helpers import (
    _connect, _fetchall_dict, _f, _f_strict, _i, _mkt_code,
    _norm_text, _norm_internal_sku,
    _warnings_reset, _warnings_append, _warnings_collect,
    _result_cache_get, _result_cache_set, _result_cache_invalidate,
    RENEWED_SKU_FILTER,
)

log = structlog.get_logger(__name__)

def _find_official_price_workbook() -> Path | None:
    candidates = [
        Path(r"N:\Analityka"),
        Path(r"C:\ACC\tmp"),
    ]
    patterns = [
        "00. Oficjalne ceny zakupu dla sprzed*.xlsx",
        "*Oficjalne ceny zakupu*.xlsx",
    ]
    for base in candidates:
        if not base.exists():
            continue
        for pattern in patterns:
            matches = sorted(base.glob(pattern), key=lambda p: p.stat().st_mtime, reverse=True)
            if matches:
                return matches[0]
    return None


def _load_google_sku_to_isk_rows() -> list[dict[str, str]]:
    now = time.time()
    if now - float(_SKU_TO_ISK_CACHE["loaded_at"]) < _SUGGESTION_CACHE_TTL_SECONDS:
        return list(_SKU_TO_ISK_CACHE["rows"])

    rows: list[dict[str, str]] = []
    try:
        conn = _connect()
        try:
            cur = conn.cursor()
            cur.execute(
                """
                IF OBJECT_ID('dbo.acc_amazon_listing_registry', 'U') IS NOT NULL
                BEGIN
                    SELECT DISTINCT
                        merchant_sku,
                        merchant_sku_alt,
                        internal_sku,
                        ean,
                        asin,
                        product_name
                    FROM dbo.acc_amazon_listing_registry WITH (NOLOCK)
                    WHERE ISNULL(internal_sku, '') <> ''
                END
                """
            )
            if cur.description:
                for row in cur.fetchall():
                    rows.append({
                        "merchant_sku": _norm_text(row[0]),
                        "merchant_sku_alt": _norm_text(row[1]),
                        "internal_sku": _norm_internal_sku(row[2]),
                        "ean": _norm_text(row[3]),
                        "asin": _norm_text(row[4]),
                        "title": _norm_text(row[5]),
                    })
        finally:
            conn.close()

        if not rows:
            with urlopen(_GOOGLE_SHEET_URL, timeout=20) as response:
                payload = response.read().decode("utf-8-sig", errors="replace")
            reader = csv.DictReader(io.StringIO(payload))
            for row in reader:
                merchant_sku = _norm_text(row.get("Merchant SKU"))
                nr_art = _norm_internal_sku(row.get("Nr art."))
                ean = _norm_text(row.get("EAN"))
                asin = _norm_text(row.get("ASIN (ADSY)"))
                title = _norm_text(row.get("Nazwa"))
                if not nr_art:
                    continue
                rows.append({
                    "merchant_sku": merchant_sku,
                    "merchant_sku_alt": (
                        merchant_sku.replace("MAG_", "FBA_", 1) if merchant_sku.startswith("MAG_")
                        else merchant_sku.replace("FBA_", "MAG_", 1) if merchant_sku.startswith("FBA_")
                        else ""
                    ),
                    "internal_sku": nr_art,
                    "ean": ean,
                    "asin": asin,
                    "title": title,
                })
    except Exception as exc:
        log.warning("profit_engine.google_sku_map_unavailable", error=str(exc))
        rows = []

    _SKU_TO_ISK_CACHE["loaded_at"] = now
    _SKU_TO_ISK_CACHE["rows"] = rows
    return rows


def _load_official_price_map() -> tuple[dict[str, float], dict[str, str]]:
    now = time.time()
    if now - float(_OFFICIAL_PRICE_CACHE["loaded_at"]) < _SUGGESTION_CACHE_TTL_SECONDS:
        return dict(_OFFICIAL_PRICE_CACHE["prices"]), dict(_OFFICIAL_PRICE_CACHE["titles"])

    prices: dict[str, float] = {}
    titles: dict[str, str] = {}
    path = _find_official_price_workbook()
    if not path:
        log.warning("profit_engine.official_price_workbook_missing")
        return prices, titles

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        header_row_idx = None
        header_map: dict[str, int] = {}
        for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
            values = [_norm_text(v) for v in row]
            if "Nr artykulu" in values and "Oficialna cena" in values:
                header_row_idx = idx
                header_map = {values[i]: i for i in range(len(values))}
                break
        if header_row_idx is None:
            raise ValueError("Official price workbook header row not found")

        nr_idx = header_map["Nr artykulu"]
        price_idx = header_map["Oficialna cena"]
        name_idx = header_map.get("Nazwa")

        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            internal_sku = _norm_internal_sku(row[nr_idx] if nr_idx < len(row) else None)
            if not internal_sku:
                continue
            raw_price = row[price_idx] if price_idx < len(row) else None
            try:
                price = round(float(raw_price), 4)
            except (TypeError, ValueError):
                continue
            if not (0 < price <= 2000):
                continue
            prices[internal_sku] = price
            if name_idx is not None and name_idx < len(row):
                title = _norm_text(row[name_idx])
                if title:
                    titles[internal_sku] = title
    except Exception as exc:
        log.warning("profit_engine.official_price_workbook_unavailable", path=str(path), error=str(exc))
        prices = {}
        titles = {}

    _OFFICIAL_PRICE_CACHE["loaded_at"] = now
    _OFFICIAL_PRICE_CACHE["prices"] = prices
    _OFFICIAL_PRICE_CACHE["titles"] = titles
    return prices, titles


# ---------------------------------------------------------------------------
# FX rate helper — uses central FX service for estimation
# ---------------------------------------------------------------------------


def _fx_case(currency_col: str = "o.currency") -> str:
    """Build SQL CASE expression for FX fallback using DB-sourced rates."""
    from app.core.fx_service import build_fx_case_sql
    return build_fx_case_sql(currency_col)


# ---------------------------------------------------------------------------
# FX rate cache — loaded from acc_exchange_rate (only ~3K rows)
# Eliminates costly OUTER APPLY correlated subquery in profit queries.
# ---------------------------------------------------------------------------
_FX_CACHE: dict[str, Any] = {"loaded_at": 0.0, "rates": {}}
_FX_CACHE_TTL = 3600  # 1 hour


def _load_fx_cache() -> dict[str, list[tuple[str, float]]]:
    """Load all exchange rates into memory, grouped by currency.

    Returns dict: currency → [(rate_date_str, rate_to_pln), ...] sorted DESC by date.
    """
    now = time.monotonic()
    if now - _FX_CACHE["loaded_at"] < _FX_CACHE_TTL and _FX_CACHE["rates"]:
        return _FX_CACHE["rates"]

    conn = _connect()
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT currency, CAST(rate_date AS DATE) AS rate_date, rate_to_pln
            FROM dbo.acc_exchange_rate WITH (NOLOCK)
            ORDER BY currency, rate_date DESC
        """)
        rates: dict[str, list[tuple[str, float]]] = defaultdict(list)
        for row in cur.fetchall():
            rates[row[0]].append((str(row[1]), float(row[2])))
        _FX_CACHE["rates"] = dict(rates)
        _FX_CACHE["loaded_at"] = now
        log.info("fx_cache.loaded", currencies=len(rates),
                 total_rates=sum(len(v) for v in rates.values()))
        return _FX_CACHE["rates"]
    finally:
        conn.close()


def _fx_rate_sql_fragment() -> str:
    """Return SQL fragment for FX rate lookup via OUTER APPLY.

    This is still used for complex queries where removing the OUTER APPLY
    would require rewriting the entire query.
    The CAST is applied to the junction, not the indexed column.
    """
    return """
        OUTER APPLY (
            SELECT TOP 1 rate_to_pln
            FROM dbo.acc_exchange_rate er WITH (NOLOCK)
            WHERE er.currency = o.currency
              AND er.rate_date <= o.purchase_date
            ORDER BY er.rate_date DESC
        ) fx
    """


def _parse_csv_list(raw: str | None) -> list[str]:
    if not raw:
        return []
    return [part.strip() for part in str(raw).split(",") if part and part.strip()]


def _parse_search_tokens(raw: str | None, *, limit: int = 12) -> list[str]:
    """Split multi-line/comma search text into safe tokens."""
    if not raw:
        return []
    tokens = [t.strip() for t in re.split(r"[\s,;]+", str(raw)) if t and t.strip()]
    # Keep query complexity bounded.
    return tokens[: max(1, int(limit))]


def _fx_rate_for_currency(currency: str | None, as_of: date | None = None) -> float:
    curr = (currency or "PLN").upper()
    if curr == "PLN":
        return 1.0

    rates = _load_fx_cache().get(curr, [])
    if not rates:
        from app.core.fx_service import get_rate_safe
        return get_rate_safe(curr, as_of)

    target = (as_of or date.today()).isoformat()
    for rate_date, value in rates:
        if str(rate_date) <= target:
            try:
                return float(value)
            except Exception:
                break
    from app.core.fx_service import get_rate_safe
    return get_rate_safe(curr, as_of)


def _choose_bucket_value(bucket_map: dict[int, dict[str, float]], qty: int, key: str) -> float:
    if not bucket_map:
        return 0.0
    if qty in bucket_map:
        return _f(bucket_map[qty].get(key))
    if 1 in bucket_map:
        return _f(bucket_map[1].get(key))
    nearest = sorted(bucket_map.keys(), key=lambda x: abs(int(x) - int(qty)))[0]
    return _f(bucket_map[nearest].get(key))


def _choose_bucket_payload(bucket_map: dict[int, dict[str, float]], qty: int) -> dict[str, float]:
    if not bucket_map:
        return {}
    if qty in bucket_map:
        return bucket_map[qty]
    if 1 in bucket_map:
        return bucket_map[1]
    nearest = sorted(bucket_map.keys(), key=lambda x: abs(int(x) - int(qty)))[0]
    return bucket_map.get(nearest, {})


def _suggest_pack_qty(bucket_map: dict[int, dict[str, float]]) -> tuple[int, str]:
    if not bucket_map:
        return 1, "default"

    baseline = bucket_map.get(1)
    if baseline and _f(baseline.get("median")) > 0:
        baseline_cost = _f(baseline.get("median"))
        suggested = 1
        for qty, payload in bucket_map.items():
            if int(qty) <= 1:
                continue
            samples = _i(payload.get("samples"))
            median = _f(payload.get("median"))
            if samples >= 3 and median > 0 and median <= baseline_cost * 1.15:
                suggested = max(suggested, int(qty))
        return max(1, suggested), "historical_bucket_rule"

    candidates = [(int(q), _i(v.get("samples"))) for q, v in bucket_map.items() if int(q) > 1]
    if candidates:
        qty, samples = max(candidates, key=lambda x: (x[1], x[0]))
        if samples >= 5:
            return max(1, qty), "historical_mode"
    return 1, "default"


_SUGGESTION_CACHE_TTL_SECONDS = 15 * 60
_SKU_TO_ISK_CACHE: dict[str, Any] = {"loaded_at": 0.0, "rows": []}
_OFFICIAL_PRICE_CACHE: dict[str, Any] = {"loaded_at": 0.0, "prices": {}, "titles": {}}
_GOOGLE_SHEET_URL = (
    "https://docs.google.com/spreadsheets/d/"
    "1rRBVZUTwqYcGYZRSp28mIWXw7gMfvqes0apEE_hdpjo/export?format=csv&gid=400534387"
)

_TKL_CACHE: dict[str, Any] = {"loaded_at": 0.0, "country_cost": {}, "sku_cost": {}}
_TKL_CACHE_TTL_SECONDS = 60 * 60
_TKL_CACHE_KEY = "default"
_TKL_CACHE_LOCK = threading.Lock()
_TKL_CACHE_SCHEMA_READY = False
_OFFER_FEE_EXPECTED_SCHEMA_READY = False

# What-if logistics decision model (production heuristics).
_WHATIF_LOGISTICS_MIN_SAMPLE = 5
_WHATIF_LOGISTICS_BLEND_SAMPLE = 15
_WHATIF_LOGISTICS_STABLE_P75_RATIO_MAX = 1.35
_WHATIF_LOGISTICS_BLEND_TKL_WEIGHT = 0.60
_WHATIF_LOGISTICS_BLEND_OBS_WEIGHT = 0.40
_WHATIF_LOGISTICS_DRIFT_SAMPLE = 10
_WHATIF_LOGISTICS_DRIFT_MEDIAN_RATIO = 1.10
_WHATIF_LOGISTICS_DRIFT_P75_RATIO = 1.20


def _parse_tkl_number(v: Any) -> float:
    txt = str(v or "").strip().replace(" ", "").replace(",", ".")
    if not txt:
        return 0.0
    try:
        return float(txt)
    except Exception:
        return 0.0


def _extract_pack_qty_from_name(name: str | None) -> int:
    raw = str(name or "")
    if not raw:
        return 0
    import re

    bracket_nums = [int(x) for x in re.findall(r"<\s*(\d{1,3})\s*>", raw)]
    # Prefer small realistic carton quantities.
    for n in bracket_nums:
        if 1 <= n <= 20:
            return n
    return 0


def _find_latest_tkl_file(pattern: str) -> Path | None:
    base = Path(r"N:\Analityka\TKL")
    if not base.exists():
        return None
    candidates = sorted(base.glob(pattern), key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0] if candidates else None


def ensure_profit_tkl_cache_schema() -> None:
    """No-op — schema managed by Alembic migration eb025."""


def _tkl_file_metadata(path: Path | None) -> tuple[str | None, datetime | None, str]:
    if not path:
        return None, None, "none"
    try:
        stat = path.stat()
        return str(path), datetime.utcfromtimestamp(stat.st_mtime), f"{path}|{int(stat.st_mtime_ns)}|{int(stat.st_size)}"
    except Exception:
        # SF-14: explicitly warn when logistics cost files are missing
        log.warning("profit_engine.tkl_file_missing", path=str(path),
                    msg="TKL logistics cost file unreadable — logistics costs will be zero for affected SKUs")
        return str(path), None, f"{path}|missing"


def _tkl_signature(path_courier: Path | None, path_tkl: Path | None) -> str:
    _, _, courier_sig = _tkl_file_metadata(path_courier)
    _, _, tkl_sig = _tkl_file_metadata(path_tkl)
    return f"courier={courier_sig};tkl={tkl_sig}"


def _load_tkl_maps_from_sql(
    *,
    signature: str | None = None,
    allow_any_signature: bool = False,
) -> tuple[dict[tuple[str, str], dict[str, Any]], dict[str, dict[str, Any]], dict[str, Any]] | None:
    ensure_profit_tkl_cache_schema()
    conn = _connect()
    try:
        cur = conn.cursor()
        cur.execute(
            """
            SELECT signature, source_courier_path, source_tkl_path, loaded_at
            FROM dbo.acc_tkl_cache_meta WITH (NOLOCK)
            WHERE cache_key = ?
            """,
            _TKL_CACHE_KEY,
        )
        meta_row = cur.fetchone()
        if not meta_row:
            return None
        cached_signature = _norm_text(meta_row[0])
        if signature and cached_signature and cached_signature != signature and not allow_any_signature:
            return None
        cur.execute(
            """
            SELECT row_type, internal_sku, country_code, cost, courier, source, pack_qty, [rank]
            FROM dbo.acc_tkl_cache_rows WITH (NOLOCK)
            WHERE cache_key = ?
            """,
            _TKL_CACHE_KEY,
        )
        rows = cur.fetchall()
        if not rows:
            return None

        country_cost: dict[tuple[str, str], dict[str, Any]] = {}
        sku_cost: dict[str, dict[str, Any]] = {}
        for row in rows:
            row_type = _norm_text(row[0]).lower()
            internal = _norm_internal_sku(row[1])
            if not internal:
                continue
            country = _norm_text(row[2]).upper()
            payload = {
                "cost": round(_f(row[3]), 4),
                "courier": _norm_text(row[4]),
                "source": _norm_text(row[5]) or "tkl:sql_cache",
                "pack_qty": max(0, _i(row[6])),
                "rank": _i(row[7]),
            }
            if row_type == "country":
                if not country:
                    continue
                country_cost[(internal, country)] = payload
            elif row_type == "sku":
                sku_cost[internal] = payload

        meta = {
            "signature": cached_signature,
            "source_courier_path": _norm_text(meta_row[1]),
            "source_tkl_path": _norm_text(meta_row[2]),
            "loaded_at": meta_row[3],
        }
        return country_cost, sku_cost, meta
    except Exception as exc:
        log.warning("profit_engine.tkl_cache_sql_read_error", error=str(exc))
        return None
    finally:
        conn.close()


def _save_tkl_maps_to_sql(
    *,
    path_courier: Path | None,
    path_tkl: Path | None,
    signature: str,
    country_cost: dict[tuple[str, str], dict[str, Any]],
    sku_cost: dict[str, dict[str, Any]],
) -> None:
    ensure_profit_tkl_cache_schema()
    conn = _connect()
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM dbo.acc_tkl_cache_rows WHERE cache_key = ?", _TKL_CACHE_KEY)

        rows: list[tuple[Any, ...]] = []
        for (internal, country), payload in country_cost.items():
            rows.append(
                (
                    _TKL_CACHE_KEY,
                    "country",
                    internal,
                    country,
                    round(_f(payload.get("cost")), 4),
                    _norm_text(payload.get("courier")),
                    _norm_text(payload.get("source")),
                    max(0, _i(payload.get("pack_qty"))),
                    _i(payload.get("rank")),
                )
            )
        for internal, payload in sku_cost.items():
            rows.append(
                (
                    _TKL_CACHE_KEY,
                    "sku",
                    internal,
                    None,
                    round(_f(payload.get("cost")), 4),
                    _norm_text(payload.get("courier")),
                    _norm_text(payload.get("source")),
                    max(0, _i(payload.get("pack_qty"))),
                    _i(payload.get("rank")),
                )
            )
        if rows:
            cur.fast_executemany = True
            cur.executemany(
                """
                INSERT INTO dbo.acc_tkl_cache_rows
                (cache_key, row_type, internal_sku, country_code, cost, courier, source, pack_qty, [rank])
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                rows,
            )

        courier_path, courier_mtime, _ = _tkl_file_metadata(path_courier)
        tkl_path, tkl_mtime, _ = _tkl_file_metadata(path_tkl)
        cur.execute(
            """
            MERGE dbo.acc_tkl_cache_meta AS tgt
            USING (
                SELECT
                    ? AS cache_key,
                    ? AS signature,
                    ? AS source_courier_path,
                    ? AS source_courier_mtime,
                    ? AS source_tkl_path,
                    ? AS source_tkl_mtime
            ) AS src
            ON tgt.cache_key = src.cache_key
            WHEN MATCHED THEN
                UPDATE SET
                    signature = src.signature,
                    source_courier_path = src.source_courier_path,
                    source_courier_mtime = src.source_courier_mtime,
                    source_tkl_path = src.source_tkl_path,
                    source_tkl_mtime = src.source_tkl_mtime,
                    loaded_at = SYSUTCDATETIME()
            WHEN NOT MATCHED THEN
                INSERT (cache_key, signature, source_courier_path, source_courier_mtime, source_tkl_path, source_tkl_mtime, loaded_at)
                VALUES (src.cache_key, src.signature, src.source_courier_path, src.source_courier_mtime, src.source_tkl_path, src.source_tkl_mtime, SYSUTCDATETIME());
            """,
            _TKL_CACHE_KEY,
            signature,
            courier_path,
            courier_mtime,
            tkl_path,
            tkl_mtime,
        )
        conn.commit()
    except Exception as exc:
        log.warning("profit_engine.tkl_cache_sql_write_error", error=str(exc))
    finally:
        conn.close()


def _parse_tkl_priority_maps_from_files(
    *,
    path_courier: Path | None,
    path_tkl: Path | None,
) -> tuple[dict[tuple[str, str], dict[str, Any]], dict[str, dict[str, Any]]]:

    country_cost: dict[tuple[str, str], dict[str, Any]] = {}
    sku_cost: dict[str, dict[str, Any]] = {}

    # 1) Country+SKU costs from "00. Wyliczanie Kurierów 2.0.xlsx"
    # Priority: ZMIANA > NOWI > base sheet.
    if path_courier:
        try:
            wb = load_workbook(path_courier, read_only=True, data_only=True)
            sheet_priority = [
                ("Kurierzy do wgrania ZMIANA", 3),
                ("Kurierzy do wgrania NOWI", 2),
                ("Kurierzy do wgrania", 1),
            ]
            for sheet_name, rank in sheet_priority:
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                header_row = None
                for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
                    vals = [_norm_text(x) for x in row]
                    if "Symbol_Kod" in vals and "Ceny.Kod KRAJU" in vals:
                        header_row = vals
                        break
                if not header_row:
                    continue
                idx = {header_row[i]: i for i in range(len(header_row))}
                for row in ws.iter_rows(min_row=2, values_only=True):
                    country = _norm_text(row[idx.get("Ceny.Kod KRAJU", -1)] if idx.get("Ceny.Kod KRAJU", -1) >= 0 else "")
                    internal = _norm_internal_sku(row[idx.get("Symbol_Kod", -1)] if idx.get("Symbol_Kod", -1) >= 0 else "")
                    if not country or not internal:
                        continue
                    price_a = _parse_tkl_number(row[idx.get("Wybrany kurier", -1)] if idx.get("Wybrany kurier", -1) >= 0 else 0)
                    price_b = _parse_tkl_number(row[idx.get("Cena obecny kuriera", -1)] if idx.get("Cena obecny kuriera", -1) >= 0 else 0)
                    cost = price_a if price_a > 0 else price_b
                    if cost <= 0:
                        continue
                    courier = _norm_text(row[idx.get("Kurier", -1)] if idx.get("Kurier", -1) >= 0 else "")
                    title = _norm_text(row[idx.get("Nazwa", -1)] if idx.get("Nazwa", -1) >= 0 else "")
                    pack_qty = _extract_pack_qty_from_name(title)
                    key = (internal, country.upper())
                    existing = country_cost.get(key)
                    if not existing or rank > _i(existing.get("rank")):
                        country_cost[key] = {
                            "cost": round(cost, 4),
                            "courier": courier,
                            "source": f"tkl:{sheet_name}",
                            "pack_qty": max(0, pack_qty),
                            "rank": rank,
                        }
        except Exception as exc:
            log.warning("profit_engine.tkl_courier_parse_error", path=str(path_courier), error=str(exc))

    # 2) Generic SKU-level fallback from "00. Tabela Kosztów Logistycznych.xlsx" / Lista TKL
    if path_tkl:
        try:
            wb = load_workbook(path_tkl, read_only=True, data_only=True)
            if "Lista TKL" in wb.sheetnames:
                ws = wb["Lista TKL"]
                header = None
                for row in ws.iter_rows(min_row=1, max_row=4, values_only=True):
                    vals = [_norm_text(x) for x in row]
                    if "Nr artykułu" in vals and "KOSZT TKL" in vals:
                        header = vals
                        break
                if header:
                    idx = {header[i]: i for i in range(len(header))}
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        internal = _norm_internal_sku(row[idx.get("Nr artykułu", -1)] if idx.get("Nr artykułu", -1) >= 0 else "")
                        if not internal:
                            continue
                        cost = _parse_tkl_number(row[idx.get("KOSZT TKL", -1)] if idx.get("KOSZT TKL", -1) >= 0 else 0)
                        if cost <= 0:
                            continue
                        title = _norm_text(row[idx.get("Nazwa", -1)] if idx.get("Nazwa", -1) >= 0 else "")
                        pack_qty = _extract_pack_qty_from_name(title)
                        existing = sku_cost.get(internal)
                        if not existing:
                            sku_cost[internal] = {
                                "cost": round(cost, 4),
                                "source": "tkl:Lista TKL",
                                "pack_qty": max(0, pack_qty),
                            }
        except Exception as exc:
            log.warning("profit_engine.tkl_lista_parse_error", path=str(path_tkl), error=str(exc))

    return country_cost, sku_cost


def refresh_tkl_sql_cache(*, force: bool = False) -> dict[str, Any]:
    """Refresh TKL SQL cache from XLSX source files (manual/system trigger)."""
    if force:
        _TKL_CACHE["loaded_at"] = 0.0
    country_cost, sku_cost = _load_tkl_priority_maps()
    return {
        "country_pairs": len(country_cost),
        "sku_rows": len(sku_cost),
        "cache_loaded_at": _TKL_CACHE.get("loaded_at"),
    }


def _load_tkl_priority_maps() -> tuple[dict[tuple[str, str], dict[str, Any]], dict[str, dict[str, Any]]]:
    now = time.time()
    if now - float(_TKL_CACHE["loaded_at"]) < _TKL_CACHE_TTL_SECONDS:
        return dict(_TKL_CACHE["country_cost"]), dict(_TKL_CACHE["sku_cost"])

    with _TKL_CACHE_LOCK:
        now = time.time()
        if now - float(_TKL_CACHE["loaded_at"]) < _TKL_CACHE_TTL_SECONDS:
            return dict(_TKL_CACHE["country_cost"]), dict(_TKL_CACHE["sku_cost"])

        path_courier = _find_latest_tkl_file("00. Wyliczanie Kurier*2.0.xlsx")
        path_tkl = _find_latest_tkl_file("00. Tabela Koszt*Logistycznych.xlsx")
        signature = _tkl_signature(path_courier, path_tkl)

        # Fast path: SQL cache with matching source signature.
        sql_match = _load_tkl_maps_from_sql(signature=signature, allow_any_signature=False)
        if sql_match:
            country_cost, sku_cost, meta = sql_match
            _TKL_CACHE["loaded_at"] = now
            _TKL_CACHE["country_cost"] = country_cost
            _TKL_CACHE["sku_cost"] = sku_cost
            log.info(
                "profit_engine.tkl_loaded",
                source="sql_cache_signature_match",
                country_pairs=len(country_cost),
                sku_rows=len(sku_cost),
                signature=meta.get("signature"),
            )
            return dict(country_cost), dict(sku_cost)

        # Parse files only if needed.
        country_cost, sku_cost = _parse_tkl_priority_maps_from_files(
            path_courier=path_courier,
            path_tkl=path_tkl,
        )
        source = "file_parse"
        if country_cost or sku_cost:
            _save_tkl_maps_to_sql(
                path_courier=path_courier,
                path_tkl=path_tkl,
                signature=signature,
                country_cost=country_cost,
                sku_cost=sku_cost,
            )
            source = "file_parse_and_sql_refresh"
        else:
            # If source files are temporarily unavailable, prefer stale SQL cache over empty results.
            sql_fallback = _load_tkl_maps_from_sql(signature=None, allow_any_signature=True)
            if sql_fallback:
                country_cost, sku_cost, _ = sql_fallback
                source = "sql_cache_stale_fallback"
                # SF-14: warn that stale data is being used
                log.warning("profit_engine.tkl_stale_cache",
                            msg="TKL source files unavailable — using stale SQL cache. Logistics costs may be outdated.")
            else:
                # SF-14: no data at all — all logistics will be zero
                log.error("profit_engine.tkl_no_data",
                          msg="No TKL data available (files + SQL cache empty). All logistics costs will be ZERO.")
                source = "empty_no_source"

        _TKL_CACHE["loaded_at"] = now
        _TKL_CACHE["country_cost"] = country_cost
        _TKL_CACHE["sku_cost"] = sku_cost
        log.info(
            "profit_engine.tkl_loaded",
            source=source,
            country_pairs=len(country_cost),
            sku_rows=len(sku_cost),
            path_courier=str(path_courier) if path_courier else None,
            path_tkl=str(path_tkl) if path_tkl else None,
            signature=signature,
        )
        return dict(country_cost), dict(sku_cost)

# ---------------------------------------------------------------------------
# In-memory result cache for heavy queries (profit table, data quality)
# ---------------------------------------------------------------------------
_RESULT_CACHE: dict[str, tuple[float, Any]] = {}
_RESULT_CACHE_TTL = 180  # 3 minutes
_RESULT_CACHE_MAX = 50



# ---------------------------------------------------------------------------
# Schemas
# ---------------------------------------------------------------------------

def ensure_profit_data_quality_schema() -> None:
    """No-op — schema managed by Alembic migration eb023."""


# ---------------------------------------------------------------------------
# Profit cost model schema/helpers (CM1/CM2/NP explicit components)
# ---------------------------------------------------------------------------

def ensure_profit_cost_model_schema() -> None:
    """No-op — schema managed by Alembic migration eb024."""


def _get_cost_config_decimal(cur: pyodbc.Cursor, key: str, default: float = 0.0) -> float:
    try:
        cur.execute(
            """
            SELECT TOP 1 value_decimal
            FROM dbo.acc_profit_cost_config WITH (NOLOCK)
            WHERE config_key = ?
            """,
            [key],
        )
        row = cur.fetchone()
        return _f(row[0], default) if row else default
    except Exception:
        return default



# ---------------------------------------------------------------------------
# Fee classification wrappers
# ---------------------------------------------------------------------------

def _classify_finance_charge(charge_type: Any, transaction_type: Any) -> dict | None:
    """Classify a finance transaction charge_type into a P&L bucket.

    Delegates to the unified fee taxonomy (core/fee_taxonomy.py) so that
    the profit engine and the finance ledger always agree.

    Returns dict with keys:
        layer: "cm2" or "np"
        bucket: specific cost bucket name
        sign: 1 (cost, increases expense) or -1 (recovery, reduces expense)

    Returns None when the charge should be skipped (revenue, CM1 fees,
    cash-flow timing entries).
    """
    from app.core.fee_taxonomy import get_profit_classification
    return get_profit_classification(charge_type, transaction_type)


def _classify_fba_component(charge_type: Any, transaction_type: Any) -> str | None:
    """Legacy wrapper — maps to CM2 FBA buckets for existing allocation code."""
    result = _classify_finance_charge(charge_type, transaction_type)
    if result is None:
        return None
    bucket = result["bucket"]
    if bucket == "fba_storage":
        return "storage"
    if bucket == "fba_aged":
        return "aged"
    if bucket == "fba_removal":
        return "removal"
    if bucket == "fba_liquidation":
        return "liquidation"
    return None



# ---------------------------------------------------------------------------
# Price source helpers & suggestions
# ---------------------------------------------------------------------------

_PRICE_SOURCE_PRIORITY_CASE = """
CASE {alias}.source
    WHEN 'manual'         THEN 1
    WHEN 'import_xlsx'    THEN 2
    WHEN 'xlsx_oficjalne' THEN 3
    WHEN 'holding'        THEN 4
    WHEN 'erp_holding'    THEN 5
    WHEN 'import_csv'     THEN 6
    WHEN 'cogs_xlsx'      THEN 7
    WHEN 'acc_product'    THEN 8
    WHEN 'ai_match'       THEN 9
    ELSE 99
END
"""


def _price_source_label(source: str | None) -> str | None:
    labels = {
        "manual": "Ręczna cena zakupu",
        "import_xlsx": "Import XLSX",
        "xlsx_oficjalne": "Oficjalny XLSX",
        "xlsx_oficjalne_live": "Oficjalny XLSX (live)",
        "holding": "Holding",
        "erp_holding": "ERP Holding",
        "import_csv": "Import CSV",
        "cogs_xlsx": "COGS XLSX",
        "acc_product": "Kartoteka produktu",
        "ai_match": "AI match",
    }
    return labels.get(str(source or "").strip().lower(), source)


def _lookup_best_price_for_internal_sku(cur, internal_sku: str | None) -> dict[str, Any] | None:
    if not internal_sku:
        return None

    cur.execute(
        f"""
        SELECT TOP 1
            CAST(pp.netto_price_pln AS FLOAT) AS price_pln,
            pp.source
        FROM dbo.acc_purchase_price pp WITH (NOLOCK)
        WHERE pp.internal_sku = ?
          AND pp.netto_price_pln > 0
          AND pp.netto_price_pln <= 2000
        ORDER BY {_PRICE_SOURCE_PRIORITY_CASE.format(alias='pp')},
                 pp.valid_from DESC,
                 pp.updated_at DESC
        """,
        [internal_sku],
    )
    row = cur.fetchone()
    if row:
        return {
            "price_pln": float(row[0]),
            "source": row[1],
            "source_label": _price_source_label(row[1]),
        }

    cur.execute(
        """
        SELECT TOP 1 CAST(p.netto_purchase_price_pln AS FLOAT) AS price_pln
        FROM dbo.acc_product p WITH (NOLOCK)
        WHERE p.internal_sku = ?
          AND p.netto_purchase_price_pln IS NOT NULL
          AND p.netto_purchase_price_pln > 0
          AND p.netto_purchase_price_pln <= 2000
        """,
        [internal_sku],
    )
    row = cur.fetchone()
    if row:
        return {
            "price_pln": float(row[0]),
            "source": "acc_product",
            "source_label": _price_source_label("acc_product"),
        }

    cur.execute(
        """
        SELECT TOP 1 CAST(ol.purchase_price_pln AS FLOAT) AS price_pln
        FROM dbo.acc_order_line ol WITH (NOLOCK)
        INNER JOIN dbo.acc_order o WITH (NOLOCK) ON o.id = ol.order_id
        INNER JOIN dbo.acc_product p WITH (NOLOCK) ON p.id = ol.product_id
        WHERE p.internal_sku = ?
          AND ol.purchase_price_pln IS NOT NULL
          AND ol.purchase_price_pln > 0
          AND ol.purchase_price_pln <= 2000
          AND ISNULL(ol.quantity_ordered, 0) > 0
          AND ISNULL(o.sales_channel, 'Amazon.com') != 'Non-Amazon'
          AND o.amazon_order_id NOT LIKE 'S02-%'
        ORDER BY o.purchase_date DESC
        """,
        [internal_sku],
    )
    row = cur.fetchone()
    if row:
        return {
            "price_pln": float(row[0]),
            "source": "same_internal_sku_history",
            "source_label": "Historia tego internal_sku",
        }

    official_prices, official_titles = _load_official_price_map()
    official_price = official_prices.get(str(internal_sku).strip())
    if official_price is not None:
        return {
            "price_pln": float(official_price),
            "source": "xlsx_oficjalne_live",
            "source_label": _price_source_label("xlsx_oficjalne_live"),
            "title": official_titles.get(str(internal_sku).strip()),
        }
    return None


def _find_unique_candidate_by_field(
    cur,
    *,
    field_name: str,
    field_value: str | None,
    exclude_internal_sku: str | None = None,
) -> dict[str, Any] | None:
    if field_name not in {"asin", "ean"} or not field_value:
        return None

    cur.execute(
        f"""
        SELECT DISTINCT TOP 3 p.internal_sku
        FROM dbo.acc_product p WITH (NOLOCK)
        WHERE p.{field_name} = ?
          AND p.internal_sku IS NOT NULL
          AND p.internal_sku <> ''
          AND (? IS NULL OR p.internal_sku <> ?)
        ORDER BY p.internal_sku
        """,
        [field_value, exclude_internal_sku, exclude_internal_sku],
    )
    unique_isks = sorted({str(row[0]).strip() for row in cur.fetchall() if row and row[0]})
    if len(unique_isks) != 1:
        return None

    suggested_internal_sku = unique_isks[0]
    price = _lookup_best_price_for_internal_sku(cur, suggested_internal_sku)

    cur.execute(
        """
        SELECT TOP 1 p.title
        FROM dbo.acc_product p WITH (NOLOCK)
        WHERE p.internal_sku = ?
          AND p.title IS NOT NULL
          AND p.title <> ''
        ORDER BY p.sku
        """,
        [suggested_internal_sku],
    )
    title_row = cur.fetchone()

    return {
        "suggested_internal_sku": suggested_internal_sku,
        "suggested_price_pln": price["price_pln"] if price else None,
        "source_type": f"{field_name}_exact_unique",
        "source_label": "Exact EAN" if field_name == "ean" else "Exact ASIN",
        "note": title_row[0] if title_row else f"Jednoznaczny match po {field_name.upper()}",
        "is_hard_source": True,
    }


def _lookup_ai_candidate(cur, sku: str) -> dict[str, Any] | None:
    cur.execute(
        """
        SELECT TOP 1
            matched_internal_sku,
            matched_title,
            confidence,
            reasoning
        FROM dbo.acc_product_match_suggestion WITH (NOLOCK)
        WHERE unmapped_sku = ?
          AND matched_internal_sku IS NOT NULL
          AND status IN ('pending', 'approved')
        ORDER BY
            CASE status WHEN 'approved' THEN 1 ELSE 2 END,
            confidence DESC,
            created_at DESC
        """,
        [sku],
    )
    row = cur.fetchone()
    if not row:
        return None

    hard_price = _lookup_best_price_for_internal_sku(cur, str(row[0]))
    return {
        "matched_internal_sku": str(row[0]),
        "matched_title": row[1],
        "confidence": float(row[2] or 0),
        "reasoning": row[3],
        "hard_price_pln": hard_price["price_pln"] if hard_price else None,
        "hard_price_source": hard_price["source_label"] if hard_price else None,
    }


def _find_same_ean_sibling_suggestion(
    cur,
    *,
    sku: str,
    ean: str | None,
    exclude_internal_sku: str | None = None,
) -> dict[str, Any] | None:
    if not ean:
        return None

    cur.execute(
        """
        SELECT DISTINCT p.internal_sku
        FROM dbo.acc_product p WITH (NOLOCK)
        WHERE p.ean = ?
          AND p.sku <> ?
          AND p.internal_sku IS NOT NULL
          AND p.internal_sku <> ''
          AND (? IS NULL OR p.internal_sku <> ?)
        ORDER BY p.internal_sku
        """,
        [ean, sku, exclude_internal_sku, exclude_internal_sku],
    )
    sibling_isks = [str(row[0]).strip() for row in cur.fetchall() if row and row[0]]
    unique_isks = sorted(set(sibling_isks))
    if len(unique_isks) != 1:
        return None

    suggested_internal_sku = unique_isks[0]
    price = _lookup_best_price_for_internal_sku(cur, suggested_internal_sku)
    if not price:
        return None

    cur.execute(
        """
        SELECT TOP 1 p.title
        FROM dbo.acc_product p WITH (NOLOCK)
        WHERE p.internal_sku = ?
          AND p.ean = ?
          AND p.title IS NOT NULL
          AND p.title <> ''
        ORDER BY p.sku
        """,
        [suggested_internal_sku, ean],
    )
    title_row = cur.fetchone()

    return {
        "suggested_internal_sku": suggested_internal_sku,
        "suggested_price_pln": price["price_pln"],
        "source_type": "same_ean_sibling_variant",
        "source_label": "Sibling po tym samym EAN",
        "note": title_row[0] if title_row else "Jednoznaczny sibling po tym samym EAN",
        "is_hard_source": True,
    }


def _find_google_sheet_official_suggestion(
    cur,
    *,
    sku: str,
    asin: str | None,
    ean: str | None,
    exclude_internal_sku: str | None = None,
) -> dict[str, Any] | None:
    sheet_rows = _load_google_sku_to_isk_rows()
    sku_norm = _norm_text(sku)
    asin_norm = _norm_text(asin)
    ean_norm = _norm_text(ean)

    match = None
    for row in sheet_rows:
        if row["merchant_sku"] == sku_norm:
            match = row
            break
    if match is None:
        for row in sheet_rows:
            if row["merchant_sku_alt"] and row["merchant_sku_alt"] == sku_norm:
                match = row
                break
    if match is None and asin_norm:
        for row in sheet_rows:
            if row["asin"] and row["asin"] == asin_norm:
                match = row
                break
    if match is None and ean_norm:
        for row in sheet_rows:
            if row["ean"] and row["ean"] == ean_norm:
                match = row
                break
    if not match:
        return None

    suggested_internal_sku = match["internal_sku"]
    if exclude_internal_sku and suggested_internal_sku == exclude_internal_sku:
        return None

    price = _lookup_best_price_for_internal_sku(cur, suggested_internal_sku)
    if not price:
        return None

    return {
        "suggested_internal_sku": suggested_internal_sku,
        "suggested_price_pln": price["price_pln"],
        "source_type": "google_sheet_to_official_price",
        "source_label": "SKU -> ISK -> Oficjalny XLSX",
        "note": match.get("title") or "Mapowanie z arkusza Amazon + oficjalny cennik",
        "is_hard_source": True,
    }


def _build_missing_cogs_suggestions(
    cur,
    *,
    sku: str,
    asin: str | None,
    ean: str | None,
    internal_sku: str | None,
) -> tuple[dict[str, Any] | None, str | None, dict[str, Any] | None, dict[str, Any] | None]:
    current_price = _lookup_best_price_for_internal_sku(cur, internal_sku)
    hard_suggestion = None

    if internal_sku and current_price and current_price.get("source") == "xlsx_oficjalne_live":
        hard_suggestion = {
            "suggested_internal_sku": internal_sku,
            "suggested_price_pln": current_price["price_pln"],
            "source_type": "same_internal_sku_official_xlsx",
            "source_label": "Oficjalny XLSX",
            "note": current_price.get("title") or "Cena z oficjalnego cennika po internal_sku",
            "is_hard_source": True,
        }

    if not current_price and not hard_suggestion:
        hard_suggestion = _find_google_sheet_official_suggestion(
            cur,
            sku=sku,
            asin=asin,
            ean=ean,
            exclude_internal_sku=internal_sku,
        )

    if not current_price and not hard_suggestion:
        hard_suggestion = _find_unique_candidate_by_field(
            cur,
            field_name="ean",
            field_value=ean,
            exclude_internal_sku=internal_sku,
        )
        if not hard_suggestion:
            hard_suggestion = _find_unique_candidate_by_field(
                cur,
                field_name="asin",
                field_value=asin,
                exclude_internal_sku=internal_sku,
            )
        if not hard_suggestion:
            hard_suggestion = _find_same_ean_sibling_suggestion(
                cur,
                sku=sku,
                ean=ean,
                exclude_internal_sku=internal_sku,
            )

    ai_candidate = _lookup_ai_candidate(cur, sku)
    return (
        current_price,
        current_price["source_label"] if current_price else None,
        hard_suggestion,
        ai_candidate,
    )



# ---------------------------------------------------------------------------
# Purchase price management
# ---------------------------------------------------------------------------

def _apply_manual_price_to_internal_sku(
    cur,
    *,
    internal_sku: str,
    netto_price_pln: float,
) -> None:
    """Apply a manual purchase price immediately to product + order lines."""
    cur.execute(
        """
        UPDATE dbo.acc_product
        SET netto_purchase_price_pln = ?,
            updated_at = GETDATE()
        WHERE internal_sku = ?
        """,
        [netto_price_pln, internal_sku],
    )

    cur.execute(
        """
        UPDATE ol
        SET ol.purchase_price_pln = ?,
            ol.cogs_pln = ROUND(? * ISNULL(ol.quantity_ordered, 1), 4),
            ol.price_source = 'manual'
        FROM dbo.acc_order_line ol
        INNER JOIN dbo.acc_order o ON o.id = ol.order_id
        INNER JOIN dbo.acc_product p ON p.id = ol.product_id
        WHERE p.internal_sku = ?
          AND ISNULL(ol.quantity_ordered, 0) > 0
          AND ISNULL(o.sales_channel, 'Amazon.com') != 'Non-Amazon'
          AND o.amazon_order_id NOT LIKE 'S02-%'
        """,
        [netto_price_pln, netto_price_pln, internal_sku],
    )

    cur.execute(
        """
        UPDATE o
        SET o.cogs_pln = ISNULL(agg.total_cogs, 0)
        FROM dbo.acc_order o
        CROSS APPLY (
            SELECT SUM(ISNULL(ol.cogs_pln, 0)) AS total_cogs
            FROM dbo.acc_order_line ol WITH (NOLOCK)
            WHERE ol.order_id = o.id
        ) agg
        WHERE EXISTS (
            SELECT 1
            FROM dbo.acc_order_line ol2
            INNER JOIN dbo.acc_product p2 ON p2.id = ol2.product_id
            WHERE ol2.order_id = o.id
              AND p2.internal_sku = ?
        )
          AND ISNULL(o.sales_channel, 'Amazon.com') != 'Non-Amazon'
        """,
        [internal_sku],
    )

    manual_logistics_join_sql = profit_logistics_join_sql(order_alias="o", fact_alias="olf")
    manual_logistics_value_sql = profit_logistics_value_sql(order_alias="o", fact_alias="olf")
    cur.execute(
        f"""
        UPDATE o
        SET contribution_margin_pln = ROUND(
                ISNULL(revenue_pln, 0) - ISNULL(cogs_pln, 0)
                - ISNULL(amazon_fees_pln, 0) - ISNULL(ads_cost_pln, 0)
                - {manual_logistics_value_sql}, 2),
            cm_percent = CASE
                WHEN ISNULL(revenue_pln, 0) > 0 THEN
                    ROUND(
                        (ISNULL(revenue_pln, 0) - ISNULL(cogs_pln, 0)
                         - ISNULL(amazon_fees_pln, 0) - ISNULL(ads_cost_pln, 0)
                         - {manual_logistics_value_sql})
                        / NULLIF(revenue_pln, 0) * 100, 4)
                ELSE 0
            END
        FROM dbo.acc_order o
        {manual_logistics_join_sql}
        WHERE EXISTS (
            SELECT 1
            FROM dbo.acc_order_line ol2
            INNER JOIN dbo.acc_product p2 ON p2.id = ol2.product_id
            WHERE ol2.order_id = o.id
              AND p2.internal_sku = ?
        )
          AND ISNULL(o.sales_channel, 'Amazon.com') != 'Non-Amazon'
        """,
        [internal_sku],
    )


def upsert_purchase_price(
    *,
    internal_sku: str,
    netto_price_pln: float,
) -> dict[str, Any]:
    """Insert or update a manual purchase price for given internal_sku."""
    if netto_price_pln > 2000:
        raise ValueError(f"Price {netto_price_pln} PLN exceeds max cap (2000 PLN). Likely data error.")
    conn = _connect()
    try:
        cur = conn.cursor()
        cur.execute("SET LOCK_TIMEOUT 30000")

        # Check if a manual entry already exists
        cur.execute("""
            SELECT id FROM dbo.acc_purchase_price
            WHERE internal_sku = ? AND source = 'manual'
        """, [internal_sku])
        existing = cur.fetchone()

        now_str = date.today().isoformat()

        if existing:
            cur.execute("""
                UPDATE dbo.acc_purchase_price
                SET netto_price_pln = ?,
                    valid_from = ?,
                    updated_at = GETDATE()
                WHERE id = ?
            """, [netto_price_pln, now_str, existing[0]])
        else:
            cur.execute("""
                INSERT INTO dbo.acc_purchase_price
                    (internal_sku, netto_price_pln, valid_from, source, source_document, created_at, updated_at)
                VALUES (?, ?, ?, 'manual', 'data_quality_ui', GETDATE(), GETDATE())
            """, [internal_sku, netto_price_pln, now_str])

        _apply_manual_price_to_internal_sku(
            cur,
            internal_sku=internal_sku,
            netto_price_pln=netto_price_pln,
        )
        conn.commit()
        _result_cache_invalidate()
        return {
            "internal_sku": internal_sku,
            "netto_price_pln": netto_price_pln,
            "status": "updated" if existing else "created",
        }
    finally:
        conn.close()


def _resolve_manual_mapping_target(
    cur,
    *,
    sku: str,
    internal_sku: str,
) -> dict[str, Any] | None:
    """Resolve the best acc_product target for manual SKU mapping."""
    # 1. Exact product SKU if it already exists for this internal SKU.
    cur.execute(
        """
        SELECT TOP 1
            p.id,
            p.sku,
            p.internal_sku,
            p.asin
        FROM dbo.acc_product p WITH (NOLOCK)
        WHERE p.sku = ?
          AND p.internal_sku = ?
        ORDER BY p.updated_at DESC, p.created_at DESC
        """,
        [sku, internal_sku],
    )
    row = cur.fetchone()
    if row:
        return {
            "product_id": str(row[0]),
            "product_sku": row[1],
            "internal_sku": row[2],
            "asin": row[3],
            "source": "exact_sku",
        }

    # 2. Use the dominant ASIN seen on order lines for this SKU.
    cur.execute(
        """
        SELECT TOP 1
            ol.asin,
            COUNT(*) AS row_count
        FROM dbo.acc_order_line ol WITH (NOLOCK)
        INNER JOIN dbo.acc_order o WITH (NOLOCK) ON o.id = ol.order_id
        WHERE ol.sku = ?
          AND ol.asin IS NOT NULL
          AND ol.asin <> ''
          AND ISNULL(ol.quantity_ordered, 0) > 0
          AND ISNULL(o.sales_channel, 'Amazon.com') != 'Non-Amazon'
          AND o.amazon_order_id NOT LIKE 'S02-%'
        GROUP BY ol.asin
        ORDER BY COUNT(*) DESC, ol.asin
        """,
        [sku],
    )
    asin_row = cur.fetchone()
    dominant_asin = asin_row[0] if asin_row else None
    if dominant_asin:
        cur.execute(
            """
            SELECT TOP 1
                p.id,
                p.sku,
                p.internal_sku,
                p.asin
            FROM dbo.acc_product p WITH (NOLOCK)
            WHERE p.internal_sku = ?
              AND p.asin = ?
            ORDER BY p.updated_at DESC, p.created_at DESC
            """,
            [internal_sku, dominant_asin],
        )
        row = cur.fetchone()
        if row:
            return {
                "product_id": str(row[0]),
                "product_sku": row[1],
                "internal_sku": row[2],
                "asin": row[3],
                "source": "dominant_asin",
            }

    # 3. If there is exactly one product for this internal SKU, use it.
    cur.execute(
        """
        SELECT
            COUNT(*) AS cnt,
            MIN(CAST(id AS NVARCHAR(36))) AS any_id,
            MIN(sku) AS any_sku,
            MIN(asin) AS any_asin
        FROM dbo.acc_product WITH (NOLOCK)
        WHERE internal_sku = ?
        """,
        [internal_sku],
    )
    row = cur.fetchone()
    if row and int(row[0] or 0) == 1:
        return {
            "product_id": row[1],
            "product_sku": row[2],
            "internal_sku": internal_sku,
            "asin": row[3],
            "source": "unique_internal_sku",
        }
    return None


def map_and_price(
    *,
    sku: str,
    internal_sku: str,
    netto_price_pln: float,
) -> dict[str, Any]:
    """Map product SKU → internal_sku AND upsert purchase price in one call."""
    conn = _connect()
    try:
        cur = conn.cursor()
        cur.execute("SET LOCK_TIMEOUT 30000")

        # Step 1: Update internal_sku on acc_product for this SKU
        # --- Controlling: log mapping change ---
        try:
            from app.services.controlling import log_mapping_change
            cur.execute(
                "SELECT CAST(id AS VARCHAR(36)), internal_sku, mapping_source "
                "FROM dbo.acc_product WHERE sku = ?", [sku]
            )
            for prow in cur.fetchall():
                log_mapping_change(
                    conn,
                    product_id=str(prow[0]),
                    sku=sku,
                    old_internal_sku=str(prow[1]) if prow[1] else None,
                    new_internal_sku=internal_sku,
                    old_source=str(prow[2]) if prow[2] else None,
                    new_source="manual_dq",
                    change_type="set" if not prow[1] else "update",
                    reason="manual data quality override via map_and_price",
                )
        except Exception:
            pass  # controlling is non-blocking

        cur.execute("""
            UPDATE dbo.acc_product
            SET internal_sku = ?, mapping_source = 'manual_dq', updated_at = GETDATE()
            WHERE sku = ? AND (internal_sku IS NULL OR internal_sku = '')
        """, [internal_sku, sku])
        mapped_count = cur.rowcount

        # Step 1b: If order lines for this SKU are still unmapped, bind them to the
        # best existing product candidate resolved from exact SKU / dominant ASIN /
        # unique internal SKU. Without this, price restamping misses NULL product_id lines.
        binding_target = _resolve_manual_mapping_target(
            cur,
            sku=sku,
            internal_sku=internal_sku,
        )
        bound_lines = 0
        if binding_target:
            params: list[Any] = [binding_target["product_id"], sku]
            asin_filter_sql = ""
            if binding_target.get("asin"):
                asin_filter_sql = " AND ol.asin = ?"
                params.append(binding_target["asin"])
            cur.execute(
                f"""
                UPDATE ol
                SET ol.product_id = CAST(? AS UNIQUEIDENTIFIER)
                FROM dbo.acc_order_line ol
                INNER JOIN dbo.acc_order o ON o.id = ol.order_id
                WHERE ol.sku = ?
                  AND ol.product_id IS NULL
                  AND ISNULL(ol.quantity_ordered, 0) > 0
                  AND ISNULL(o.sales_channel, 'Amazon.com') != 'Non-Amazon'
                  AND o.amazon_order_id NOT LIKE 'S02-%'
                  {asin_filter_sql}
                """,
                params,
            )
            bound_lines = cur.rowcount

        # Step 2: Upsert purchase price
        cur.execute("""
            SELECT id, netto_price_pln FROM dbo.acc_purchase_price
            WHERE internal_sku = ? AND source = 'manual'
        """, [internal_sku])
        existing = cur.fetchone()
        old_manual_price = float(existing[1]) if existing and existing[1] else None

        now_str = date.today().isoformat()

        if existing:
            cur.execute("""
                UPDATE dbo.acc_purchase_price
                SET netto_price_pln = ?,
                    valid_from = ?,
                    updated_at = GETDATE()
                WHERE id = ?
            """, [netto_price_pln, now_str, existing[0]])
        else:
            cur.execute("""
                INSERT INTO dbo.acc_purchase_price
                    (internal_sku, netto_price_pln, valid_from, source, source_document, created_at, updated_at)
                VALUES (?, ?, ?, 'manual', 'data_quality_ui', GETDATE(), GETDATE())
            """, [internal_sku, netto_price_pln, now_str])

        # --- Controlling: log price change ---
        try:
            from app.services.controlling import log_price_change
            log_price_change(
                conn,
                internal_sku=internal_sku,
                old_price=old_manual_price,
                new_price=netto_price_pln,
                source="manual",
                source_document="data_quality_ui",
            )
        except Exception:
            pass  # controlling is non-blocking

        _apply_manual_price_to_internal_sku(
            cur,
            internal_sku=internal_sku,
            netto_price_pln=netto_price_pln,
        )
        conn.commit()
        _result_cache_invalidate()
        return {
            "sku": sku,
            "internal_sku": internal_sku,
            "netto_price_pln": netto_price_pln,
            "mapped_products": mapped_count,
            "bound_lines": bound_lines,
            "binding_source": binding_target["source"] if binding_target else None,
            "price_status": "updated" if existing else "created",
        }
    finally:
        conn.close()

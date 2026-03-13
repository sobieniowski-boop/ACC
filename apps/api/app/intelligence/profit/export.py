"""Profit Engine - XLSX export.

Extracted from the monolithic profit_engine.py (Sprint 3).
"""
from __future__ import annotations

from datetime import date
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.intelligence.profit.query import get_product_profit_table

def export_product_profit_xlsx(
    *,
    date_from: date,
    date_to: date,
    marketplace_id: str | None = None,
    brand: str | None = None,
    sku_search: str | None = None,
    fulfillment: str | None = None,
    profit_mode: str = "cm1",
    sort_by: str = "cm1_profit",
    sort_dir: str = "desc",
    group_by: str = "asin_marketplace",
    columns: list[str] | None = None,
) -> tuple[bytes, str]:
    """Generate XLSX file bytes for product profit table."""
    from io import BytesIO

    data = get_product_profit_table(
        date_from=date_from,
        date_to=date_to,
        marketplace_id=marketplace_id,
        brand=brand,
        sku_search=sku_search,
        fulfillment=fulfillment,
        profit_mode=profit_mode,
        sort_by=sort_by,
        sort_dir=sort_dir,
        group_by=group_by,
        page=1,
        page_size=100000,
    )

    default_columns = [
        "sku",
        "asin",
        "marketplace_code",
        "fulfillment_channel",
        "units",
        "revenue_pln",
        "cogs_per_unit",
        "fees_per_unit",
        "cm1_profit",
        "cm1_percent",
        "confidence_score",
        "loss_orders_pct",
    ]
    selected = columns or default_columns

    wb = Workbook()
    ws = wb.active
    ws.title = "ProductProfit"
    ws.append(selected)

    for item in data.get("items", []):
        ws.append([item.get(col) for col in selected])

    for idx, col in enumerate(selected, start=1):
        width = min(max(len(col) + 2, 12), 42)
        ws.column_dimensions[get_column_letter(idx)].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"product_profit_{group_by}_{date_from}_{date_to}.xlsx"
    return buf.getvalue(), filename



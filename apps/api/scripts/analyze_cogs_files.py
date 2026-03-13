"""Analyse structure of all XLSX files in 'cogs from sell' folder."""
import sys, os, glob
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl

FOLDER = r"C:\ACC\cogs from sell"

for fpath in sorted(glob.glob(os.path.join(FOLDER, "*.xlsx"))):
    fname = os.path.basename(fpath)
    print(f"\n{'=' * 70}")
    print(f"FILE: {fname}")
    print(f"{'=' * 70}")
    
    wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n  Sheet: '{sheet_name}'  (rows ~{ws.max_row}, cols ~{ws.max_column})")
        
        # Print first 8 rows to understand structure
        for i, row in enumerate(ws.iter_rows(max_row=8, values_only=False)):
            vals = []
            for cell in row:
                v = cell.value
                if v is not None:
                    vals.append(f"[{cell.column_letter}] {repr(v)[:60]}")
            if vals:
                print(f"    Row {i+1}: {' | '.join(vals)}")
    wb.close()

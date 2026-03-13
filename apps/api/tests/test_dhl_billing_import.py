from __future__ import annotations

from datetime import datetime
from unittest.mock import patch

from openpyxl import Workbook

from app.services.dhl_billing_import import (
    ImportedShipmentSeed,
    LocalPackageCandidate,
    _collect_package_candidates,
    _build_parcel_map_payloads,
    _collect_seed_links,
    _insert_parcel_map_payloads,
    _is_lock_timeout_error,
    _normalize_parcel_number,
    _resolve_seed_parcel_bases,
    _run_with_lock_retry,
    _should_skip_import_file,
    parse_dhl_invoice_xlsx,
    parse_dhl_jj_xlsx,
    parse_dhl_manifest_xlsx,
)
from app.services.bl_order_lookup import ResolvedBlOrder


def test_normalize_parcel_number_splits_suffix():
    parcel = _normalize_parcel_number("29907226125/ZW")
    assert parcel.raw == "29907226125/ZW"
    assert parcel.base == "29907226125"
    assert parcel.suffix == "ZW"


def test_parse_dhl_manifest_xlsx_reads_document_rows(tmp_path):
    path = tmp_path / "Raport faktur.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Lista faktur"
    ws.append(
        [
            "Klient",
            "Płatnik",
            "Typ dokumentu",
            "Numer dokumentu",
            "Data wystawienia",
            "Data wysyłki",
            "Termin płatności",
            "Kwota netto",
            "Kwota VAT",
            "Kwota brutto",
        ]
    )
    ws.append([2446860, 2446860, "Załącznik do dokumentu", 1106711106, "2026-02-15", "2026-02-16", "2026-03-17", 100.0, 23.0, 123.0])
    wb.save(path)

    rows = parse_dhl_manifest_xlsx(path)

    assert len(rows) == 1
    assert rows[0]["document_number"] == "1106711106"
    assert rows[0]["document_type"] == "Załącznik do dokumentu"
    assert rows[0]["gross_amount"] == 123.0


def test_parse_dhl_invoice_xlsx_normalizes_return_suffix(tmp_path):
    path = tmp_path / "DHL_Dokument nr 2100118278.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "2100118278"
    ws.append(
        [
            "Numer załącznika do dokumentu",
            "Data wystawienia",
            "Data sprzedaży",
            "Termin płatności",
            "Numer przesyłki",
            "Data doręczenia",
            "Ilość",
            "Kod produktu",
            "Opis",
            "Waga",
            "Rodzaj wagi",
            "Nadawca/Odbiorca",
            "Kto płaci?",
            "SAP zlec.",
            "MPK",
            "PKWiU",
            "Uwagi",
            "Razem netto",
            "Opłata podstawowa",
            "Rabat od opłaty podstawowej/zaokrąglenia",
            "Opłata sezonowa",
            "Opłata paliwowa i drogowa",
            "Przystąpienie do ubezp. DHL",
            "Pobranie",
            "Dopłata za wydruk etykiety",
            "Dopłata za wagę wolum. pow. 31,5 kg",
        ]
    )
    ws.append(
        [
            "2100118278",
            datetime(2026, 2, 12),
            datetime(2025, 12, 31),
            datetime(2026, 3, 14),
            "29907226125/ZW",
            datetime(2025, 12, 29),
            1,
            "EK.NL.0",
            "DHL Parcel Connect do 3 kg",
            0.05,
            "W",
            "NET FOX",
            "N",
            "2446861",
            "",
            "49.4",
            "",
            24.87,
            81.0,
            -60.86,
            0.8,
            3.93,
            0.0,
            0.0,
            0.0,
            0.0,
        ]
    )
    wb.save(path)

    rows = parse_dhl_invoice_xlsx(path)

    assert len(rows) == 1
    assert rows[0]["document_number"] == "2100118278"
    assert rows[0]["parcel_number"] == "29907226125/ZW"
    assert rows[0]["parcel_number_base"] == "29907226125"
    assert rows[0]["parcel_number_suffix"] == "ZW"
    assert rows[0]["sap_order"] == "2446861"
    assert rows[0]["net_amount"] == 24.87


def test_parse_dhl_jj_xlsx_reads_mapping(tmp_path):
    path = tmp_path / "EX_DHL_test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "dane"
    ws.append(
        [
            "NUMER_PRZESYLKI",
            "PACZKA_NUMER_DOWOLNY_2",
            "RODZAJ_PRZESYLKI",
            "NADANIE_DATACZAS_BORSOFT",
            "DORECZENIE_DATACZAS_BORSOFT",
            "OSTAT_ZDARZ_NAZWA",
            "OSTAT_ZDARZ_DATACZAS",
        ]
    )
    ws.append(
        [
            "30167116285",
            "JJD000030206530000003167275",
            "EK",
            datetime(2026, 2, 5, 8, 21),
            datetime(2026, 2, 11, 18, 58),
            "DRPDOR",
            datetime(2026, 2, 11, 18, 58),
        ]
    )
    wb.save(path)

    rows = parse_dhl_jj_xlsx(path)

    assert len(rows) == 1
    assert rows[0]["parcel_number_base"] == "30167116285"
    assert rows[0]["jjd_number"] == "JJD000030206530000003167275"
    assert rows[0]["last_event_code"] == "DRPDOR"


def test_collect_seed_links_prefers_jjd_match():
    seed = ImportedShipmentSeed(
        parcel_number="30167116285",
        parcel_number_base="30167116285",
        jjd_number="JJD000030206530000003167275",
        shipment_type="EK",
        ship_date=datetime(2026, 2, 5, 8, 21),
        delivery_date=datetime(2026, 2, 11, 18, 58),
        last_event_code="DRPDOR",
        last_event_at=datetime(2026, 2, 11, 18, 58),
        product_code="EK.NL.0",
        description="DHL Parcel Connect do 3 kg",
        issue_date=None,
        sales_date=None,
        total_net_amount=24.87,
        line_count=1,
    )
    candidate = LocalPackageCandidate(
        amazon_order_id="ORDER-1",
        acc_order_id="00000000-0000-0000-0000-000000000111",
        bl_order_id=123,
        package_order_id=123,
        courier_package_nr="JJD000030206530000003167275",
        courier_inner_number="30167116285",
    )

    links = _collect_seed_links(
        seed=seed,
        by_tracking={"JJD000030206530000003167275": [candidate]},
        by_inner={"30167116285": [candidate]},
        by_jjd_netfox={},
    )

    assert len(links) >= 1
    assert links[0]["amazon_order_id"] == "ORDER-1"
    assert any(link["link_method"] == "billing_jjd" for link in links)
    assert any(link["is_primary"] for link in links)


def test_collect_package_candidates_uses_resolved_bl_order_lookup():
    class FakeCursor:
        def execute(self, *_args, **_kwargs):
            return None

        def fetchall(self):
            return [
                (
                    316051289,
                    32174790,
                    "JJD000030214472000000232044",
                    "30252354464",
                    "dhl",
                    "de",
                )
            ]

    bucket: dict[str, list[LocalPackageCandidate]] = {}
    with patch(
        "app.services.dhl_billing_import.resolve_bl_orders_to_acc_orders",
        return_value={
            316051289: ResolvedBlOrder(
                bl_order_id=316051289,
                amazon_order_id="ORDER-1",
                acc_order_id="00000000-0000-0000-0000-000000000111",
            )
        },
    ):
        _collect_package_candidates(
            FakeCursor(),
            column_name="courier_inner_number",
            values=["30252354464"],
            result_bucket=bucket,
        )

    assert "30252354464" in bucket
    assert bucket["30252354464"][0].amazon_order_id == "ORDER-1"
    assert bucket["30252354464"][0].bl_order_id == 316051289


def test_is_lock_timeout_error_detects_sql_1222():
    assert _is_lock_timeout_error(Exception("(1222, b'Lock request time out period exceeded')")) is True
    assert _is_lock_timeout_error(Exception("other db error")) is False


def test_run_with_lock_retry_retries_and_succeeds():
    class FakeConn:
        def __init__(self):
            self.rollbacks = 0

        def rollback(self):
            self.rollbacks += 1

    calls = {"count": 0}

    def flaky():
        calls["count"] += 1
        if calls["count"] < 3:
            raise Exception("(1222, b'Lock request time out period exceeded')")
        return "ok"

    conn = FakeConn()
    with patch("app.services.dhl_billing_import.time.sleep") as mocked_sleep:
        result = _run_with_lock_retry(
            conn,
            flaky,
            op_name="jj_insert_chunk",
            source_file="test.xlsx",
        )

    assert result == "ok"
    assert calls["count"] == 3
    assert conn.rollbacks == 2
    assert mocked_sleep.call_count == 2


def test_should_skip_import_file_only_for_imported_same_snapshot():
    mtime = datetime(2026, 3, 6, 14, 16, 44)

    assert (
        _should_skip_import_file(
            (123, mtime, "imported"),
            file_size=123,
            file_mtime=mtime,
            force_reimport=False,
        )
        is True
    )
    assert (
        _should_skip_import_file(
            (123, mtime, "failed"),
            file_size=123,
            file_mtime=mtime,
            force_reimport=False,
        )
        is False
    )
    assert (
        _should_skip_import_file(
            (123, mtime, "imported"),
            file_size=123,
            file_mtime=mtime,
            force_reimport=True,
        )
        is False
    )


def test_resolve_seed_parcel_bases_uses_changed_scope_by_default():
    result = _resolve_seed_parcel_bases(
        None,
        changed_parcels={" 30167116285 ", "", "29907226125/ZW"},
        seed_all_existing=False,
    )

    assert result == {"30167116285", "29907226125/ZW"}


def test_resolve_seed_parcel_bases_can_load_all_existing():
    with patch(
        "app.services.dhl_billing_import._load_all_billing_parcel_bases",
        return_value={"A", "B"},
    ) as mocked:
        result = _resolve_seed_parcel_bases(
            object(),
            changed_parcels={"X"},
            seed_all_existing=True,
        )

    assert result == {"A", "B"}
    mocked.assert_called_once()


def test_build_and_insert_parcel_map_payloads_uses_all_rows():
    rows = [
        {
            "parcel_number": "30167116285",
            "parcel_number_base": "30167116285",
            "parcel_number_suffix": None,
            "jjd_number": "JJD0001",
            "shipment_type": "EK",
            "ship_date": datetime(2026, 2, 5, 8, 21),
            "delivery_date": datetime(2026, 2, 11, 18, 58),
            "last_event_code": "DRPDOR",
            "last_event_at": datetime(2026, 2, 11, 18, 58),
            "source_row_no": 2,
        },
        {
            "parcel_number": "30167116286",
            "parcel_number_base": "30167116286",
            "parcel_number_suffix": None,
            "jjd_number": "JJD0002",
            "shipment_type": "EK",
            "ship_date": datetime(2026, 2, 6, 8, 21),
            "delivery_date": datetime(2026, 2, 12, 18, 58),
            "last_event_code": "DRPDOR",
            "last_event_at": datetime(2026, 2, 12, 18, 58),
            "source_row_no": 3,
        },
    ]
    payloads = _build_parcel_map_payloads(source_file="source.xlsx", rows=rows)

    class FakeCursor:
        def __init__(self):
            self.payloads = None

        def executemany(self, sql, payloads):
            self.payloads = payloads

    cur = FakeCursor()
    count = _insert_parcel_map_payloads(cur, payloads)

    assert count == 2
    assert cur.payloads is not None
    assert len(cur.payloads) == 2
    assert cur.payloads[0][3] == "JJD0001"
